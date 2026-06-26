"""
SyndikPro — app.py v2
Flask backend — SaaS Platform
"""
from flask import Flask, request, jsonify, session, render_template
from flask_wtf.csrf import CSRFProtect
from flask_migrate import Migrate

from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from functools import wraps
from models import (AnnouncementDismissal, PlatformSettings,db, User, Residence, Apartment, Building, Payment, Complaint,
                    Expense, Worker, Resident, Notification,
                    GlobalNotification, SubscriptionPlan,
                    Assembly, Vote, VoteResponse, Document,
                    ReserveFund, FundTransaction, Announcement,
                    OnlinePaymentRequest, SupportTicket, NeighborPost, NeighborPostMessage,
                    PasswordResetToken, DailyFact)
import os, json
from PIL import Image
from io import BytesIO
import logging
from logging.handlers import RotatingFileHandler

# قراءة .env
def load_env():
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())
load_env()
from security import (rate_limit, sanitize_request_data,
    check_session_security, add_security_headers,
    secure_file_upload, log_suspicious, get_security_log)

app = Flask(__name__)
csrf = CSRFProtect(app)
migrate = Migrate(app, db)

app.config['WTF_CSRF_CHECK_DEFAULT'] = False

@app.after_request
def advanced_security_headers(response):
    # إخفاء معلومات السيرفر
    response.headers.pop('Server', None)
    response.headers.pop('X-Powered-By', None)
    # منع MIME sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # منع Clickjacking
    response.headers['X-Frame-Options'] = 'DENY'
    # XSS Protection
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # HSTS
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'
    # منع cache للبيانات الحساسة
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'
    # Referrer Policy
    response.headers['Referrer-Policy'] = 'no-referrer'
    # Permissions Policy
    response.headers['Permissions-Policy'] = 'geolocation=(), camera=(), microphone=(), payment=()'
    return response

@app.before_request
def block_suspicious_requests():
    import re
    ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr or '0.0.0.0').split(',')[0].strip()

    # منع User Agents الخطيرة
    ua = request.headers.get('User-Agent', '').lower()
    bad_agents = ['sqlmap', 'nikto', 'nmap', 'masscan', 'zgrab', 'gobuster', 'dirbuster', 'hydra', 'burpsuite', 'metasploit']
    for agent in bad_agents:
        if agent in ua:
            return jsonify({'error': 'forbidden'}), 403

    # منع path traversal
    path = request.path
    if '..' in path or '/etc/' in path or '/proc/' in path or '/root/' in path:
        return jsonify({'error': 'forbidden'}), 403

    # منع SQL Injection في كل الطلبات
    suspicious = ["'", '"', ';--', 'union select', 'drop table', 'insert into',
                  'delete from', 'update set', 'exec(', 'execute(', 'xp_cmd',
                  '<script', 'javascript:', 'vbscript:', 'onload=', 'onerror=']

    check_data = (request.path + '?' + request.query_string.decode()).lower()
    for s in suspicious:
        if s in check_data:
            return jsonify({'error': 'invalid request'}), 400

    # فحص JSON body إذا وجد
    if request.is_json:
        try:
            body = str(request.get_json()).lower()
            for s in ['<script', 'javascript:', 'union select', 'drop table']:
                if s in body:
                    return jsonify({'error': 'invalid request'}), 400
        except:
            pass


app.secret_key = os.environ.get('SECRET_KEY')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_COOKIE_NAME'] = 'syndikpro_session'
app.config['SESSION_COOKIE_SECURE'] = False  # PythonAnywhere HTTPS handled by proxy
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['SESSION_COOKIE_PATH'] = '/'

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
# Logging احترافي
if not app.debug:
    log_formatter = logging.Formatter('[%(asctime)s] %(levelname)s in %(module)s: %(message)s')
    log_file = os.path.join(BASE_DIR, 'logs', 'syndikpro.log')
    os.makedirs(os.path.join(BASE_DIR, 'logs'), exist_ok=True)
    file_handler = RotatingFileHandler(log_file, maxBytes=1024*1024, backupCount=5)
    file_handler.setFormatter(log_formatter)
    file_handler.setLevel(logging.WARNING)
    app.logger.addHandler(file_handler)
    app.logger.setLevel(logging.WARNING)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'databasesyndic.db') + '?timeout=30'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXT = {'png','jpg','jpeg','pdf'}

def allowed_file(f):
    if '.' not in f:
        return False
    ext = f.rsplit('.',1)[1].lower()
    if ext not in ALLOWED_EXT:
        return False
    # منع Double Extension مثل file.php.jpg
    if f.count('.') > 1:
        parts = f.lower().split('.')
        dangerous_ext = ['php','py','js','sh','exe','bat','cmd','asp','jsp']
        for p in parts[:-1]:
            if p in dangerous_ext:
                return False
    return True

def secure_save_file(file_obj, upload_folder):
    import time, re
    original = file_obj.filename
    ext = original.rsplit('.',1)[1].lower()
    # اسم عشوائي كامل - لا نحتفظ بالاسم الأصلي
    safe_name = f"{int(time.time())}_{secrets.token_hex(8)}.{ext}"
    # التحقق من Magic Bytes
    header = file_obj.read(8)
    file_obj.seek(0)
    magic = {
        'jpg': b'\xff\xd8\xff', 'jpeg': b'\xff\xd8\xff',
        'png': b'\x89PNG', 'pdf': b'%PDF'
    }
    if ext in magic and not header.startswith(magic[ext]):
        return None, 'محتوى الملف لا يطابق نوعه'
    # فحص الحجم (5MB max)
    file_obj.seek(0, 2)
    if file_obj.tell() > 5 * 1024 * 1024:
        return None, 'حجم الملف كبير (الحد 5MB)'
    file_obj.seek(0)
    import os
    path = os.path.join(upload_folder, safe_name)
    file_obj.save(path)
    return f'static/uploads/{safe_name}', 'ok'

db.init_app(app)

# ══════════════════════════════════════
#  SECURITY MIDDLEWARE
# ══════════════════════════════════════
from collections import defaultdict
import time as _time_module
import hashlib

_login_attempts = defaultdict(list)
_blocked_ips = {}

def get_client_ip():
    return request.environ.get('HTTP_X_FORWARDED_FOR',
           request.remote_addr or '0.0.0.0').split(',')[0].strip()

def is_ip_blocked(ip):
    if ip in _blocked_ips:
        if _time_module.time() < _blocked_ips[ip]:
            return True
        else:
            del _blocked_ips[ip]
    return False

def record_failed_attempt(ip, max_attempts=5, block_minutes=15):
    now = _time_module.time()
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < 300]
    _login_attempts[ip].append(now)
    if len(_login_attempts[ip]) >= max_attempts:
        _blocked_ips[ip] = now + (block_minutes * 60)
        return True
    return False

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), camera=(), microphone=()'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers.pop('Server', None)
    return response

@app.before_request
def check_session_timeout():
    # تجاهل المسارات العامة
    public = ['/api/login','/api/logout','/api/demo','/api/admin/login',
              '/api/syndic/register','/api/resident/login','/api/resident/logout',
              '/api/resident/register','/api/public/','/api/syndic/forgot',
              '/api/syndic/reset','/api/resident/forgot','/api/resident/reset',
              '/static/']
    path = request.path
    for p in public:
        if path.startswith(p):
            return None

    # فحص timeout السانديك/المدير
    if 'user_id' in session:
        last = session.get('last_activity')
        now = _time_module.time()
        if last and (now - last) > 30 * 60:
            session.clear()
            return jsonify({'error':'انتهت جلستك، يرجى تسجيل الدخول مجدداً','code':401}), 401
        session['last_activity'] = now

    # فحص timeout الساكن
    if 'resident_id' in session:
        last = session.get('resident_last_activity')
        now = _time_module.time()
        if last and (now - last) > 30 * 60:
            session.clear()
            return jsonify({'error':'انتهت جلستك، يرجى تسجيل الدخول مجدداً','code':401}), 401
        session['resident_last_activity'] = now

    return None

@app.before_request
def security_check():
    ip = get_client_ip()
    # فحص IP محظور
    if is_ip_blocked(ip):
        return jsonify({'error': 'تم حظر هذا العنوان مؤقتاً'}), 429
    # فحص محاولات SQL Injection في URL
    path = request.path.lower()
    args = request.query_string.decode().lower()
    dangerous = ["'", '"', ';--', 'union select', 'drop table', '<script', 'javascript:']
    for d in dangerous:
        if d in path or d in args:
            return jsonify({'error': 'طلب غير صالح'}), 400



@app.after_request
def security_headers(response):
    from security import add_security_headers
    return response

@app.before_request
def check_security():
    from security import check_session_security, log_suspicious
    import re
    # فحص SQL Injection في URL
    suspicious_patterns = ['union', 'select', 'drop', 'insert', '--', 'xp_', '<script']
    path = request.path.lower()
    for p in suspicious_patterns:
        if p in path:
            log_suspicious('SQL_INJECTION_ATTEMPT', f'Path: {request.path}')
            return jsonify({'error': 'طلب غير صالح'}), 400

# تسجيل Blueprints

with app.app_context():
    db.create_all()
    # Migration automatique
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)

    # residents
    res_cols = [c['name'] for c in inspector.get_columns('residents')]
    with db.engine.connect() as conn:
        for col, typ in [('country','VARCHAR(80)'),('floor','INTEGER'),('apt_number','VARCHAR(20)')]:
            if col not in res_cols:
                conn.execute(text(f"ALTER TABLE residents ADD COLUMN {col} {typ}"))
        conn.commit()

    # users — plan column
    user_cols = [c['name'] for c in inspector.get_columns('users')]
    with db.engine.connect() as conn:
        if 'plan' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN plan VARCHAR(20) DEFAULT 'free'"))
        if 'neighborhood' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN neighborhood VARCHAR(80)"))
        if 'country' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN country VARCHAR(80) DEFAULT 'المغرب'"))
        if 'status' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN status VARCHAR(20) DEFAULT 'active'"))
        if 'subscription_start' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN subscription_start DATETIME"))
        if 'subscription_end' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN subscription_end DATETIME"))
        if 'payment_proof' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN payment_proof VARCHAR(300)"))
        if 'payment_date' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN payment_date DATETIME"))
        if 'pay_method' not in user_cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN pay_method VARCHAR(30)"))
        conn.commit()

    # buildings table migration
    try:
        inspector.get_columns('buildings')
    except Exception:
        with db.engine.connect() as conn:
            conn.execute(text("""CREATE TABLE buildings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                residence_id INTEGER NOT NULL,
                name VARCHAR(50) NOT NULL,
                total_floors INTEGER DEFAULT 1,
                created_at DATETIME,
                FOREIGN KEY(residence_id) REFERENCES residences(id)
            )"""))
            conn.commit()

    # apartments — building_id column
    apt_cols = [c['name'] for c in inspector.get_columns('apartments')]
    with db.engine.connect() as conn:
        if 'building_id' not in apt_cols:
            conn.execute(text("ALTER TABLE apartments ADD COLUMN building_id INTEGER"))
            conn.commit()

    # residences — neighborhood and syndic_id columns
    res_cols2 = [c['name'] for c in inspector.get_columns('residences')]
    with db.engine.connect() as conn:
        if 'neighborhood' not in res_cols2:
            conn.execute(text("ALTER TABLE residences ADD COLUMN neighborhood VARCHAR(80)"))
        if 'country' not in res_cols2:
            conn.execute(text("ALTER TABLE residences ADD COLUMN country VARCHAR(80) DEFAULT 'المغرب'"))
        if 'syndic_id' not in res_cols2:
            conn.execute(text("ALTER TABLE residences ADD COLUMN syndic_id INTEGER"))
        if 'status' not in res_cols2:
            conn.execute(text("ALTER TABLE residences ADD COLUMN status VARCHAR(20) DEFAULT 'active'"))
        conn.commit()

    # payments — method column
    pay_cols = [c['name'] for c in inspector.get_columns('payments')]
    with db.engine.connect() as conn:
        if 'method' not in pay_cols:
            conn.execute(text("ALTER TABLE payments ADD COLUMN method VARCHAR(30) DEFAULT 'نقدي'"))
        conn.commit()

    # جداول جديدة
    for table in ['assemblies','votes','vote_responses','documents','reserve_fund','fund_transactions','announcements','support_tickets']:
        try:
            inspector.get_columns(table)
        except Exception:
            pass  # db.create_all() already handles creation

    # خطط الاشتراك الافتراضية
    if SubscriptionPlan.query.count() == 0:
        plans = [
        SubscriptionPlan(name='free',     label='مجانية',  price_monthly=0,   max_residences=1,   max_apartments=20,  features='["إقامة واحدة","6 عمارات","20 شقة كحد أقصى","الدعم الأساسي"]'),
        SubscriptionPlan(name='basic',    label='أساسية',  price_monthly=49,  max_residences=1,   max_apartments=100, features='["إقامة واحدة","6 عمارات","100 شقة","تقارير أساسية","الدعم عبر البريد"]'),
        SubscriptionPlan(name='pro',      label='احترافية',price_monthly=89,  max_residences=3,   max_apartments=500, features='["3 إقامات","6 عمارات لكل إقامة","500 شقة","تقارير متقدمة","أداء إلكتروني","الدعم المميز"]'),
        SubscriptionPlan(name='enterprise',label='غير محدود',price_monthly=119,max_residences=999,max_apartments=9999,features='["إقامات غير محدودة","6 عمارات لكل إقامة","دعم مخصص","تقارير كاملة","أولوية الدعم"]'),
        ]
        for p in plans: db.session.add(p)
        db.session.commit()

    # Admin افتراضي
    if not User.query.filter_by(role='admin').first():
        admin = User(
            username='admin', full_name='المدير العام',
            password_hash=generate_password_hash(os.environ.get('ADMIN_PASSWORD','SyndikPro2026@!')),
            role='admin', status='active', plan='enterprise',
            subscription_start=datetime.utcnow(),
            subscription_end=datetime(2099,12,31)
        )
        db.session.add(admin)
        db.session.commit()

# ═══════════════════════════════
#  DECORATORS
# ═══════════════════════════════
def login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if 'user_id' not in session:
            return jsonify({'error':'غير مصرح','code':401}),401
        return f(*a,**kw)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if 'user_id' not in session:
            return jsonify({'error':'غير مصرح','code':401}),401
        u = User.query.get(session['user_id'])
        if not u or u.role != 'admin':
            return jsonify({'error':'صلاحيات المدير مطلوبة','code':403}),403
        return f(*a,**kw)
    return decorated

def resident_required(f):
    @wraps(f)
    def decorated(*a, **kw):

        if 'resident_id' not in session:
            return jsonify({'error':'غير مصرح','code':401}),401
        return f(*a,**kw)
    return decorated

def get_current_user():
    return User.query.get(session['user_id'])

# ═══════════════════════════════
#  PAGE PRINCIPALE
# ═══════════════════════════════
@app.route('/')
def index():
    return render_template('index.html')

# ═══════════════════════════════
#  AUTH
# ═══════════════════════════════

@app.route('/api/admin/login', methods=['POST'])
@rate_limit(max_calls=3, period=600)
def admin_login():
    data = request.get_json() or {}
    u = User.query.filter_by(username=data.get('username','')).first()
    ip = get_client_ip()
    if not u or not check_password_hash(u.password_hash, data.get('password','')):
        blocked = record_failed_attempt(ip, max_attempts=3, block_minutes=30)
        return jsonify({'error':'بيانات غير صحيحة', 'blocked': blocked}),401
    if u.role != 'admin':
        record_failed_attempt(ip)
        return jsonify({'error':'غير مصرح'}),403
    session['user_id'] = u.id
    session['role'] = u.role
    session.permanent = True
    return jsonify({'ok':True,'role':'admin','name':u.full_name})

@app.route('/api/login', methods=['POST'])
@rate_limit(max_calls=5, period=300)
def login():
    data = request.get_json() or {}
    u = User.query.filter_by(username=data.get('username','')).first()
    if not u or not check_password_hash(u.password_hash, data.get('password','')):
        return jsonify({'error':'بيانات غير صحيحة'}),401
    ip = get_client_ip()
    if u.role == 'admin':
        record_failed_attempt(ip)
        return jsonify({'error':'غير مصرح'}),403

    session['user_id'] = u.id
    session['role'] = u.role
    session.permanent = True
    return jsonify({'ok':True,'role':u.role,'name':u.full_name or u.username})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    resp = jsonify({'ok':True})
    resp.delete_cookie('syndikpro_session')
    return resp

@app.route('/api/me')
@login_required
def me():
    u = get_current_user()
    return jsonify(u.to_dict())

@app.route('/api/syndic/register', methods=['POST'])
def syndic_register():
    full_name = request.form.get('full_name','').strip()
    username = request.form.get('username','').strip()
    email = request.form.get('email','').strip()
    phone = request.form.get('phone','').strip()
    city = request.form.get('city','').strip()
    neighborhood = request.form.get('neighborhood','').strip()
    country = request.form.get('country','المغرب').strip()
    password = request.form.get('password','')

    if not all([full_name, username, email, phone, password]):
        return jsonify({'error':'جميع الحقول مطلوبة'}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({'error':'البريد موجود'}), 400

    import random
    otp_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    otp_expires = datetime.utcnow() + timedelta(minutes=10)

    u = User(username=username, full_name=full_name, email=email, phone=phone,
             city=city, neighborhood=neighborhood, country=country,
             password_hash=generate_password_hash(password), role='syndic',
             status='pending_email_verification', otp_code=otp_code,
             otp_expires_at=otp_expires, email_verified=False)
    db.session.add(u)
    db.session.commit()

    try:
        send_email(email, 'رمز التحقق', f'الرمز: {otp_code}')
    except:
        pass

    return jsonify({'ok':True, 'user_id':u.id, 'message':'رمز أرسل'}), 200

@app.route('/api/syndic/verify-otp', methods=['POST'])
def verify_syndic_otp():
    user_id = request.form.get('user_id', type=int)
    otp_code = request.form.get('otp_code', '').strip()
    if not user_id or not otp_code:
        return jsonify({'error':'مطلوب'}), 400
    u = User.query.get(user_id)
    if not u or u.otp_code != otp_code:
        return jsonify({'error':'رمز خطأ'}), 400
    if datetime.utcnow() > u.otp_expires_at:
        return jsonify({'error':'انتهى الوقت'}), 400
    u.email_verified = True
    u.status = 'pending'
    u.otp_code = None
    db.session.commit()
    # إشعار الأدمين فقط — بدون تدخل في الحساب
    try:
        admins = User.query.filter_by(role='admin').all()
        for admin in admins:
            n = Notification(user_id=admin.id,
                             message=f'سنديك جديد مسجل: {u.full_name} ({u.email})',
                             type='info')
            db.session.add(n)
        db.session.commit()
    except:
        pass
    return jsonify({'ok':True, 'message':'تم التحقق، يمكنك الولوج الآن'}), 200

@app.route('/api/syndic/resend-otp', methods=['POST'])
def resend_syndic_otp():
    user_id = request.form.get('user_id', type=int)
    if not user_id:
        return jsonify({'error': 'مطلوب'}), 400
    u = User.query.get(user_id)
    if not u:
        return jsonify({'error': 'غير موجود'}), 404
    import random
    otp_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    u.otp_code = otp_code
    u.otp_expires_at = datetime.utcnow() + timedelta(minutes=10)
    db.session.commit()
    try:
        send_email(u.email, 'رمز جديد', f'الرمز: {otp_code}')
    except:
        pass
    return jsonify({'ok': True, 'message': 'أعيد'}), 200

@app.route('/api/check-syndic-exists')
def check_syndic_exists():
    email = request.args.get('email', '').strip()
    phone = request.args.get('phone', '').strip()
    username = request.args.get('username', '').strip()
    
    result = {
        'email_exists': False,
        'phone_exists': False,
        'username_exists': False
    }
    
    if email:
        result['email_exists'] = bool(User.query.filter_by(email=email).first())
    if phone:
        result['phone_exists'] = bool(User.query.filter_by(phone=phone).first())
    if username:
        result['username_exists'] = bool(User.query.filter_by(username=username).first())
    
    return jsonify(result), 200


@app.route('/api/subscription-plans')
def get_subscription_plans():
    try:
        plans_list = SubscriptionPlan.query.filter_by(is_active=True).all()
        result = []
        if plans_list:
            for p in plans_list:
                result.append({'name': p.name, 'label': p.label or p.name, 'price_monthly': p.price_monthly, 'desc': f'{p.max_residences} إقامة، {p.max_apartments} شقة'})
        else:
            result = [{'name': 'free', 'label': 'مجانية', 'price_monthly': 0, 'desc': '1 إقامة، 20 شقة'}, {'name': 'basic', 'label': 'أساسية', 'price_monthly': 299, 'desc': '1 إقامة، 100 شقة'}, {'name': 'pro', 'label': 'احترافية', 'price_monthly': 599, 'desc': '3 إقامات، 500 شقة'}, {'name': 'enterprise', 'label': 'غير محدود', 'price_monthly': 1299, 'desc': 'إقامات غير محدودة'}]
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/syndic/choose-plan', methods=['POST'])
def syndic_select_plan():
    if 'user_id' not in session:
        return jsonify({'error':'غير مصرح'}), 401
    u = User.query.get(session['user_id'])
    if not u or u.role != 'syndic':
        return jsonify({'error':'غير مصرح'}), 403
    plan = request.form.get('plan','starter')
    pay_method = request.form.get('pay_method','')
    proof_path = ''
    if 'payment_proof' in request.files:
        f = request.files['payment_proof']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            fn = secure_filename(f.filename)
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            proof_path = f'static/uploads/{fn}'
    u.plan = plan
    u.pay_method = pay_method
    u.payment_method = pay_method
    if proof_path:
        u.payment_proof = proof_path
    u.payment_date = datetime.utcnow()
    u.total_amount = float(request.form.get('total_amount', 0) or 0)
    u.duration_months = int(request.form.get('duration_months') or request.form.get('months') or 1)
    u.subscription_confirmed = False
    db.session.commit()
    # إشعار داخلي للأدمين
    try:
        gn = GlobalNotification(
            title=f'طلب اشتراك جديد — {u.full_name}',
            body=f'الخطة: {plan} | الأداء: {pay_method} | {u.phone or "—"} | {u.city or "—"}',
            type='payment',
            target='admin'
        )
        db.session.add(gn)
        db.session.commit()
    except Exception as e:
        print(f'notif error: {e}')
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        admins = User.query.filter_by(role='admin').all()
        for adm in admins:
            if not adm.email:
                continue
            msg = MIMEMultipart('alternative')
            msg['Subject'] = 'SyndikPro — طلب اشتراك جديد'
            msg['From'] = GMAIL_USER
            msg['To'] = adm.email
            body = (
                '<div dir="rtl" style="font-family:Arial,sans-serif;max-width:500px;margin:auto;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden">'
                '<div style="background:linear-gradient(135deg,#1e40af,#3b82f6);padding:20px;text-align:center">'
                '<h1 style="color:#fff;margin:0;font-size:17px">🏢 SyndikPro — طلب اشتراك جديد</h1></div>'
                '<div style="padding:20px">'
                f'<p><b>الاسم:</b> {u.full_name}</p>'
                f'<p><b>البريد:</b> {u.email}</p>'
                f'<p><b>الهاتف:</b> {u.phone or "—"}</p>'
                f'<p><b>الخطة:</b> {plan}</p>'
                f'<p><b>طريقة الأداء:</b> {pay_method}</p>'
                '<div style="text-align:center;margin-top:14px">'
                '<a href="https://hicham.pythonanywhere.com" style="background:#1e40af;color:#fff;padding:10px 24px;border-radius:8px;text-decoration:none;font-weight:700">لوحة التحكم ←</a>'
                '</div></div>'
                '<div style="background:#f1f5f9;padding:10px;text-align:center;font-size:11px;color:#94a3b8">SyndikPro</div></div>'
            )
            msg.attach(MIMEText(body, 'html', 'utf-8'))
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(GMAIL_USER, GMAIL_PASS)
                smtp.sendmail(GMAIL_USER, adm.email, msg.as_string())
    except Exception as e:
        print(f'admin email error: {e}')
    return jsonify({'ok':True,'message':'تم إرسال طلب الاشتراك، في انتظار تأكيد الأداء'})

# ═══════════════════════════════
#  ADMIN — لوحة المدير العام
# ═══════════════════════════════

# إحصائيات عامة
@app.route('/api/admin/stats')
@admin_required
def admin_stats():
    # إجمالي السانديك
    total_syndics   = User.query.filter_by(role='syndic').count()
    # النشطون: subscription_confirmed=True
    active_syndics  = User.query.filter_by(role='syndic', subscription_confirmed=True).count()
    # بانتظار القبول: طلب اشتراك pending
    from models import SubscriptionRequest
    pending_ids     = [r.user_id for r in SubscriptionRequest.query.filter_by(status='pending').all()]
    pending_syndics = len(set(pending_ids))
    # منتهي الاشتراك
    expired_syndics = User.query.filter_by(role='syndic', status='expired').count()
    
    total_res      = Residence.query.count()
    total_apts     = Apartment.query.count()
    total_residents= Resident.query.count()
    total_complaints= Complaint.query.count()
    open_complaints = Complaint.query.filter_by(status='open').count()
    total_payments  = Payment.query.filter_by(status='paid').count()
    revenue_total   = db.session.query(db.func.sum(Payment.amount)).filter_by(status='paid').scalar() or 0

    # إيرادات هذا الشهر
    now = datetime.utcnow()
    revenue_month = db.session.query(db.func.sum(Payment.amount)).filter(
        Payment.status=='paid', Payment.month==now.month, Payment.year==now.year
    ).scalar() or 0

    # خطط الاشتراك
    plan_stats = {}
    for plan in ['free','basic','pro','enterprise']:
        plan_stats[plan] = User.query.filter_by(role='syndic', status='active', plan=plan).count()

    return jsonify({
        'syndics': {'total':total_syndics,'active':active_syndics,'pending':pending_syndics,'expired':expired_syndics},
        'platform': {'residences':total_res,'apartments':total_apts,'residents':total_residents},
        'complaints': {'total':total_complaints,'open':open_complaints},
        'payments': {'total':total_payments,'revenue_total':revenue_total,'revenue_month':revenue_month},
        'plans': plan_stats,
    })


# إيرادات شهرية تفصيلية
@app.route('/api/admin/revenue/monthly')
@admin_required
def admin_revenue_monthly():
    year = request.args.get('year', datetime.utcnow().year, type=int)
    month = request.args.get('month', datetime.utcnow().month, type=int)
    payments = db.session.query(Payment, User, Residence).join(
        Apartment, Payment.apartment_id == Apartment.id
    ).join(
        Residence, Apartment.residence_id == Residence.id
    ).join(
        User, Residence.user_id == User.id
    ).filter(
        Payment.status == 'paid',
        Payment.month == month,
        Payment.year == year
    ).all()
    result = {}
    for pay, syndic, res in payments:
        key = syndic.id
        if key not in result:
            result[key] = {
                'syndic_id': syndic.id,
                'syndic_name': syndic.full_name or syndic.username,
                'syndic_phone': syndic.phone or '',
                'syndic_email': syndic.email or '',
                'residence': res.name,
                'city': syndic.city or '',
                'total': 0,
                'payments_count': 0
            }
        result[key]['total'] += pay.amount
        result[key]['payments_count'] += 1
    months_data = []
    for y in range(2024, datetime.utcnow().year + 1):
        for m in range(1, 13):
            total = db.session.query(db.func.sum(Payment.amount)).filter(
                Payment.status == 'paid', Payment.month == m, Payment.year == y
            ).scalar() or 0
            if total > 0:
                months_data.append({'year': y, 'month': m, 'total': float(total)})
    return jsonify({
        'details': list(result.values()),
        'month_total': sum(r['total'] for r in result.values()),
        'history': months_data
    })
# قائمة السانديك
@app.route('/api/admin/syndics')
@admin_required
def admin_syndics():
    status = request.args.get('status','')
    q = User.query.filter_by(role='syndic')
    if status: q = q.filter_by(status=status)
    syndics = q.order_by(User.created_at.desc()).all()
    return jsonify([s.to_dict() for s in syndics])

# قبول/رفض السانديك
@app.route('/api/admin/syndics/<int:sid>', methods=['GET'])
@admin_required
def admin_get_syndic(sid):
    u = User.query.get_or_404(sid)
    return jsonify(u.to_dict())

@app.route('/api/admin/syndics/<int:sid>/approve', methods=['POST'])
@admin_required
def admin_approve_syndic(sid):
    data = request.get_json() or {}
    u = User.query.get_or_404(sid)
    action = data.get('action','approve')
    plan   = data.get('plan','basic')
    months = int(data.get('months', 1))

    if action == 'approve':
        u.status = 'active'
        u.subscription_confirmed = True
        u.renewal_pending = False
        u.plan   = plan
        u.approved_by = session['user_id']
        u.approved_at = datetime.utcnow()
        u.subscription_start = datetime.utcnow()
        u.subscription_end   = datetime.utcnow() + timedelta(days=30*months)
    elif action == 'reject':
        u.status = 'rejected'
    elif action == 'expire':
        u.status = 'expired'
    elif action == 'suspend':
        u.status = 'suspended'

    db.session.commit()

    if action == 'approve' and u.email:
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            pl = {'free':'مجانية','basic':'أساسية','pro':'احترافية','enterprise':'غير محدود'}
            msg = MIMEMultipart('alternative')
            msg['Subject'] = "SyndikPro - تم قبول حسابك"
            msg['From'] = GMAIL_USER
            msg['To']   = u.email
            end_date = u.subscription_end.strftime('%d/%m/%Y')
            body = (
                '<div dir="rtl" style="font-family:Arial,sans-serif;max-width:560px;margin:auto;background:#f8fafc;padding:0;border-radius:16px;overflow:hidden;border:1px solid #e2e8f0">'

                '<div style="background:linear-gradient(135deg,#1e40af,#3b82f6);padding:32px 24px;text-align:center">'
                '<div style="font-size:48px;margin-bottom:8px">🏢</div>'
                '<h1 style="color:#fff;margin:0;font-size:22px;font-weight:900">SyndikPro</h1>'
                '<p style="color:rgba(255,255,255,0.8);margin:4px 0 0;font-size:13px">نظام تدبير الإقامات السكنية الذكي</p>'
                '</div>'

                '<div style="padding:28px 24px">'

                '<p style="font-size:15px;color:#0f172a;margin:0 0 6px">السلام عليكم ورحمة الله،</p>'
                '<p style="font-size:14px;color:#334155;margin:0 0 20px">الأستاذ/ة <b>' + u.full_name + '</b>،</p>'

                '<p style="font-size:14px;color:#334155;line-height:1.8;margin:0 0 20px">'
                'يسعدنا أن نُبلغكم بأن طلب انضمامكم إلى منصة <b>SyndikPro</b> قد تمت مراجعته والموافقة عليه رسمياً. '
                'نشكركم على ثقتكم بنا ونرحب بكم في عائلة SyndikPro.'
                '</p>'

                '<div style="background:#d1fae5;border:1px solid #6ee7b7;border-radius:12px;padding:16px;text-align:center;margin-bottom:24px">'
                '<div style="font-size:28px">✅</div>'
                '<b style="color:#065f46;font-size:16px">تم تفعيل حسابكم بنجاح!</b>'
                '</div>'

                '<div style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden;margin-bottom:24px">'
                '<div style="background:#1e40af;padding:10px 16px">'
                '<b style="color:#fff;font-size:13px">📋 تفاصيل اشتراككم</b>'
                '</div>'
                '<table style="width:100%;font-size:13px;border-collapse:collapse">'
                '<tr style="border-bottom:1px solid #f1f5f9"><td style="color:#64748b;padding:10px 16px;width:45%">اسم المستخدم</td><td style="padding:10px 16px;font-weight:700;color:#1e40af">' + u.username + '</td></tr>'
                '<tr style="border-bottom:1px solid #f1f5f9;background:#f8fafc"><td style="color:#64748b;padding:10px 16px">خطة الاشتراك</td><td style="padding:10px 16px;font-weight:700">' + pl.get(plan,plan) + '</td></tr>'
                '<tr style="border-bottom:1px solid #f1f5f9"><td style="color:#64748b;padding:10px 16px">مدة الاشتراك</td><td style="padding:10px 16px;font-weight:700">' + str(months) + ' شهر</td></tr>'
                '<tr style="background:#f8fafc"><td style="color:#64748b;padding:10px 16px">تاريخ الانتهاء</td><td style="padding:10px 16px;font-weight:700;color:#dc2626">' + end_date + '</td></tr>'
                '</table>'
                '</div>'

                '<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:12px;padding:18px;margin-bottom:24px">'
                '<p style="font-weight:900;color:#1e40af;margin:0 0 12px;font-size:14px">🚀 كيف تبدأ؟</p>'
                '<p style="margin:6px 0;font-size:13px;color:#334155">1️⃣ توجه إلى التطبيق وانقر على <b>دخول</b></p>'
                '<p style="margin:6px 0;font-size:13px;color:#334155">2️⃣ أدخل اسم المستخدم: <b style="color:#1e40af;background:#dbeafe;padding:2px 8px;border-radius:6px">' + u.username + '</b></p>'
                '<p style="margin:6px 0;font-size:13px;color:#334155">3️⃣ أدخل كلمة المرور التي اخترتها عند التسجيل</p>'
                '<p style="margin:6px 0;font-size:13px;color:#334155">4️⃣ أضف إقاماتك وعماراتك وشققك وابدأ الإدارة</p>'
                '</div>'

                '<div style="text-align:center;margin-bottom:24px">'
                '<a href="https://syndikpro.ma" style="display:inline-block;background:linear-gradient(135deg,#1e40af,#3b82f6);color:#fff;padding:14px 40px;border-radius:12px;text-decoration:none;font-weight:900;font-size:15px">الدخول إلى التطبيق ←</a>'
                '</div>'

                '<p style="font-size:13px;color:#64748b;line-height:1.8;margin:0 0 8px">'
                'إذا واجهتم أي صعوبة أو كان لديكم أي استفسار، فريق الدعم لدينا في خدمتكم على مدار الساعة.'
                '</p>'
                '<p style="font-size:13px;color:#334155;margin:0">مع تحيات فريق <b>SyndikPro</b> 🏢</p>'
                '</div>'

                '<div style="background:#f1f5f9;padding:16px 24px;text-align:center;border-top:1px solid #e2e8f0">'
                '<p style="color:#94a3b8;font-size:11px;margin:0">SyndikPro — نظام تدبير الإقامات السكنية الذكي</p>'
                '<p style="color:#94a3b8;font-size:11px;margin:4px 0 0">هذا البريد أُرسل تلقائياً، يرجى عدم الرد عليه مباشرة</p>'
                '</div>'
                '</div>'
            )
            msg.attach(MIMEText(body, 'html', 'utf-8'))
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(GMAIL_USER, GMAIL_PASS)
                smtp.sendmail(GMAIL_USER, u.email, msg.as_string())
            print("email sent to " + u.email)
        except Exception as e:
            print("email error: " + str(e))

    return jsonify({'ok':True,'status':u.status})

# تعديل خطة السانديك
@app.route('/api/admin/syndics/<int:sid>/plan', methods=['PUT'])
@admin_required
def admin_update_plan(sid):
    data = request.get_json() or {}
    u = User.query.get_or_404(sid)
    u.plan = data.get('plan', u.plan)
    months = int(data.get('months', 1))
    base = u.subscription_end if u.subscription_end and u.subscription_end > datetime.utcnow() else datetime.utcnow()
    u.subscription_end = base + timedelta(days=30*months)
    u.status = 'active'
    db.session.commit()
    return jsonify({'ok':True})

# حذف السانديك
@app.route('/api/admin/syndics/<int:sid>', methods=['DELETE'])
@admin_required
def admin_delete_syndic(sid):
    u = User.query.get_or_404(sid)
    db.session.delete(u)
    db.session.commit()
    return jsonify({'ok':True})

# خطط الاشتراك
@app.route('/api/admin/plans')
@admin_required
def admin_plans():
    plans = SubscriptionPlan.query.filter_by(is_active=True).order_by(SubscriptionPlan.id).all()
    return jsonify([p.to_dict() for p in plans])

@app.route('/api/admin/plans/<int:pid>', methods=['PUT'])
@admin_required
def admin_update_plan_def(pid):
    data = request.get_json() or {}
    p = SubscriptionPlan.query.get_or_404(pid)
    
    # تحديث الحقول مع معالجة القيم الفارغة
    if 'label' in data and data['label']:
        p.label = data['label']
    if 'price_monthly' in data:
        p.price_monthly = float(data['price_monthly']) if data['price_monthly'] else 0
    if 'max_residences' in data:
        p.max_residences = int(data['max_residences']) if data['max_residences'] else 1
    if 'max_buildings' in data:
        p.max_buildings = int(data['max_buildings']) if data['max_buildings'] else 6
    if 'max_apartments' in data:
        # القيمة الافتراضية 1 إذا كانت فارغة
        p.max_apartments = int(data['max_apartments']) if data['max_apartments'] else 1
    if 'features' in data:
        p.features = data['features'] if data['features'] else ''
    
    db.session.commit()
    return jsonify({'ok':True})

# إشعار جماعي
@app.route('/api/admin/notify', methods=['POST'])
@admin_required
def admin_notify():
    data = request.get_json() or {}
    n = GlobalNotification(
        title=data.get('title','إشعار'),
        body=data.get('body',''),
        type=data.get('type','info'),
        target=data.get('target','all'),
    )
    db.session.add(n)
    db.session.commit()
    if data.get('target') == 'residents':
        from models import Resident, Notification
        residents = Resident.query.filter_by(status='approved').all()
        for r in residents:
            notif = Notification(resident_id=r.id, title=data.get('title','إشعار'), body=data.get('body',''), type=data.get('type','info'))
            db.session.add(notif)
        db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/admin/global-notifications/<int:nid>', methods=['DELETE'])
@admin_required
def admin_delete_global_notif(nid):
    g = GlobalNotification.query.get(nid)
    if g:
        db.session.delete(g)
        db.session.commit()
        return jsonify({'ok':True})
    return jsonify({'error':'غير موجود'}), 404

# admin_notifications old removed

# إحصائيات النمو الشهري
@app.route('/api/admin/growth')
@admin_required
def admin_growth():
    result = []
    now = datetime.utcnow()
    for i in range(6,0,-1):
        d = now - timedelta(days=30*i)
        count = User.query.filter(
            User.role=='syndic',
            User.created_at >= d,
            User.created_at < d + timedelta(days=30)
        ).count()
        result.append({'month': d.strftime('%m/%Y'), 'count': count})
    return jsonify(result)

# ═══════════════════════════════
#  SYNDIC — إقامات
# ═══════════════════════════════
@app.route('/api/residences', methods=['GET'])
@login_required
def get_residences():
    u = get_current_user()
    if u.role == 'admin':
        res = Residence.query.all()
    else:
        res = Residence.query.filter_by(user_id=u.id).all()
    result = []
    for r in res:
        total = len(r.apartments)
        paid  = sum(1 for a in r.apartments if Payment.query.filter_by(apartment_id=a.id,month=datetime.utcnow().month,year=datetime.utcnow().year,status='paid').first())
        complaints_open = sum(Complaint.query.filter_by(apartment_id=a.id,status='open').count() for a in r.apartments)
        result.append({
            'id':r.id,'name':r.name,'address':r.address or '','city':r.city or '',
            'neighborhood':r.neighborhood or '','total_floors':r.total_floors,
            'total_apartments':total,'paid_this_month':paid,
            'unpaid_this_month':total-paid,'complaints_open':complaints_open,
        })
    return jsonify(result)

@app.route('/api/residences', methods=['POST'])
@login_required
def add_residence():
    u = get_current_user()
    data = request.get_json() or {}
    plan_obj = SubscriptionPlan.query.filter_by(name=u.plan).first()
    current_count = Residence.query.filter_by(user_id=u.id).count()
    if plan_obj and current_count >= plan_obj.max_residences:
        return jsonify({'error': 'وصلت للحد الاقصى (' + str(plan_obj.max_residences) + ' اقامة) في خطتك. يرجى الترقية.'}), 403
    r = Residence(
        user_id=u.id, name=data['name'],
        address=data.get('address',''), city=data.get('city',''),
        neighborhood=data.get('neighborhood',''),
        total_floors=int(data.get('total_floors',0))
    )
    db.session.add(r)
    db.session.commit()
    return jsonify({'ok':True,'id':r.id})

@app.route('/api/residences/<int:rid>', methods=['PUT'])
@login_required
def update_residence(rid):
    u = get_current_user()
    r = Residence.query.get_or_404(rid)
    if u.role != 'admin' and r.user_id != u.id:
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json() or {}
    for k in ['name','address','city','neighborhood','total_floors']:
        if k in data: setattr(r, k, data[k])
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/residences/<int:rid>', methods=['DELETE'])
@login_required
def delete_residence(rid):
    r = Residence.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'ok':True})


# ═══════════════════════════════
#  BUILDINGS — عمارات
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/buildings')
@login_required
def get_buildings(rid):
    buildings = Building.query.filter_by(residence_id=rid).all()
    return jsonify([b.to_dict() for b in buildings])

@app.route('/api/buildings', methods=['POST'])
@login_required
def add_building():
    data = request.get_json() or {}
    rid = data.get('residence_id')
    res = Residence.query.get_or_404(rid)
    user = res.syndic
    plan = SubscriptionPlan.query.filter_by(name=user.plan).first()
    if not plan:
        return jsonify({'error': 'خطة غير موجودة'}), 400
    current = Building.query.filter_by(residence_id=rid).count()
    if current >= plan.max_buildings:
        return jsonify({'error': f'الحد الأقصى: {plan.max_buildings} عمارات'}), 403
    b = Building(residence_id=rid, name=data.get('name'), total_floors=int(data.get('total_floors', 1)))
    db.session.add(b)
    db.session.commit()
    return jsonify({'ok':True, 'id':b.id})

@app.route('/api/buildings/<int:bid>', methods=['PUT'])
@login_required
def update_building(bid):
    b = Building.query.get_or_404(bid)
    data = request.get_json() or {}
    for k in ['name','total_floors']:
        if k in data: setattr(b, k, data[k])
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/buildings/<int:bid>', methods=['DELETE'])
@login_required
def delete_building(bid):
    b = Building.query.get_or_404(bid)
    db.session.delete(b)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/residences/<int:rid>/stats')
@login_required
def residence_stats(rid):
    buildings = Building.query.filter_by(residence_id=rid).all()
    total_apts = sum(len(b.apartments) for b in buildings)
    return jsonify({
        'buildings_count': len(buildings),
        'apartments_count': total_apts,
        'buildings': [b.to_dict() for b in buildings]
    })

# ═══════════════════════════════
#  APARTMENTS
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/apartments')
@login_required
def get_apartments(rid):
    # جلب الشقق من خلال العمارات فقط
    buildings = Building.query.filter_by(residence_id=rid).all()
    apts = []
    for b in buildings:
        apts.extend(b.apartments)
    result = []
    for a in apts:
        d = a.to_dict()
        resident = Resident.query.filter_by(apartment_id=a.id, status='approved').first()
        d['has_account'] = resident is not None
        d['resident_phone'] = resident.phone if resident else (a.owner_phone or '')
        cur_year = datetime.utcnow().year
        d['paid_months'] = Payment.query.filter_by(apartment_id=a.id, year=cur_year, status='paid').count()
        result.append(d)
    return jsonify(result)

@app.route('/api/apartments/<int:aid>', methods=['GET'])
@login_required
def get_apartment(aid):
    u = get_current_user()
    a = Apartment.query.get_or_404(aid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=a.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    return jsonify(a.to_dict())


@app.route('/api/apartments/<int:aid>/yearly-payments')
@login_required
def get_yearly_payments(aid):
    year = int(request.args.get('year', datetime.utcnow().year))
    pays = Payment.query.filter_by(apartment_id=aid, year=year).all()
    return jsonify([{'month':p.month,'status':p.status,'amount':p.amount} for p in pays])

@app.route('/api/apartments', methods=['POST'])
@login_required
def add_apartment():
    data = request.get_json() or {}
    if not data.get('building_id'):
        return jsonify({'error': 'building_id مطلوب'}), 400
    data = request.get_json() or {}
    res = Residence.query.get(data.get('residence_id'))
    if res:
        u = User.query.get(res.user_id)
        plan_obj = SubscriptionPlan.query.filter_by(name=u.plan).first() if u else None
        if plan_obj:
            all_res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
            total_apts = Apartment.query.filter(Apartment.residence_id.in_(all_res_ids)).count()
            if total_apts >= plan_obj.max_apartments:
                return jsonify({'error': 'وصلت للحد الاقصى (' + str(plan_obj.max_apartments) + ' شقة) في خطتك. يرجى الترقية.'}), 403
    a = Apartment(
        residence_id=data['residence_id'], number=data['number'],
        floor=int(data.get('floor',0)), owner_name=data.get('owner_name',''),
        owner_phone=data.get('owner_phone',''), tenant_name=data.get('tenant_name',''),
        tenant_phone=data.get('tenant_phone',''),
        monthly_fee=float(data.get('monthly_fee',250)),
        apt_type=data.get('apt_type','سكني'), notes=data.get('notes',''),
        building_id=int(data['building_id'])
    )
    db.session.add(a)
    db.session.commit()
    return jsonify({'ok':True,'id':a.id})

@app.route('/api/apartments/<int:aid>', methods=['PUT'])
@login_required
def update_apartment(aid):
    a = Apartment.query.get_or_404(aid)
    data = request.get_json() or {}
    for k in ['number','floor','owner_name','owner_phone','tenant_name','tenant_phone','monthly_fee','apt_type','notes']:
        if k in data: setattr(a, k, data[k])
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/apartments/<int:aid>', methods=['DELETE'])
@login_required
def delete_apartment(aid):
    a = Apartment.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  PAYMENTS
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/payments')
@login_required
def get_payments(rid):
    month = request.args.get('month')
    year  = int(request.args.get('year', datetime.utcnow().year))
    apts  = Apartment.query.filter_by(residence_id=rid).all()
    result = []
    for a in apts:
        resident = Resident.query.filter_by(apartment_id=a.id, status='approved').first()
        if month:
            pays_list = Payment.query.filter_by(apartment_id=a.id, month=int(month), year=year).all()
        else:
            pays_list = Payment.query.filter_by(apartment_id=a.id, year=year).all()
        if not pays_list:
            pays_list = [None]
        # نجلب طلبات الأداء المعلقة
        pending_reqs = OnlinePaymentRequest.query.filter_by(apartment_id=a.id, year=year).all()
        pending_months = {r.month: r for r in pending_reqs}
        paid_months = {pay.month for pay in pays_list if pay}
        # نضيف الأشهر pending اللي ما عندهاش Payment
        for mn, req in pending_months.items():
            if mn not in paid_months:
                result.append({
                    'apartment_id': a.id,
                    'number': a.number,
                    'owner_name': a.owner_name or '',
                    'month': mn,
                    'monthly_fee': a.monthly_fee,
                    'building_id': a.building_id,
                    'status': 'pending',
                    'amount': req.amount if req else 0,
                    'date_paid': None,
                    'method': '',
                    'payment_id': None,
                    'has_account': resident is not None,
            'resident_id': resident.id if resident else None,
            'owner_phone': a.owner_phone or '',
                })
        for pay in pays_list:
            if pay is None and not month:
                continue
            mn = pay.month if pay else (int(month) if month else 0)
            status = pay.status if pay else 'unpaid'
            result.append({
                'apartment_id': a.id,
                'number': a.number,
                'owner_name': a.owner_name or '',
                'month': mn,
                'monthly_fee': a.monthly_fee,
                'building_id': a.building_id,
                'status': status,
                'amount': pay.amount if pay else 0,
                'date_paid': pay.date_paid.isoformat() if pay and pay.date_paid else None,
                'method': pay.method if pay else '',
                'payment_id': pay.id if pay else None,
                'has_account': resident is not None,
            'resident_id': resident.id if resident else None,
            'owner_phone': a.owner_phone or '',
            })
    return jsonify(result)
@app.route('/api/payments', methods=['POST'])
@login_required
def add_payment():
    data = request.get_json() or {}
    pay = Payment.query.filter_by(
        apartment_id=data['apartment_id'],
        month=data['month'], year=data['year']
    ).first()
    if pay:
        pay.amount   = float(data.get('amount', pay.amount))
        pay.status   = data.get('status','paid')
        pay.method   = data.get('method','نقدي')
        pay.note     = data.get('note','')
        pay.date_paid= datetime.strptime(data['date_paid'],'%Y-%m-%d').date() if data.get('date_paid') else date.today()
    else:
        pay = Payment(
            apartment_id=int(data['apartment_id']),
            month=int(data['month']), year=int(data['year']),
            amount=float(data.get('amount',0)),
            status=data.get('status','paid'),
            method=data.get('method','نقدي'),
            note=data.get('note',''),
            date_paid=datetime.strptime(data['date_paid'],'%Y-%m-%d').date() if data.get('date_paid') else date.today()
        )
        db.session.add(pay)
    db.session.commit()

    # إشعار للقاطن عند تسجيل الدفعة
    if data.get('status', 'paid') == 'paid':
        try:
            from models import Resident, Notification
            res = Resident.query.filter_by(apartment_id=int(data['apartment_id']), status='approved').first()
            if res:
                months_ar = ['يناير','فبراير','مارس','أبريل','ماي','يونيو','يوليوز','غشت','شتنبر','أكتوبر','نونبر','دجنبر']
                mn = int(data['month'])
                notif = Notification(
                    resident_id=res.id,
                    title='✅ تم تسجيل دفعتك',
                    body=f'تمت المصادقة على دفعة شهر {months_ar[mn-1]} بمبلغ {data.get("amount", 0)} درهم',
                    type='payment'
                )
                db.session.add(notif)
                db.session.commit()
        except:
            pass

    return jsonify({'ok':True,'id':pay.id})

@app.route('/api/payments/<int:pid>', methods=['DELETE'])
@login_required
def delete_payment(pid):
    u = get_current_user()
    pay = Payment.query.get_or_404(pid)
    res = Residence.query.filter_by(id=pay.residence_id, user_id=u.id).first()
    if u.role != 'admin' and not res:
        return jsonify({'error': 'غير مصرح'}), 403
    db.session.delete(pay)
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  COMPLAINTS
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/complaints')
@login_required
def get_complaints(rid):
    apts = [a.id for a in Apartment.query.filter_by(residence_id=rid).all()]
    status = request.args.get('status','')
    q = Complaint.query.filter(Complaint.apartment_id.in_(apts))
    if status: q = q.filter_by(status=status)
    return jsonify([c.to_dict() for c in q.order_by(Complaint.date_created.desc()).all()])

@app.route('/api/complaints', methods=['POST'])
@login_required
def add_complaint():
    import time as _t
    photo_path = ''
    if request.content_type and 'multipart' in request.content_type:
        data = request.form.to_dict()
        if 'photo' in request.files:
            f = request.files['photo']
            if f and f.filename and allowed_file(f.filename):
                from werkzeug.utils import secure_filename
                fn = f"{int(_t.time())}_{secure_filename(f.filename)}"
                fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
                f.save(fp)
                photo_path = f'static/uploads/{fn}'
    else:
        data = request.get_json() or {}
    c = Complaint(
        apartment_id=data['apartment_id'], title=data['title'],
        description=data.get('description',''),
        priority=data.get('priority','medium'), status='open',
        photo_path=photo_path
    )
    db.session.add(c)
    db.session.commit()
    return jsonify({'ok':True,'id':c.id})

@app.route('/api/complaints/<int:cid>', methods=['PUT'])
@login_required
def update_complaint(cid):
    u = get_current_user()
    c = Complaint.query.get_or_404(cid)
    if u.role not in ('admin', 'syndic'):
        return jsonify({'error': 'غير مصرح'}), 403
    if u.role == 'syndic':
        from models import Apartment
        apt = Apartment.query.get(c.apartment_id)
        res = Residence.query.filter_by(id=apt.residence_id, user_id=u.id).first() if apt else None
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json() or {}
    if 'status' in data:
        old_status = c.status
        c.status = data['status']
        if data['status'] == 'closed': c.date_closed = datetime.utcnow()
    if 'priority' in data: c.priority = data['priority']
    if 'admin_note' in data: c.admin_note = data['admin_note']
    db.session.commit()
    # إشعار القاطن عند تغيير الحالة
    try:
        if 'status' in data and old_status != data['status']:
            from models import Resident, Notification
            st_ar = {'open':'مفتوحة','progress':'قيد المعالجة','closed':'تمت المعالجة'}
            res = Resident.query.filter_by(apartment_id=c.apartment_id, status='approved').first()
            if res:
                notif = Notification(
                    user_id=res.id, user_type='resident',
                    title=f'تحديث شكايتك: {c.title}',
                    body=f'الحالة الجديدة: {st_ar.get(data["status"],data["status"])}' + (f' — {data["admin_note"]}' if data.get('admin_note') else ''),
                    type='complaint', read=False
                )
                db.session.add(notif)
                db.session.commit()
    except Exception as e:
        print('notif error:', e)
    return jsonify({'ok':True})

@app.route('/api/complaints/<int:cid>', methods=['DELETE'])
@login_required
def delete_complaint(cid):
    u = get_current_user()
    c = Complaint.query.get_or_404(cid)
    if u.role not in ('admin', 'syndic'):
        return jsonify({'error': 'غير مصرح'}), 403
    if u.role == 'syndic':
        from models import Apartment
        apt = Apartment.query.get(c.apartment_id)
        res = Residence.query.filter_by(id=apt.residence_id, user_id=u.id).first() if apt else None
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  EXPENSES
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/expenses')
@login_required
def get_expenses(rid):
    from flask import request as freq
    month = freq.args.get('month', '')
    year = freq.args.get('year', '')
    exps = Expense.query.filter_by(residence_id=rid).order_by(Expense.date.desc()).all()
    if year:
        exps = [e for e in exps if str(e.date).startswith(year)]
    if month:
        exps = [e for e in exps if str(e.date).startswith(year+'-'+month.zfill(2))]
    return jsonify([e.to_dict() for e in exps])

@app.route('/api/expenses', methods=['POST'])
@login_required
def add_expense():
    u = get_current_user()
    data = request.get_json() or {}
    if u.role != 'admin':
        res = Residence.query.filter_by(id=data.get('residence_id'), user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    e = Expense(
        residence_id=data['residence_id'], title=data['title'],
        amount=float(data['amount']),
        category=data.get('category','أخرى'),
        date=datetime.strptime(data['date'],'%Y-%m-%d').date(),
        note=data.get('note','')
    )
    db.session.add(e)
    db.session.commit()
    return jsonify({'ok':True,'id':e.id})

@app.route('/api/expenses/<int:eid>', methods=['DELETE'])
@login_required
def delete_expense(eid):
    u = get_current_user()
    e = Expense.query.get_or_404(eid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=e.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    db.session.delete(e)
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  WORKERS
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/workers')
@login_required
def get_workers(rid):
    u = get_current_user()
    if u.role != 'admin':
        res = Residence.query.filter_by(id=rid, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    workers = Worker.query.filter_by(residence_id=rid).all()
    return jsonify([w.to_dict() for w in workers])

@app.route('/api/workers', methods=['POST'])
@login_required
def add_worker():
    u = get_current_user()
    data = request.get_json() or {}
    if u.role != 'admin':
        res = Residence.query.filter_by(id=data.get('residence_id'), user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    w = Worker(
        residence_id=data['residence_id'], full_name=data['full_name'],
        role=data.get('role','حارس'), phone=data.get('phone',''),
        cin=data.get('cin',''), cnss_number=data.get('cnss_number',''),
        cnss_status=data.get('cnss_status','غير مسجل'),
        salary=float(data.get('salary',0)),
        start_date=datetime.strptime(data['start_date'],'%Y-%m-%d').date() if data.get('start_date') else None,
        status=data.get('status','active'), notes=data.get('notes','')
    )
    db.session.add(w)
    db.session.commit()
    return jsonify({'ok':True,'id':w.id})


@app.route('/api/workers/<int:wid>/photo', methods=['POST'])
@login_required
def upload_worker_photo(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        chk = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not chk:
            return jsonify({'error': 'غير مصرح'}), 403
    if 'photo' not in request.files:
        return jsonify({'error': 'لا توجد صورة'}), 400
    f = request.files['photo']
    if not f or f.filename == '':
        return jsonify({'error': 'لم يتم اختيار ملف'}), 400
    from werkzeug.utils import secure_filename
    import time as _t
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'jpg'
    if ext not in ('png', 'jpg', 'jpeg'):
        return jsonify({'error': 'صيغة غير مدعومة'}), 400
    fn = secure_filename(f"worker_{wid}_{int(_t.time())}.{ext}")
    fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
    f.save(fp)
    w.photo = f'static/uploads/{fn}'
    db.session.commit()
    return jsonify({'ok': True, 'photo': w.photo})


def _ar_text(text):
    import arabic_reshaper
    from bidi.algorithm import get_display
    return get_display(arabic_reshaper.reshape(str(text)))


def _register_arabic_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    try:
        pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('Arabic-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    except Exception:
        pass


@app.route('/api/workers/<int:wid>/certificate/work', methods=['GET'])
@login_required
def worker_certificate_work(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        chk = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not chk:
            return jsonify({'error': 'غير مصرح'}), 403
    res = Residence.query.get(w.residence_id)

    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

    _register_arabic_fonts()
    ar = _ar_text

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2.5*cm, bottomMargin=2.5*cm, leftMargin=2.5*cm, rightMargin=2.5*cm)

    title_style = ParagraphStyle('title_ar_w', fontName='Arabic-Bold', fontSize=16, alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor('#1e40af'))
    title_en_style = ParagraphStyle('title_en_w', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, spaceAfter=22, textColor=colors.HexColor('#64748b'))
    header_style = ParagraphStyle('header_w', fontName='Arabic-Bold', fontSize=12, alignment=TA_CENTER, spaceAfter=2)
    sub_style = ParagraphStyle('sub_w', fontName='Arabic', fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#64748b'), spaceAfter=18)
    body_style = ParagraphStyle('body_w', fontName='Arabic', fontSize=11, alignment=TA_JUSTIFY, leading=20, spaceAfter=12)

    elements = []
    elements.append(Paragraph(ar(res.name if res else 'الإقامة'), header_style))
    elements.append(Paragraph(ar(f"{res.address or ''} - {res.city or ''}") if res else '', sub_style))

    elements.append(Paragraph(ar('شهادة عمل'), title_style))
    elements.append(Paragraph('ATTESTATION DE TRAVAIL', title_en_style))

    start_date_str = w.start_date.strftime('%Y/%m/%d') if w.start_date else '—'
    today_str = datetime.now().strftime('%Y/%m/%d')
    cin_part = f"، الحامل لبطاقة التعريف الوطنية رقم {w.cin}" if w.cin else ''

    body_text = (
        f"نشهد نحن الموقعين أسفله، متصرفي عقار {res.name if res else ''}، "
        f"بأن السيد(ة) {w.full_name}{cin_part}، "
        f"يعمل لدينا بصفة {w.role} منذ تاريخ {start_date_str}، "
        f"ولا يزال يزاول مهامه بكل جدية وإخلاص إلى حدود تاريخ إصدار هذه الشهادة."
    )
    elements.append(Paragraph(ar(body_text), body_style))
    elements.append(Paragraph(ar('سُلمت هذه الشهادة للمعني(ة) بالأمر للإدلاء بها عند الطلب.'), body_style))
    elements.append(Spacer(1, 30))

    info_data = [
        [ar('الاسم الكامل'), w.full_name, ar('الوظيفة'), ar(w.role)],
        [ar('رقم البطاقة الوطنية'), w.cin or '—', ar('رقم CNSS'), w.cnss_number or '—'],
        [ar('تاريخ الالتحاق'), start_date_str, ar('الحالة'), ar('نشط' if w.status == 'active' else w.status)],
    ]
    t = Table(info_data, colWidths=[3.6*cm, 4.4*cm, 3.0*cm, 4.0*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#eff6ff')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#eff6ff')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 50))

    footer_data = [[ar(f'حُررت بـ {res.city if res else ""}، بتاريخ {today_str}'), ar('إمضاء وختم المتصرف')]]
    ft = Table(footer_data, colWidths=[7.5*cm, 7.5*cm])
    ft.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    elements.append(ft)

    doc.build(elements)
    buf.seek(0)
    safe_name = w.full_name.replace(' ', '_')
    return send_file(buf, mimetype='application/pdf', as_attachment=False,
                      download_name=f'attestation_travail_{safe_name}.pdf')


@app.route('/api/workers/<int:wid>/certificate/salary', methods=['GET'])
@login_required
def worker_certificate_salary(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        chk = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not chk:
            return jsonify({'error': 'غير مصرح'}), 403
    res = Residence.query.get(w.residence_id)

    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY

    _register_arabic_fonts()
    ar = _ar_text

    salary = w.salary or 0
    cnss_employee = round(salary * 0.0464, 2)
    net_salary = round(salary - cnss_employee, 2)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2.5*cm, bottomMargin=2.5*cm, leftMargin=2.5*cm, rightMargin=2.5*cm)

    title_style = ParagraphStyle('title_ar_s', fontName='Arabic-Bold', fontSize=16, alignment=TA_CENTER, spaceAfter=4, textColor=colors.HexColor('#059669'))
    title_en_style = ParagraphStyle('title_en_s', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, spaceAfter=22, textColor=colors.HexColor('#64748b'))
    header_style = ParagraphStyle('header_s', fontName='Arabic-Bold', fontSize=12, alignment=TA_CENTER, spaceAfter=2)
    sub_style = ParagraphStyle('sub_s', fontName='Arabic', fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#64748b'), spaceAfter=18)
    body_style = ParagraphStyle('body_s', fontName='Arabic', fontSize=11, alignment=TA_JUSTIFY, leading=20, spaceAfter=12)

    elements = []
    elements.append(Paragraph(ar(res.name if res else 'الإقامة'), header_style))
    elements.append(Paragraph(ar(f"{res.address or ''} - {res.city or ''}") if res else '', sub_style))

    elements.append(Paragraph(ar('شهادة الأجرة'), title_style))
    elements.append(Paragraph('ATTESTATION DE SALAIRE', title_en_style))

    start_date_str = w.start_date.strftime('%Y/%m/%d') if w.start_date else '—'
    today_str = datetime.now().strftime('%Y/%m/%d')
    cin_part = f"، الحامل لبطاقة التعريف الوطنية رقم {w.cin}" if w.cin else ''

    body_text = (
        f"نشهد نحن الموقعين أسفله، متصرفي عقار {res.name if res else ''}، "
        f"بأن السيد(ة) {w.full_name}{cin_part}، "
        f"يعمل لدينا بصفة {w.role} منذ تاريخ {start_date_str}، "
        f"ويتقاضى أجرة شهرية إجمالية قدرها {salary:,.2f} درهم، "
        f"يُخصم منها اقتطاع الصندوق الوطني للضمان الاجتماعي (CNSS) بقيمة {cnss_employee:,.2f} درهم، "
        f"لتبلغ بذلك أجرته الصافية الشهرية {net_salary:,.2f} درهم."
    )
    elements.append(Paragraph(ar(body_text), body_style))
    elements.append(Paragraph(ar('سُلمت هذه الشهادة بطلب من المعني(ة) بالأمر للإدلاء بها لدى الجهات المختصة.'), body_style))
    elements.append(Spacer(1, 20))

    salary_data = [
        [ar('البيان'), ar('المبلغ (درهم)')],
        [ar('الأجرة الشهرية الإجمالية'), f"{salary:,.2f}"],
        [ar('اقتطاع CNSS'), f"-{cnss_employee:,.2f}"],
        [ar('الأجرة الصافية الشهرية'), f"{net_salary:,.2f}"],
    ]
    st = Table(salary_data, colWidths=[8*cm, 6*cm])
    st.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Arabic-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d1fae5')),
        ('FONTNAME', (0, -1), (-1, -1), 'Arabic-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 40))

    footer_data = [[ar(f'حُررت بـ {res.city if res else ""}، بتاريخ {today_str}'), ar('إمضاء وختم المتصرف')]]
    ft = Table(footer_data, colWidths=[7.5*cm, 7.5*cm])
    ft.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
    ]))
    elements.append(ft)

    doc.build(elements)
    buf.seek(0)
    safe_name = w.full_name.replace(' ', '_')
    return send_file(buf, mimetype='application/pdf', as_attachment=False,
                      download_name=f'attestation_salaire_{safe_name}.pdf')


@app.route('/api/workers/<int:wid>', methods=['DELETE'])
@login_required
def delete_worker(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    db.session.delete(w)
    db.session.commit()
    return jsonify({'ok':True})

def calculate_morocco_payroll(base_salary, allowances=0, num_children=0):
    """حساب الأجرة وفق التشريع المغربي 2026: CNSS 4.48% (مسقوف 6000), AMO 2.26% (بدون سقف), IR"""
    gross = round((base_salary or 0) + (allowances or 0), 2)
    cnss_base = min(gross, 6000.0)
    cnss_employee = round(cnss_base * 0.0448, 2)
    amo_employee = round(gross * 0.0226, 2)
    social_total = cnss_employee + amo_employee

    taxable_before = max(0, gross - social_total)
    abattement = min(taxable_before * 0.35, 2916.67)
    net_taxable = max(0, round(taxable_before - abattement, 2))

    brackets = [
        (0,        3333.33,      0.00, 0.00),
        (3333.34,  5000.00,      0.10, 333.33),
        (5000.01,  6666.67,      0.20, 833.33),
        (6666.68,  8333.33,      0.30, 1500.00),
        (8333.34,  15000.00,     0.34, 1833.33),
        (15000.01, float('inf'), 0.37, 2283.33),
    ]
    ir_raw = 0.0
    for lo, hi, rate, ded in brackets:
        if lo <= net_taxable <= hi:
            ir_raw = net_taxable * rate - ded
            break

    family_ded = min(num_children or 0, 6) * 30.0
    ir_amount = max(0, round(ir_raw - family_ded, 2))

    total_ded = round(cnss_employee + amo_employee + ir_amount, 2)
    net_salary = round(gross - total_ded, 2)

    return {
        'gross_salary': gross, 'cnss_employee': cnss_employee,
        'amo_employee': amo_employee, 'ir_amount': ir_amount,
        'total_deductions': total_ded, 'net_salary': net_salary,
    }

@app.route('/api/workers/<int:wid>', methods=['GET'])
@login_required
def get_worker(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    return jsonify(w.to_dict())

@app.route('/api/workers/<int:wid>', methods=['PUT'])
@login_required
def update_worker(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json() or {}
    if 'full_name' in data: w.full_name = data['full_name']
    if 'role' in data: w.role = data['role']
    if 'phone' in data: w.phone = data['phone']
    if 'cin' in data: w.cin = data['cin']
    if 'cnss_number' in data: w.cnss_number = data['cnss_number']
    if 'cnss_status' in data: w.cnss_status = data['cnss_status']
    if 'salary' in data: w.salary = float(data['salary'] or 0)
    if 'status' in data: w.status = data['status']
    if 'notes' in data: w.notes = data['notes']
    if 'contract_type' in data: w.contract_type = data['contract_type']
    if 'weekly_hours' in data: w.weekly_hours = int(data['weekly_hours'] or 0)
    if 'address' in data: w.address = data['address']
    if 'family_status' in data: w.family_status = data['family_status']
    if 'num_children' in data: w.num_children = int(data['num_children'] or 0)
    if 'allowances' in data: w.allowances = float(data['allowances'] or 0)
    if data.get('start_date'):
        w.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    if data.get('birth_date'):
        w.birth_date = datetime.strptime(data['birth_date'], '%Y-%m-%d').date()
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/workers/<int:wid>/wages', methods=['GET'])
@login_required
def get_worker_wages(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    wages = WorkerWage.query.filter_by(worker_id=wid).order_by(WorkerWage.year.desc(), WorkerWage.month.desc()).all()
    return jsonify([wg.to_dict() for wg in wages])

@app.route('/api/workers/<int:wid>/wages/generate', methods=['POST'])
@login_required
def generate_worker_wage(wid):
    u = get_current_user()
    w = Worker.query.get_or_404(wid)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json() or {}
    month = int(data.get('month') or datetime.utcnow().month)
    year = int(data.get('year') or datetime.utcnow().year)

    existing = WorkerWage.query.filter_by(worker_id=wid, month=month, year=year).first()
    if existing:
        return jsonify({'error': 'الأجرة لهذا الشهر مُسجّلة بالفعل'}), 400

    calc = calculate_morocco_payroll(w.salary, w.allowances, w.num_children)
    wg = WorkerWage(
        worker_id=wid, month=month, year=year,
        base_salary=w.salary or 0, allowances=w.allowances or 0,
        gross_salary=calc['gross_salary'], cnss_employee=calc['cnss_employee'],
        amo_employee=calc['amo_employee'], ir_amount=calc['ir_amount'],
        total_deductions=calc['total_deductions'], net_salary=calc['net_salary'],
        status='غير مدفوعة'
    )
    db.session.add(wg)
    db.session.commit()
    return jsonify({'ok': True, 'id': wg.id, 'wage': wg.to_dict()})

@app.route('/api/workers/wages/<int:wgid>/pay', methods=['PUT'])
@login_required
def pay_worker_wage(wgid):
    u = get_current_user()
    wg = WorkerWage.query.get_or_404(wgid)
    w = Worker.query.get_or_404(wg.worker_id)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=w.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    wg.status = 'مدفوعة'
    wg.payment_date = datetime.utcnow().date()
    db.session.commit()
    return jsonify({'ok': True})

# ═══════════════════════════════
#  RESIDENTS
# ═══════════════════════════════
@app.route('/api/residents')
@login_required
def get_residents():
    u = get_current_user()
    if u.role == 'admin':
        residents = Resident.query.all()
        apts = Apartment.query.all()
    else:
        res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
        residents = Resident.query.filter(Resident.residence_id.in_(res_ids)).all()
        apts = Apartment.query.filter(Apartment.residence_id.in_(res_ids)).all()

    # IDs of apartments already linked to a real account
    linked_apt_ids = set(r.apartment_id for r in residents if r.apartment_id)
    linked_apt_keys = set((r.apt_number, r.residence_id) for r in residents if r.apt_number)

    # لكل شقة نأخذ الساكن الأحدث فقط (approved)
    seen_apartments = {}
    for r in residents:
        key = r.apartment_id
        if key is None:
            seen_apartments[f'no_apt_{r.id}'] = r
        elif key not in seen_apartments:
            seen_apartments[key] = r
        else:
            existing = seen_apartments[key]
            r_date = r.approved_at or r.created_at or 0
            e_date = existing.approved_at or existing.created_at or 0
            if r_date and e_date and r_date > e_date:
                seen_apartments[key] = r
    result = []
    for r in seen_apartments.values():
        d = r.to_dict()
        d['has_account'] = True
        result.append(d)

    # Add apartments with owner/tenant data but no real account
    for apt in apts:
        if apt.id in linked_apt_ids or (apt.number, apt.residence_id) in linked_apt_keys:
            continue
        if not apt.owner_name and not apt.tenant_name:
            continue
        if apt.owner_name or apt.tenant_name:
            result.append({
                'id': f'apt_{apt.id}',
                'full_name': apt.owner_name or apt.tenant_name or '',
                'first_name': (apt.owner_name or apt.tenant_name or '').split()[0],
                'last_name': ' '.join((apt.owner_name or apt.tenant_name or '').split()[1:]),
                'phone': apt.owner_phone or apt.tenant_phone or '',
                'owner_name': apt.owner_name or '',
                'owner_phone': apt.owner_phone or '',
                'tenant_name': apt.tenant_name or '',
                'tenant_phone': apt.tenant_phone or '',
                'email': '',
                'resident_type': 'owner',
                'status': 'no_account',
                'apartment_id': apt.id,
                'apartment_number': str(apt.number),
                'residence_id': apt.residence_id,
                'has_account': False,
            })

    return jsonify(result)

@app.route('/api/residents/<int:rid2>', methods=['PUT'])
@login_required
def update_resident(rid2):
    u = get_current_user()
    r = Resident.query.get_or_404(rid2)
    if u.role != 'admin':
        res = Residence.query.filter_by(id=r.residence_id, user_id=u.id).first()
        if not res:
            return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json() or {}
    if 'status' in data:
        r.status = data['status']
        if data['status'] == 'approved':
            r.approved_at = datetime.utcnow()
            # حذف الحساب القديم المصادق عليه لنفس الشقة ونفس النوع (مالك/قاطن وحيد)
            if r.apartment_id:
                old_residents = Resident.query.filter(
                    Resident.apartment_id == r.apartment_id,
                    Resident.status == 'approved',
                    Resident.id != r.id
                ).all()
                for old_r in old_residents:
                    db.session.delete(old_r)
            # ربط الشقة تلقائياً عند المصادقة
            if r.apartment_id:
                apt = Apartment.query.get(r.apartment_id)
                if apt:
                    apt.tenant_name = f'{r.first_name} {r.last_name}'
                    apt.tenant_phone = r.phone
            # ربط معلومات القاطن بالشقة
            if r.apartment_id:
                apt = Apartment.query.get(r.apartment_id)
                if apt:
                    if r.resident_type == 'owner':
                        apt.owner_name = f'{r.first_name} {r.last_name}'
                        apt.owner_phone = r.phone or ''
                    else:
                        apt.tenant_name = f'{r.first_name} {r.last_name}'
                        apt.tenant_phone = r.phone or ''
            # إرسال إيميل للقاطن
            if r.email and GMAIL_USER and GMAIL_PASS:
                try:
                    import smtplib
                    from email.mime.text import MIMEText
                    from email.mime.multipart import MIMEMultipart
                    res = Residence.query.get(r.residence_id)
                    msg = MIMEMultipart('alternative')
                    msg['Subject'] = "SyndikPro - Votre compte a ete approuve"
                    msg['From'] = f"SyndikPro <{GMAIL_USER}>"
                    msg['To'] = r.email
                    body = f"""<div dir="rtl" style="font-family:Arial;max-width:500px;margin:auto;background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1)"><div style="background:linear-gradient(135deg,#0f766e,#0d9488);padding:32px;text-align:center"><div style="font-size:36px">&#127970;</div><div style="color:#fff;font-size:20px;font-weight:700;margin-top:8px">SyndikPro</div><div style="color:#99f6e4;font-size:11px;margin-top:4px">&#1606;&#1592;&#1575;&#1605; &#1578;&#1583;&#1576;&#1610;&#1585; &#1575;&#1604;&#1573;&#1602;&#1575;&#1605;&#1575;&#1578;</div></div><div style="background:#059669;padding:8px;text-align:center"><span style="color:#fff;font-size:13px;font-weight:600">&#10003; &#1578;&#1605; &#1602;&#1576;&#1608;&#1604; &#1591;&#1604;&#1576;&#1603;</span></div><div style="padding:28px"><p style="font-size:16px;font-weight:700;color:#1e293b;margin:0 0 12px">&#1605;&#1585;&#1581;&#1576;&#1575; {r.first_name} {r.last_name}&#x60;</p><p style="color:#64748b;font-size:13px;line-height:1.8;margin:0 0 20px">&#1578;&#1605; &#1602;&#1576;&#1608;&#1604; &#1591;&#1604;&#1576;&#1603; &#1601;&#1610; <strong style="color:#0f766e">{res.name if res else "&#1575;&#1604;&#1573;&#1602;&#1575;&#1605;&#1577;"}</strong> &#1585;&#1587;&#1605;&#1610;&#1575;&#1611;.</p><div style="text-align:center;margin:20px 0"><a href="https://hicham.pythonanywhere.com" style="background:#0f766e;color:#fff;text-decoration:none;padding:12px 32px;border-radius:8px;font-size:14px;font-weight:700;display:inline-block">&#1578;&#1587;&#1580;&#1610;&#1604; &#1575;&#1604;&#1583;&#1582;&#1608;&#1604;</a></div></div><div style="background:#f8fafc;padding:14px;text-align:center;border-top:1px solid #e2e8f0"><p style="color:#94a3b8;font-size:11px;margin:0">&#169; 2026 SyndikPro</p></div></div>"""
                    msg.attach(MIMEText(body, 'html', 'utf-8'))
                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                        smtp.login(GMAIL_USER, GMAIL_PASS)
                        smtp.sendmail(GMAIL_USER, r.email, msg.as_string())
                except Exception as e:
                    print(f"خطأ في إرسال الإيميل: {e}")
    if 'apartment_id' in data: r.apartment_id = data['apartment_id']
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  GLOBAL NOTIFICATIONS (for syndics)
# ═══════════════════════════════
@app.route('/api/announcements')
@login_required
def get_announcements():
    u = get_current_user()
    q = GlobalNotification.query
    if u.role != 'admin':
        q = q.filter(GlobalNotification.target.in_(['all', u.status]))
    ns = q.order_by(GlobalNotification.created_at.desc()).limit(10).all()
    return jsonify([n.to_dict() for n in ns])

# ═══════════════════════════════
#  DEMO LOGIN
# ═══════════════════════════════
@app.route('/api/demo', methods=['POST'])
def demo_login():
    u = User.query.filter_by(username='demo').first()
    if not u:
        u = User(
            username='demo', full_name='Demo سانديك',
            password_hash=generate_password_hash('demo'),
            role='syndic', status='active', plan='pro',
            subscription_start=datetime.utcnow(),
            subscription_end=datetime(2099,12,31)
        )
        db.session.add(u)
        db.session.commit()
        # إقامة تجريبية
        res = Residence(user_id=u.id, name='إقامة الياسمين', address='شارع محمد الخامس', city='الدار البيضاء', total_floors=5)
        db.session.add(res)
        db.session.commit()
        for i in range(1,9):
            a = Apartment(residence_id=res.id, number=str(i), floor=(i-1)//2+1,
                         owner_name=f'ساكن {i}', monthly_fee=300)
            db.session.add(a)
        db.session.commit()
    session['user_id'] = u.id
    session.permanent = True
    return jsonify({'ok':True,'role':u.role,'name':u.full_name})

# ═══════════════════════════════
#  PENDING ALERTS for syndics
# ═══════════════════════════════
@app.route('/api/syndic/alerts')
@login_required
def syndic_alerts():
    u = get_current_user()
    residences = Residence.query.filter_by(user_id=u.id).all()
    late_count = 0
    complaint_count = 0
    for r in residences:
        for a in r.apartments:
            now = datetime.utcnow()
            paid = Payment.query.filter_by(apartment_id=a.id,month=now.month,year=now.year,status='paid').first()
            if not paid: late_count += 1
            complaint_count += Complaint.query.filter_by(apartment_id=a.id,status='open').count()
    return jsonify({'late':late_count,'complaints':complaint_count,'residences':len(residences)})



# ═══════════════════════════════
#  ASSEMBLIES — الجمع العام
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/assemblies')
@login_required
def get_assemblies(rid):
    items = Assembly.query.filter_by(residence_id=rid).order_by(Assembly.date.desc()).all()
    return jsonify([a.to_dict() for a in items])

@app.route('/api/assemblies', methods=['POST'])
@login_required
def add_assembly():
    data = request.get_json() or {}
    a = Assembly(
        residence_id=data['residence_id'],
        title=data['title'],
        description=data.get('description',''),
        date=datetime.strptime(data['date'], '%Y-%m-%dT%H:%M') if 'T' in str(data.get('date','')) else datetime.strptime(data['date'], '%Y-%m-%d'),
        location=data.get('location',''),
        status=data.get('status','upcoming')
    )
    db.session.add(a)
    db.session.commit()
    return jsonify({'ok':True,'id':a.id})

@app.route('/api/assemblies/<int:aid>', methods=['PUT'])
@login_required
def update_assembly(aid):
    a = Assembly.query.get_or_404(aid)
    data = request.get_json() or {}
    for k in ['title','description','location','status']:
        if k in data: setattr(a, k, data[k])
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/assemblies/<int:aid>', methods=['DELETE'])
@login_required
def delete_assembly(aid):
    a = Assembly.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    return jsonify({'ok':True})

# Votes
@app.route('/api/assemblies/<int:aid>/votes')
@login_required
def get_votes(aid):
    votes = Vote.query.filter_by(assembly_id=aid).all()
    return jsonify([v.to_dict() for v in votes])

@app.route('/api/votes', methods=['POST'])
@login_required
def add_vote():
    data = request.get_json() or {}
    v = Vote(assembly_id=data['assembly_id'], question=data['question'], status='open')
    db.session.add(v)
    db.session.commit()
    
    # إرسال إشعارات للسكان
    try:
        assembly = Assembly.query.get(data['assembly_id'])
        if assembly:
            residents = Resident.query.filter_by(residence_id=assembly.residence_id, status='approved').all()
            for r in residents:
                notif = Notification(
                    resident_id=r.id,
                    title='🗳️ تصويت جديد متاح',
                    body=f'{data.get("question", "سؤال جديد")}',
                    type='vote'
                )
                db.session.add(notif)
            db.session.commit()
    except:
        pass
    
    return jsonify({'ok':True,'id':v.id})


@app.route('/api/syndic/assemblies')
@login_required
def get_syndic_assemblies():
    u = get_current_user()
    res = Residence.query.filter_by(user_id=u.id).first()
    if not res: return jsonify([])
    assemblies = Assembly.query.filter_by(residence_id=res.id).order_by(Assembly.id.desc()).all()
    return jsonify([a.to_dict() for a in assemblies])

@app.route('/api/syndic/all-votes')
@login_required
def get_syndic_all_votes():
    u = get_current_user()
    res = Residence.query.filter_by(user_id=u.id).first()
    result = []
    assemblies = Assembly.query.filter_by(residence_id=res.id).all() if res else Assembly.query.all()
    for a in assemblies:
        for v in Vote.query.filter_by(assembly_id=a.id).all():
            d = v.to_dict()
            d['assembly_title'] = a.title or ''
            result.append(d)
    return jsonify(result)

@app.route('/api/resident/meetings')
@resident_required
def get_resident_meetings():
    r = get_current_resident()
    if not r:
        return jsonify([])
    apt = Apartment.query.get(r.apartment_id) if r.apartment_id else None
    residence_id = apt.residence_id if apt else r.residence_id
    if not residence_id:
        return jsonify([])
    assemblies = Assembly.query.filter_by(residence_id=residence_id).order_by(Assembly.date.desc()).all()
    result = []
    for a in assemblies:
        result.append({
            'id': a.id,
            'title': a.title or 'اجتماع',
            'date': str(a.date)[:10] if a.date else '',
            'location': a.location or '',
            'status': a.status or '',
            'decisions': a.decisions or '',
            'report_notes': a.report_notes or '',
        })
    return jsonify(result)

@app.route('/api/resident/votes')
@resident_required
def get_resident_votes():
    r = get_current_resident()
    if not r:
        return jsonify([])
    apt = Apartment.query.get(r.apartment_id) if r.apartment_id else None
    residence_id = apt.residence_id if apt else r.residence_id
    if not residence_id:
        return jsonify([])
    assemblies = Assembly.query.filter_by(residence_id=residence_id).all()
    result = []
    for a in assemblies:
        for v in Vote.query.filter_by(assembly_id=a.id).all():
            d = v.to_dict()
            resp = VoteResponse.query.filter_by(vote_id=v.id, resident_id=r.id).first()
            d['my_vote'] = resp.choice if resp else None
            result.append(d)
    return jsonify(result)

@app.route('/api/votes/<int:vid>/cast', methods=['POST'])
@resident_required
def cast_vote(vid):
    r = get_current_resident()
    data = request.get_json() or {}
    choice = data.get('choice','')
    if choice not in ['yes','no','abstain']:
        return jsonify({'error':'اختيار غير صالح'}),400
    existing = VoteResponse.query.filter_by(vote_id=vid, resident_id=r.id).first()
    if existing:
        return jsonify({'error':'لقد صوت مسبقاً'}),400
    v = Vote.query.get_or_404(vid)
    if v.status != 'open':
        return jsonify({'error':'التصويت مغلق'}),400
    resp = VoteResponse(vote_id=vid, resident_id=r.id, choice=choice)
    db.session.add(resp)
    if choice == 'yes': v.yes_count += 1
    elif choice == 'no': v.no_count += 1
    else: v.abstain += 1
    db.session.commit()
    return jsonify({'ok':True})


@app.route('/api/votes/<int:vid>/details')
@login_required
def vote_details(vid):
    v = Vote.query.get_or_404(vid)
    a = Assembly.query.get(v.assembly_id)
    res = Residence.query.filter_by(user_id=get_current_user().id).first()
    if not res:
        return jsonify({'error':'لا توجد إقامة'}),400
    apts = Apartment.query.filter_by(residence_id=res.id).all()
    responses = VoteResponse.query.filter_by(vote_id=vid).all()
    voted_ids = {r.resident_id for r in responses}
    voters = []
    non_voters = []
    for apt in apts:
        resident = Resident.query.filter_by(apartment_id=apt.id).first()
        if not resident:
            non_voters.append({'apt': apt.number, 'name': apt.owner_name or '—'})
            continue
        resp = VoteResponse.query.filter_by(vote_id=vid, resident_id=resident.id).first()
        if resp:
            voters.append({'apt': apt.number, 'name': apt.owner_name or resident.full_name or '—', 'choice': resp.choice})
        else:
            non_voters.append({'apt': apt.number, 'name': apt.owner_name or resident.full_name or '—'})
    return jsonify({
        'vote': v.to_dict(),
        'assembly': a.title if a else '',
        'total_apts': len(apts),
        'voters': voters,
        'non_voters': non_voters
    })


@app.route('/api/assemblies/<int:aid>/cancel', methods=['POST'])
@login_required
def cancel_assembly(aid):
    a = Assembly.query.get_or_404(aid)
    a.status = 'cancelled'
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/assemblies/<int:aid>/report', methods=['POST'])
@login_required
def save_assembly_report(aid):
    a = Assembly.query.get_or_404(aid)
    data = request.get_json() or {}
    a.attendees = data.get('attendees','')
    a.absentees = data.get('absentees','')
    a.president = data.get('president','')
    a.decisions = data.get('decisions','')
    a.report_notes = data.get('report_notes','')
    a.status = 'closed'
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/votes/<int:vid>/close', methods=['POST'])
@login_required
def close_vote(vid):
    v = Vote.query.get_or_404(vid)
    v.status = 'closed'
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  DOCUMENTS — وثائق
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/documents')
@login_required
def get_documents(rid):
    docs = Document.query.filter_by(residence_id=rid).order_by(Document.created_at.desc()).all()
    return jsonify([d.to_dict() for d in docs])

@app.route('/api/documents', methods=['POST'])
@login_required
def add_document():
    u = get_current_user()
    title    = request.form.get('title','').strip()
    rid      = request.form.get('residence_id')
    category = request.form.get('category','عام')
    is_pub   = request.form.get('is_public','true') == 'true'
    file_path = ''
    file_name = ''
    if 'file' in request.files:
        f = request.files['file']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            fn = secure_filename(f.filename)
            import time
            fn = f"{int(time.time())}_{fn}"
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            file_path = f'static/uploads/{fn}'
            file_name = f.filename
    d = Document(
        residence_id=int(rid), title=title, category=category,
        file_path=file_path, file_name=file_name,
        is_public=is_pub, uploaded_by=u.id
    )
    db.session.add(d)
    db.session.commit()
    return jsonify({'ok':True,'id':d.id})

@app.route('/api/documents/<int:did>', methods=['DELETE'])
@login_required
def delete_document(did):
    d = Document.query.get_or_404(did)
    db.session.delete(d)
    db.session.commit()
    return jsonify({'ok':True})

# وثائق للساكن
@app.route('/api/resident/documents')
@resident_required
def resident_documents():
    r = get_current_resident()
    docs = Document.query.filter_by(residence_id=r.residence_id, is_public=True).order_by(Document.created_at.desc()).all()
    return jsonify([d.to_dict() for d in docs])

# ═══════════════════════════════
#  RESERVE FUND — صندوق الاحتياط
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/fund')
@login_required
def get_fund(rid):
    fund = ReserveFund.query.filter_by(residence_id=rid).first()
    if not fund:
        fund = ReserveFund(residence_id=rid, balance=0, target=0)
        db.session.add(fund)
        db.session.commit()
    txs = FundTransaction.query.filter_by(fund_id=fund.id).order_by(FundTransaction.created_at.desc()).limit(20).all()
    return jsonify({'fund': fund.to_dict(), 'transactions': [t.to_dict() for t in txs]})

@app.route('/api/residences/<int:rid>/fund', methods=['PUT'])
@login_required
def update_fund(rid):
    fund = ReserveFund.query.filter_by(residence_id=rid).first()
    if not fund:
        fund = ReserveFund(residence_id=rid)
        db.session.add(fund)
    data = request.get_json() or {}
    if 'target' in data: fund.target = float(data['target'])
    if 'description' in data: fund.description = data['description']
    fund.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/residences/<int:rid>/fund/transaction', methods=['POST'])
@login_required
def add_fund_transaction(rid):
    fund = ReserveFund.query.filter_by(residence_id=rid).first()
    if not fund:
        fund = ReserveFund(residence_id=rid)
        db.session.add(fund)
        db.session.commit()
    data = request.get_json() or {}
    amount = float(data.get('amount', 0))
    tx_type = data.get('type', 'in')
    tx = FundTransaction(
        fund_id=fund.id, amount=amount, type=tx_type,
        note=data.get('note',''),
        date=datetime.strptime(data['date'],'%Y-%m-%d').date() if data.get('date') else date.today()
    )
    db.session.add(tx)
    if tx_type == 'in':
        fund.balance += amount
    else:
        fund.balance = max(0, fund.balance - amount)
    fund.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  ANNOUNCEMENTS — إعلانات الإقامة
# ═══════════════════════════════
@app.route('/api/residences/<int:rid>/announcements')
@login_required
def get_res_announcements(rid):
    items = Announcement.query.filter_by(residence_id=rid).order_by(Announcement.created_at.desc()).limit(20).all()
    return jsonify([a.to_dict() for a in items])

@app.route('/api/announcements', methods=['POST'])
@login_required
def add_announcement():
    u = get_current_user()
    data = request.get_json() or {}
    a = Announcement(
        residence_id=data['residence_id'],
        title=data['title'],
        body=data.get('body',''),
        type=data.get('type','info'),
        created_by=u.id
    )
    db.session.add(a)
    db.session.commit()
    return jsonify({'ok':True,'id':a.id})

@app.route('/api/announcements/<int:aid>', methods=['DELETE'])
@login_required
def delete_announcement(aid):
    a = Announcement.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    return jsonify({'ok':True})

# إعلانات للساكن
@app.route('/api/resident/announcements')
@resident_required
def resident_announcements():
    r = get_current_resident()
    items = Announcement.query.filter_by(residence_id=r.residence_id).order_by(Announcement.created_at.desc()).limit(10).all()
    result = []
    for a in items:
        d = a.to_dict()
        try:
            viewed = json.loads(a.viewed_by or '[]')
        except:
            viewed = []
        d['is_viewed'] = r.id in viewed
        result.append(d)
    # احصل على الـ dismissed من session
    dismissed = session.get('dismissed_announcements', [])
    
    # صفّي الإعلانات المرفوضة
    announcements = [a for a in announcements if a.id not in dismissed]
    
    return jsonify([{'id': a.id, 'title': a.title, 'body': a.body, 'created_at': a.created_at.strftime('%Y-%m-%d %H:%M')} for a in announcements])


@app.route('/api/resident/announcements/<int:ann_id>/mark-viewed', methods=['POST'])
@resident_required
def mark_announcement_viewed(ann_id):
    r = get_current_resident()
    a = Announcement.query.get(ann_id)
    if not a or a.residence_id != r.residence_id:
        return jsonify({'error': 'غير مصرح'}), 403
    try:
        viewed = json.loads(a.viewed_by or '[]')
        if r.id not in viewed:
            viewed.append(r.id)
            a.viewed_by = json.dumps(viewed)
            a.viewed_count = len(viewed)
            db.session.commit()
        return jsonify({'ok': True})
    except:
        db.session.rollback()
        return jsonify({'error': 'خطأ'}), 500

# الجمع العام للساكن
@app.route('/api/resident/assemblies')
@resident_required
def resident_assemblies():
    r = get_current_resident()
    items = Assembly.query.filter_by(residence_id=r.residence_id).order_by(Assembly.date.desc()).all()
    result = []
    for a in items:
        d = a.to_dict()
        # check if resident already voted in each vote
        voted_map = {}
        for v in a.votes:
            resp = VoteResponse.query.filter_by(vote_id=v.id, resident_id=r.id).first()
            voted_map[v.id] = resp.choice if resp else None
        d['votes'] = [v.to_dict() for v in a.votes]
        d['voted_map'] = voted_map
        result.append(d)
    return jsonify(result)

# ═══════════════════════════════════════════════
#  ONLINE PAYMENT — نظام الأداء الإلكتروني
# ═══════════════════════════════════════════════

# الساكن يرى شهوره غير المدفوعة
@app.route('/api/resident/unpaid-months')
@resident_required
def resident_unpaid_months():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify([])
@app.route('/api/resident/announcements/<int:ann_id>/dismiss', methods=['POST'])
@resident_required
def dismiss_announcement(ann_id):
    r = get_current_resident()
    if not r:
        return jsonify({'error': 'not found'}), 404
    
    # حفظ في session أو في جدول جديد
    if 'dismissed_announcements' not in session:
        session['dismissed_announcements'] = []
    
    if ann_id not in session['dismissed_announcements']:
        session['dismissed_announcements'].append(ann_id)
        session.modified = True
    
    return jsonify({'ok': True})

    apt = Apartment.query.get(r.apartment_id)
    now = datetime.utcnow()
    year = request.args.get('year', now.year, type=int)
    MONTHS_AR = ['يناير','فبراير','مارس','أبريل','ماي','يونيو',
                 'يوليوز','غشت','شتنبر','أكتوبر','نونبر','دجنبر']
    result = []
    for m in range(1, 13):
        pay = Payment.query.filter_by(apartment_id=r.apartment_id, month=m, year=year).first()
        pending = OnlinePaymentRequest.query.filter_by(
            apartment_id=r.apartment_id, month=m, year=year, status='pending'
        ).first()
        result.append({
            'month': m, 'year': year,
            'label': f'{MONTHS_AR[m-1]} {year}',
            'amount': apt.monthly_fee if apt else 0,
            'status': pay.status if pay else 'unpaid',
            'has_pending': pending is not None,
            'pending_id': pending.id if pending else None,
        })
    return jsonify(result)

# الساكن يرسل طلب أداء
@app.route('/api/resident/pay', methods=['POST'])
@resident_required
def resident_submit_payment():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify({'error': 'لم يتم تعيين شقة لك بعد'}), 400

    import uuid as _uuid

    method = request.form.get('method', '')
    tx_ref = request.form.get('tx_ref', '').strip()
    note   = request.form.get('note', '').strip()

    # دعم دفعة لعدة أشهر: months_json = [{"month":1,"year":2026,"amount":300}, ...]
    # مع الحفاظ على التوافقية مع الإرسال القديم (شهر واحد: month/year/amount)
    months_json = request.form.get('months_json', '').strip()
    items = []
    if months_json:
        import json as _json
        try:
            raw_items = _json.loads(months_json)
        except Exception:
            return jsonify({'error': 'صيغة الأشهر غير صحيحة'}), 400
        for it in raw_items:
            try:
                m = int(it.get('month', 0))
                y = int(it.get('year', 0))
                a = float(it.get('amount', 0))
            except Exception:
                return jsonify({'error': 'بيانات ناقصة'}), 400
            if not all([m, y, a]):
                return jsonify({'error': 'بيانات ناقصة'}), 400
            items.append({'month': m, 'year': y, 'amount': a})
    else:
        month  = int(request.form.get('month', 0))
        year   = int(request.form.get('year', 0))
        amount = float(request.form.get('amount', 0))
        if not all([method, month, year, amount]):
            return jsonify({'error': 'بيانات ناقصة'}), 400
        items = [{'month': month, 'year': year, 'amount': amount}]

    if not method or not items:
        return jsonify({'error': 'بيانات ناقصة'}), 400

    # تحقق من عدم وجود طلب معلق لأي شهر من الأشهر المطلوبة (قبل إنشاء أي سجل)
    for it in items:
        existing = OnlinePaymentRequest.query.filter_by(
            apartment_id=r.apartment_id, month=it['month'], year=it['year'], status='pending'
        ).first()
        if existing:
            return jsonify({'error': f"يوجد طلب أداء معلق لشهر {it['month']}/{it['year']}"}), 400

    proof_path = ''
    if 'proof' in request.files:
        f = request.files['proof']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            import time
            fn = f"{int(time.time())}_{secure_filename(f.filename)}"
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            proof_path = f'static/uploads/{fn}'

    batch_id = _uuid.uuid4().hex[:32] if len(items) > 1 else None
    total_amount = 0.0
    created_ids = []

    for it in items:
        req = OnlinePaymentRequest(
            resident_id=r.id,
            apartment_id=r.apartment_id,
            month=it['month'], year=it['year'], amount=it['amount'],
            method=method, tx_ref=tx_ref, note=note,
            proof_path=proof_path, status='pending',
            batch_id=batch_id
        )
        db.session.add(req)
        db.session.flush()
        created_ids.append(req.id)
        total_amount += it['amount']

    db.session.commit()

    # إشعار واحد فقط للساكن، يجمع كل الدفعة
    if len(items) > 1:
        body = f'طلبك بمبلغ {total_amount} درهم عبر {method} لعدد {len(items)} أشهر قيد المراجعة من طرف السانديك.'
    else:
        body = f'طلبك بمبلغ {total_amount} درهم عبر {method} قيد المراجعة من طرف السانديك.'

    notif = Notification(
        resident_id=r.id,
        title='تم إرسال طلب الأداء',
        body=body,
        type='info'
    )
    db.session.add(notif)
    db.session.commit()

    return jsonify({'ok': True, 'ids': created_ids, 'batch_id': batch_id})

# الساكن يرى طلباته
@app.route('/api/resident/pay-requests')
@resident_required
def resident_pay_requests():
    r = get_current_resident()
    reqs = OnlinePaymentRequest.query.filter_by(resident_id=r.id)           .order_by(OnlinePaymentRequest.created_at.desc()).limit(20).all()
    return jsonify([x.to_dict() for x in reqs])

# ─────────────────────────────────────────
#  SYNDIC — استقبال الأداءات
# ─────────────────────────────────────────

# قائمة طلبات الأداء لإقامة
@app.route('/api/residences/<int:rid>/pay-requests')
@login_required
def syndic_pay_requests(rid):
    status = request.args.get('status', '')
    apts = [a.id for a in Apartment.query.filter_by(residence_id=rid).all()]
    q = OnlinePaymentRequest.query.filter(OnlinePaymentRequest.apartment_id.in_(apts))
    if status:
        q = q.filter_by(status=status)
    items = q.order_by(OnlinePaymentRequest.created_at.desc()).all()
    return jsonify([x.to_dict() for x in items])

# جميع طلبات السانديك (كل إقاماته)
@app.route('/api/syndic/pay-requests')
@login_required
def syndic_all_pay_requests():
    u = get_current_user()
    res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
    apt_ids = [a.id for a in Apartment.query.filter(Apartment.residence_id.in_(res_ids)).all()]
    status = request.args.get('status', '')
    q = OnlinePaymentRequest.query.filter(OnlinePaymentRequest.apartment_id.in_(apt_ids))
    if status:
        q = q.filter_by(status=status)
    items = q.order_by(OnlinePaymentRequest.created_at.desc()).limit(50).all()
    return jsonify([x.to_dict() for x in items])

# قبول أو رفض طلب أداء
@app.route('/api/pay-requests/<int:req_id>/review', methods=['POST'])
@login_required
def review_pay_request(req_id):
    data = request.get_json() or {}
    pr = OnlinePaymentRequest.query.get_or_404(req_id)
    action = data.get('action', 'confirm')  # confirm | reject
    note = data.get('note', '')
    u = get_current_user()

    pr.reviewed_by = u.id
    pr.reviewed_at = datetime.utcnow()

    if action == 'confirm':
        pr.status = 'confirmed'
        # سجّل الدفع في جدول payments
        pay = Payment.query.filter_by(
            apartment_id=pr.apartment_id, month=pr.month, year=pr.year
        ).first()
        if pay:
            pay.status = 'paid'
            pay.amount = pr.amount
            pay.method = pr.method
            pay.date_paid = date.today()
        else:
            pay = Payment(
                apartment_id=pr.apartment_id,
                month=pr.month, year=pr.year,
                amount=pr.amount, status='paid',
                method=pr.method, date_paid=date.today()
            )
            db.session.add(pay)
        # إشعار الساكن
        notif = Notification(
            resident_id=pr.resident_id,
            title='✅ تم قبول أداءك',
            body=f'تم تأكيد أداء {pr.amount} درهم لشهر {pr.month}/{pr.year}.',
            type='success'
        )
        db.session.add(notif)
    else:
        pr.status = 'rejected'
        notif = Notification(
            resident_id=pr.resident_id,
            title='❌ تم رفض طلب الأداء',
            body=f'تم رفض طلب الأداء لشهر {pr.month}/{pr.year}. {note}',
            type='alert'
        )
        db.session.add(notif)

    db.session.commit()
    return jsonify({'ok': True, 'status': pr.status})

# إحصائيات الأداء الإلكتروني للسانديك
@app.route('/api/syndic/pay-stats')
@login_required
def syndic_pay_stats():
    u = get_current_user()
    res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
    apt_ids = [a.id for a in Apartment.query.filter(Apartment.residence_id.in_(res_ids)).all()]
    pending   = OnlinePaymentRequest.query.filter(
        OnlinePaymentRequest.apartment_id.in_(apt_ids),
        OnlinePaymentRequest.status=='pending').count()
    confirmed = OnlinePaymentRequest.query.filter(
        OnlinePaymentRequest.apartment_id.in_(apt_ids),
        OnlinePaymentRequest.status=='confirmed').count()
    rejected  = OnlinePaymentRequest.query.filter(
        OnlinePaymentRequest.apartment_id.in_(apt_ids),
        OnlinePaymentRequest.status=='rejected').count()
    total_confirmed = db.session.query(db.func.sum(OnlinePaymentRequest.amount)).filter(
        OnlinePaymentRequest.apartment_id.in_(apt_ids),
        OnlinePaymentRequest.status=='confirmed'
    ).scalar() or 0
    # توزيع الطرق
    methods = {}
    for m in ['paypal','cmi','virement','qr']:
        methods[m] = OnlinePaymentRequest.query.filter(
            OnlinePaymentRequest.apartment_id.in_(apt_ids),
            OnlinePaymentRequest.method==m,
            OnlinePaymentRequest.status=='confirmed'
        ).count()
    return jsonify({
        'pending': pending, 'confirmed': confirmed,
        'rejected': rejected, 'total_confirmed': total_confirmed,
        'methods': methods
    })

# وصل PDF لطلب أداء إلكتروني
@app.route('/api/pay-requests/<int:req_id>/receipt')
@login_required
def online_pay_receipt(req_id):
    pr  = OnlinePaymentRequest.query.get_or_404(req_id)
    apt = Apartment.query.get(pr.apartment_id)
    res = Residence.query.get(apt.residence_id) if apt else None
    r   = Resident.query.get(pr.resident_id)
    syndic = User.query.get(res.user_id) if res else None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    ar = ParagraphStyle('ar', fontName='Helvetica', fontSize=11, alignment=1, spaceAfter=6)
    ts = ParagraphStyle('ts', fontName='Helvetica-Bold', fontSize=16, alignment=1, spaceAfter=10)
    sm = ParagraphStyle('sm', fontName='Helvetica', fontSize=9, alignment=1, textColor=colors.grey)
    ok_s = ParagraphStyle('ok', fontName='Helvetica-Bold', fontSize=13, alignment=1,
                          textColor=colors.HexColor('#16a34a'), spaceAfter=8)

    MONTHS_FR = ['Janvier','Février','Mars','Avril','Mai','Juin',
                 'Juillet','Août','Septembre','Octobre','Novembre','Décembre']
    METHOD_LABEL = {'paypal':'PayPal','cmi':'Carte CMI','virement':'Virement bancaire','qr':'QR Code'}
    ref = f"ONL-{pr.year}-{pr.id:06d}"

    story = []
    story.append(Paragraph("SyndikPro", ts))
    story.append(Paragraph("Reçu de Paiement Électronique / وصل الأداء الإلكتروني", ar))
    story.append(Spacer(1, 0.3*cm))

    status_color = colors.HexColor('#16a34a') if pr.status=='confirmed' else colors.HexColor('#dc2626')
    status_text = '✓ Confirmé / مؤكد' if pr.status=='confirmed' else '✗ Rejeté / مرفوض'
    story.append(Paragraph(status_text, ParagraphStyle('st', fontName='Helvetica-Bold',
                            fontSize=14, alignment=1, textColor=status_color, spaceAfter=10)))

    rows = [
        ['Référence', ref],
        ['Résidence', res.name if res else '?'],
        ['Appartement', f"N° {apt.number}" if apt else '?'],
        ['Résident', f"{r.first_name} {r.last_name}" if r else '?'],
        ['Mois / الشهر', f"{MONTHS_FR[(pr.month or 1)-1]} {pr.year}"],
        ['Montant / المبلغ', f"{pr.amount:.2f} MAD"],
        ['Méthode / الطريقة', METHOD_LABEL.get(pr.method, pr.method)],
        ['Référence transaction', pr.tx_ref or '—'],
        ['Date demande', pr.created_at.strftime('%d/%m/%Y %H:%M')],
        ['Date confirmation', pr.reviewed_at.strftime('%d/%m/%Y %H:%M') if pr.reviewed_at else '—'],
    ]
    t = Table(rows, colWidths=[7*cm, 10*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#dbeafe')),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    qr_data = f"SYNDIKPRO-ONLINE|{ref}|{pr.amount}|{pr.status}"
    qr_b64 = gen_qr_base64(qr_data)
    qr_buf = io.BytesIO(base64.b64decode(qr_b64))
    story.append(Paragraph("QR de vérification", sm))
    story.append(Image(qr_buf, width=3*cm, height=3*cm))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Syndic: {syndic.full_name if syndic else '?'} | {syndic.phone if syndic else ''}", sm))
    story.append(Paragraph(f"Généré par SyndikPro le {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}", sm))

    doc.build(story)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="recu_online_{ref}.pdf"'
    return resp


# ─────────────────────────────────────────
#  SYNDIC UPDATE PROFILE
# ─────────────────────────────────────────
@app.route('/api/syndic/profile', methods=['PUT'])
@login_required
def syndic_update_profile():
    u = get_current_user()
    data = request.get_json() or {}
    if 'full_name'     in data: u.full_name     = data['full_name']
    if 'phone'         in data: u.phone         = data['phone']
    if 'city'          in data: u.city          = data['city']
    if 'neighborhood'  in data: u.neighborhood  = data['neighborhood']
    if 'country'       in data: u.country       = data['country']
    db.session.commit()
    return jsonify({'ok': True})

# ─────────────────────────────────────────
#  SYNDIC PAYMENT INFO (RIB, PayPal)
# ─────────────────────────────────────────
@app.route('/api/syndic/payment-info', methods=['GET','POST'])
@login_required
def syndic_payment_info():
    u = get_current_user()
    if request.method == 'GET':
        return jsonify({
            'paypal_email': u.paypal_email or '',
            'bank_rib':     u.bank_rib or '',
            'bank_name':    u.bank_name or '',
        })
    data = request.get_json() or {}
    if 'paypal_email' in data: u.paypal_email = data['paypal_email']
    if 'bank_rib'     in data: u.bank_rib     = data['bank_rib']
    if 'bank_name'    in data: u.bank_name    = data['bank_name']
    db.session.commit()
    return jsonify({'ok': True})

# Resident fetches syndic payment info for their residence
@app.route('/api/resident/syndic-payment-info')
@resident_required
def resident_syndic_payment_info():
    r = get_current_resident()
    res = Residence.query.get(r.residence_id)
    if not res:
        return jsonify({'paypal_email':'','bank_rib':'','bank_name':''})
    u = User.query.get(res.user_id)
    return jsonify({
        'paypal_email': u.paypal_email or '' if u else '',
        'bank_rib':     u.bank_rib or '' if u else '',
        'bank_name':    u.bank_name or '' if u else '',
        'syndic_name':  u.full_name or '' if u else '',
    })



@app.route('/api/syndic/subscription-payment', methods=['POST'])
@login_required
def syndic_subscription_payment():
    from models import SubscriptionRequest, SubscriptionPlan
    u = get_current_user()
    plan = request.form.get('plan', u.plan)
    months = int(request.form.get('months', 1))
    total = float(request.form.get('total_amount', '0'))
    method = request.form.get('method', '')
    proof_path = ''
    
    if 'proof' in request.files:
        f = request.files['proof']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            import time
            fn = f"{int(time.time())}_{secure_filename(f.filename)}"
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            proof_path = f'static/uploads/{fn}'
    
    plan_obj = SubscriptionPlan.query.filter_by(name=plan).first()
    if not plan_obj:
        return jsonify({'error': 'خطة غير موجودة'}), 404
    
    req = SubscriptionRequest(
        user_id=u.id,
        plan_id=plan_obj.id,
        months=months,
        base_price=plan_obj.price_monthly * months,
        discount_percent=0,
        discount_amount=0,
        final_price=total,
        status='pending',
        payment_method=method,
        receipt_path=proof_path
    )
    db.session.add(req)
    db.session.commit()
    
    # إشعار واحد للأدمين
    try:
        gn = GlobalNotification(
            title=f'💳 طلب اشتراك جديد — {u.full_name}',
            body=(f'الخطة: {plan} | المدة: {months} شهر | المجموع: {total} درهم | الطريقة: {method}' + (f' | وصل: /{proof_path}' if proof_path else '')),
            type='payment', target='admin'
        )
        db.session.add(gn)
        db.session.commit()
    except Exception as e:
        print(f'Admin notification error: {e}')

    u.renewal_pending = True
    db.session.commit()
    return jsonify({'ok': True, 'request_id': req.id})


@app.route('/api/resident/payment-request', methods=['POST'])
@resident_required
def resident_payment_request():
    r = get_current_resident()
    method = request.form.get('method','cash')
    ref = request.form.get('ref','')
    proof_path = ''
    if 'proof' in request.files:
        f = request.files['proof']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            import time
            fn = f"{int(time.time())}_{secure_filename(f.filename)}"
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            proof_path = f'static/uploads/{fn}'
    res = Residence.query.get(r.residence_id)
    apt = Apartment.query.get(r.apartment_id) if r.apartment_id else None
    if res:
        syndic_notif = Notification(
            user_id=res.user_id,
            title=f'💰 طلب أداء من ساكن — {r.first_name} {r.last_name}',
            body=f'الإقامة: {res.name} | شقة: {apt.number if apt else "?"} | الطريقة: {method} | المرجع: {ref or "-"}',
            type='payment'
        )
        db.session.add(syndic_notif)
    notif = Notification(
        resident_id=r.id,
        title='✅ تم إرسال طلب الأداء',
        body=f'طلبك عبر {method} قيد المراجعة من طرف السانديك.',
        type='info'
    )
    db.session.add(notif)
    db.session.commit()
    return jsonify({'ok': True})


# ═══ إشعارات المدير ═══
@app.route('/api/admin/notifications')
@admin_required
def admin_get_notifications():
    result = []
    # 1) تسجيلات السانديك الجدد
    new_syndics = User.query.filter_by(role='syndic', status='pending').order_by(User.created_at.desc()).limit(10).all()
    for s in new_syndics:
        result.append({'id':f's{s.id}','title':f'📋 تسجيل سانديك جديد','body':f'{s.full_name} | {s.city or ""} | {s.phone or ""}','type':'register','read':False,'created_at':s.created_at.strftime('%d/%m/%Y %H:%M')})
    # 2) تذاكر الدعم المفتوحة
    tickets = SupportTicket.query.filter_by(status='open').order_by(SupportTicket.created_at.desc()).limit(10).all()
    for t in tickets:
        s = User.query.get(t.syndic_id)
        result.append({'id':f't{t.id}','title':f'📩 تذكرة دعم: {t.title}','body':f'{s.full_name if s else "?"} | {t.category}','type':'complaint','read':False,'created_at':t.created_at.strftime('%d/%m/%Y %H:%M')})
    # 3) طلبات تجديد الاشتراك (GlobalNotification type=payment)
    pays = GlobalNotification.query.filter(GlobalNotification.type.in_(['payment','alert'])).order_by(GlobalNotification.created_at.desc()).limit(10).all()
    for p in pays:
        result.append({'id':f'g{p.id}','title':p.title,'body':p.body or '','type':'payment','read':False,'created_at':p.created_at.strftime('%d/%m/%Y %H:%M')})
    # 4) السانديك المنتهية اشتراكاتهم
    now = datetime.utcnow()
    expiring = User.query.filter(User.role=='syndic', User.status=='active', User.subscription_end!=None, User.subscription_end < now + timedelta(days=7)).all()
    for s in expiring:
        days = (s.subscription_end - now).days if s.subscription_end else 0
        result.append({'id':f'e{s.id}','title':f'⚠️ اشتراك على وشك الانتهاء','body':f'{s.full_name} | باقي {max(0,days)} يوم','type':'alert','read':False,'created_at':now.strftime('%d/%m/%Y %H:%M')})
    # ترتيب حسب التاريخ
    try:
        result.sort(key=lambda x: datetime.strptime(x['created_at'], '%d/%m/%Y %H:%M'), reverse=True)
    except Exception:
        pass
    result = [r for r in result if not r.get('read', False)]
    return jsonify(result[:30])

@app.route('/api/admin/notifications/<int:nid>/read', methods=['POST'])
@admin_required
def admin_read_notification(nid):
    n = Notification.query.get_or_404(nid)
    n.read = True
    db.session.commit()
    return jsonify({'ok':True})


@app.route('/api/admin/notifications/<int:nid>', methods=['DELETE'])
@admin_required
def admin_delete_notif(nid):
    from models import SupportTicket, GlobalNotification
    # حذف GlobalNotification مباشرة
    g = GlobalNotification.query.get(nid)
    if g:
        db.session.delete(g)
        db.session.commit()
        return jsonify({'ok':True})
    t = SupportTicket.query.get(nid)
    if t:
        db.session.delete(t)
        db.session.commit()
        return jsonify({'ok':True})
    return jsonify({'ok':True})

@app.route('/api/admin/notifications/read-all', methods=['POST'])
@admin_required
def admin_read_all_notifications():
    Notification.query.filter_by(user_id=session['user_id'], read=False).update({'read':True})
    db.session.commit()
    return jsonify({'ok':True})


# ═══════════════════════════════
#  SUPPORT TICKETS
# ═══════════════════════════════
@app.route('/api/support', methods=['POST'])
@login_required
def create_ticket():
    u = get_current_user()
    data = request.get_json() or {}
    if not data.get('title','').strip():
        return jsonify({'error':'العنوان مطلوب'}),400
    t = SupportTicket(
        syndic_id=u.id,
        category=data.get('category','other'),
        title=data['title'].strip(),
        description=data.get('description','').strip()
    )
    db.session.add(t)
    db.session.commit()
    return jsonify({'ok':True,'id':t.id})

@app.route('/api/support', methods=['GET'])
@login_required
def get_my_tickets():
    u = get_current_user()
    tickets = SupportTicket.query.filter_by(syndic_id=u.id)                .order_by(SupportTicket.created_at.desc()).all()
    return jsonify([t.to_dict() for t in tickets])

@app.route('/api/admin/support', methods=['GET'])
@admin_required
def admin_get_tickets():
    status = request.args.get('status','')
    q = SupportTicket.query
    if status and status != 'all':
        q = q.filter_by(status=status)
    return jsonify([t.to_dict() for t in q.order_by(SupportTicket.created_at.desc()).all()])

@app.route('/api/admin/support/<int:tid>', methods=['PUT'])
@admin_required
def admin_reply_ticket(tid):
    t = SupportTicket.query.get_or_404(tid)
    data = request.get_json() or {}
    if 'status' in data: t.status = data['status']
    if 'admin_reply' in data:
        t.admin_reply = data['admin_reply']
        t.replied_at  = datetime.utcnow()
    db.session.commit()
    if data.get('admin_reply') and t.syndic_id:
        from models import Notification
        n = Notification(
            user_id=t.syndic_id,
            title='💬 رد المدير على تذكرتك',
            body=data['admin_reply'][:200],
            type='info'
        )
        db.session.add(n)
        db.session.commit()
    return jsonify({'ok':True})


@app.route('/api/syndic/change-password', methods=['POST'])
@login_required
def syndic_change_password():
    u = get_current_user()
    data = request.get_json() or {}
    old_pass = data.get('old_password','')
    new_pass = data.get('new_password','')
    if not check_password_hash(u.password_hash, old_pass):
        return jsonify({'error':'كلمة المرور الحالية غير صحيحة'}),400
    import re
    if len(new_pass) < 6:
        return jsonify({'error':'يجب أن تحتوي على 6 رموز على الأقل'}),400
    if not re.search(r'[a-zA-Z]', new_pass):
        return jsonify({'error':'يجب أن تحتوي على حرف واحد على الأقل'}),400
    if not re.search(r'[0-9]', new_pass):
        return jsonify({'error':'يجب أن تحتوي على رقم واحد على الأقل'}),400
    if not re.search(r'[^a-zA-Z0-9]', new_pass):
        return jsonify({'error':'يجب أن تحتوي على رمز خاص واحد على الأقل'}),400
    u.password_hash = generate_password_hash(new_pass)
    db.session.commit()
    return jsonify({'ok':True,'msg':'تم تغيير كلمة المرور بنجاح'})


# ═══════════════════════════════
#  SYNDIC NOTIFICATIONS — إشعارات السانديك
# ═══════════════════════════════

@app.route('/api/resident/emergency', methods=['POST'])
@resident_required
def resident_emergency():
    r = get_current_resident()
    data = request.get_json() or {}
    etype = data.get('type', 'طارئ')

    apt  = Apartment.query.get(r.apartment_id) if r.apartment_id else None
    res  = Residence.query.get(r.residence_id) if r.residence_id else None
    apt_num = apt.number if apt else '—'
    floor   = apt.floor   if apt else '—'
    res_name = res.name   if res else '—'

    # شكاية طارئة بأولوية عالية — تظهر فوراً في إشعارات السانديك
    comp = Complaint(
        apartment_id = r.apartment_id,
        title        = f'🚨 طارئ: {etype}',
        description  = f'تنبيه طارئ من الشقة {apt_num} (الطابق {floor}) — الإقامة: {res_name}',
        priority     = 'high',
        status       = 'open',
        date_created = datetime.utcnow()
    )
    db.session.add(comp)

    # إشعار تأكيد للقاطن
    notif = Notification(
        resident_id = r.id,
        title       = f'✅ تم إرسال تنبيه الطوارئ',
        body        = f'تم إبلاغ السانديك بـ «{etype}» — الشقة {apt_num}',
        type        = 'alert'
    )
    db.session.add(notif)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/syndic/notifications')
@login_required
def syndic_notifications():
    u = get_current_user()
    result = []

    # 1) إشعارات السكان من جدول Notification
    res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
    apt_ids = [a.id for a in Apartment.query.filter(Apartment.residence_id.in_(res_ids)).all()] if res_ids else []
    
    if apt_ids:
        residents_list = Resident.query.filter(Resident.apartment_id.in_(apt_ids), Resident.status=='approved').all()
        res_user_ids = [r.id for r in residents_list]
        
        # إشعارات من جدول Notification
        notifications = Notification.query.filter(
            Notification.resident_id.in_(res_user_ids)
        ).order_by(Notification.created_at.desc()).limit(20).all()
        
        for n in notifications:
            result.append({
                'id': f'n{n.id}',
                'title': n.title,
                'body': n.body,
                'type': n.type,
                'read': False,
                'source': 'resident',
                'created_at': n.created_at.strftime('%d/%m/%Y %H:%M') if n.created_at else '01/01/2026 00:00'
            })

    # 2) شكايات مفتوحة جديدة
    if apt_ids:
        complaints = Complaint.query.filter(
            Complaint.apartment_id.in_(apt_ids),
            Complaint.status == 'open'
        ).order_by(Complaint.date_created.desc()).limit(10).all()
        for c in complaints:
            apt = Apartment.query.get(c.apartment_id)
            result.append({
                'id': f'c{c.id}',
                'title': f'📋 شكاية جديدة — شقة {apt.number if apt else "?"}',
                'body': c.title,
                'type': 'complaint',
                'read': c.status != 'open',
                'source': 'resident',
                'created_at': c.date_created.strftime('%d/%m/%Y %H:%M')
            })

        # سكان جدد بانتظار القبول
        pending_residents = Resident.query.filter(
            Resident.residence_id.in_(res_ids),
            Resident.status == 'pending'
        ).order_by(Resident.created_at.desc()).limit(5).all()
        for r in pending_residents:
            result.append({
                'id': f'r{r.id}',
                'title': f'👤 طلب انضمام جديد',
                'body': f'{r.first_name} {r.last_name} | {r.phone or ""} | شقة {r.apt_number or "—"} ط{r.floor or "—"}',
                'type': 'register',
                'read': False,
                'source': 'resident',
                'created_at': r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else '01/01/2026 00:00'
            })

    # ترتيب حسب التاريخ
    from datetime import datetime
    try:
        result.sort(key=lambda x: datetime.strptime(x['created_at'], '%d/%m/%Y %H:%M'), reverse=True)
    except Exception:
        pass
    result = [r for r in result if not r.get('read', False)]
    return jsonify(result[:30])



@app.route('/api/syndic/notifications/<notif_id>/read', methods=['POST'])
@login_required
def syndic_notif_read(notif_id):
    # الإشعارات ليست في جدول خاص بالسانديك - نعيد ok فقط
    return jsonify({'ok': True})

@app.route('/api/syndic/notifications/read-all', methods=['POST'])
@login_required
def syndic_notif_read_all():
    return jsonify({'ok': True})

@app.route('/api/syndic/residents/pending')
@login_required
def syndic_pending_residents():
    u = get_current_user()
    if u.role != 'syndic':
        return jsonify({'error':'ghir musarrah'}), 403
    res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
    pending = Resident.query.filter(
        Resident.residence_id.in_(res_ids),
        Resident.status == 'pending'
    ).order_by(Resident.created_at.desc()).all()
    result = []
    for r in pending:
        apt = Apartment.query.get(r.apartment_id) if r.apartment_id else None
        result.append({
            'id': r.id,
            'full_name': f'{r.first_name} {r.last_name}',
            'phone': r.phone,
            'email': r.email,
            'cin': r.cin,
            'resident_type': r.resident_type,
            'residence_id': r.residence_id,
            'apartment_id': r.apartment_id,
            'apt_number': apt.number if apt else r.apt_number,
            'created_at': r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else ''
        })
    return jsonify(result)

@app.route('/api/syndic/residents/<int:rid>/approve', methods=['POST'])
@login_required
def syndic_approve_resident(rid):
    u = get_current_user()
    if u.role != 'syndic':
        return jsonify({'error':'ghir musarrah'}), 403
    r = Resident.query.get(rid)
    if not r:
        return jsonify({'error':'ghir mawjoud'}), 404
    res_ids = [x.id for x in Residence.query.filter_by(user_id=u.id).all()]
    if r.residence_id not in res_ids:
        return jsonify({'error':'ghir musarrah'}), 403

    data = request.get_json() or {}
    apt_id = data.get('apartment_id')
    if apt_id:
        apt = Apartment.query.get(int(apt_id))
        if not apt or apt.residence_id != r.residence_id:
            return jsonify({'error':'shaqqa ghayr sahiha'}), 400
        r.apartment_id = apt.id

    if not r.apartment_id:
        return jsonify({'error':'khass tahdid achaqqa'}), 400

    if 'tenant_name' in data:
        r.tenant_name = (data.get('tenant_name') or '').strip()
    if 'tenant_phone' in data:
        r.tenant_phone = (data.get('tenant_phone') or '').strip()

    # إلغاء أي ساكن مقبول سابق في نفس الشقة
    old_residents = Resident.query.filter_by(apartment_id=r.apartment_id, status='approved').all()
    for old_r in old_residents:
        if old_r.id != r.id:
            old_r.status = 'rejected'

    r.status = 'approved'
    db.session.commit()

    apt = Apartment.query.get(r.apartment_id)
    if apt:
        if r.resident_type == 'owner':
            apt.owner_name = f"{r.first_name} {r.last_name}".strip()
            apt.owner_phone = r.phone or ''
            if r.tenant_name:
                apt.tenant_name = r.tenant_name
                apt.tenant_phone = r.tenant_phone or ''
        else:
            apt.tenant_name = f"{r.first_name} {r.last_name}".strip()
            apt.tenant_phone = r.phone or ''
        db.session.commit()

    return jsonify({'ok': True})

@app.route('/api/syndic/residents/<int:rid>/reject', methods=['POST'])
@login_required
def syndic_reject_resident(rid):
    u = get_current_user()
    if u.role != 'syndic':
        return jsonify({'error':'ghir musarrah'}), 403
    r = Resident.query.get(rid)
    if not r:
        return jsonify({'error':'ghir mawjoud'}), 404
    res_ids = [x.id for x in Residence.query.filter_by(user_id=u.id).all()]
    if r.residence_id not in res_ids:
        return jsonify({'error':'ghir musarrah'}), 403
    r.status = 'rejected'
    db.session.commit()
    return jsonify({'ok': True})


@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'المورد غير موجود'}), 404
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'خطأ في الخادم'}), 500
    return render_template('500.html'), 500

@app.errorhandler(403)
def forbidden(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'غير مصرح'}), 403
    return render_template('403.html'), 403



# ════════════════════════════════════════
# ADMIN — طلبات الاشتراك المعلقة
# ════════════════════════════════════════

@app.route('/api/admin/pending-subscriptions')
def admin_pending_subscriptions():
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    user = User.query.get(session.get('user_id'))
    if not user or user.role != 'admin':
        return jsonify({'error': 'غير مصرح'}), 403
    
    try:
        from models import SubscriptionRequest
        result = []
        for req in pending:
            user_data = User.query.get(req.user_id)
            plan_data = SubscriptionPlan.query.get(req.plan_id)
            result.append({
                'id': req.id,
                'user_name': user_data.full_name if user_data else 'مجهول',
                'user_phone': user_data.phone if user_data else '',
                'plan_name': plan_data.label if plan_data else req.plan_id,
                'months': req.months,
                'base_price': req.base_price,
                'discount': req.discount_percent,
                'final_price': req.final_price,
                'payment_method': req.payment_method,
                'receipt_path': req.receipt_path,
                'created_at': req.requested_at.isoformat() if hasattr(req, 'requested_at') else ''
            })
        return jsonify({'count': len(result), 'data': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/subscription/<int:req_id>/approve', methods=['POST'])
def approve_subscription(req_id):
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    user = User.query.get(session.get('user_id'))
    if not user or user.role != 'admin':
        return jsonify({'error': 'غير مصرح'}), 403
    
    try:
        from models import SubscriptionRequest
        req = SubscriptionRequest.query.get(req_id)
        if not req:
            return jsonify({'error': 'الطلب غير موجود'}), 404
        
        req.status = 'approved'
        syndic = User.query.get(req.user_id)
        syndic.subscription_confirmed = True
        syndic.plan = SubscriptionPlan.query.get(req.plan_id).name
        
        db.session.commit()
        
        # إشعار للـ Syndic
        notif = GlobalNotification(
            title='✅ تم قبول طلبك الاشتراك!',
            body='تم الموافقة على طلب اشتراكك. يمكنك الآن استخدام جميع الميزات.',
            type='subscription',
            target='user',
            user_id=req.user_id
        )
        db.session.add(notif)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم القبول'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/subscription/<int:req_id>/reject', methods=['POST'])
def reject_subscription(req_id):
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    user = User.query.get(session.get('user_id'))
    if not user or user.role != 'admin':
        return jsonify({'error': 'غير مصرح'}), 403
    
    try:
        from models import SubscriptionRequest
        reason = request.form.get('reason', 'لم يتم تحديد السبب')
        req = SubscriptionRequest.query.get(req_id)
        if not req:
            return jsonify({'error': 'الطلب غير موجود'}), 404
        
        req.status = 'rejected'
        db.session.commit()
        
        # إشعار للـ Syndic
        notif = GlobalNotification(
            title='❌ تم رفض طلبك الاشتراك',
            body=f'للأسف تم رفض طلبك. السبب: {reason}',
            type='subscription',
            target='user',
            user_id=req.user_id
        )
        db.session.add(notif)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'تم الرفض'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=False)

# ═══════════════════════════════
#  RESIDENT AUTH & PORTAL
# ═══════════════════════════════
@app.route('/api/resident/register', methods=['POST'])
def resident_register():
    data = request.get_json() or {}
    required = ['cin','email','phone','password','residence_id']
    if not all(data.get(k,'').strip() for k in required):
        return jsonify({'error':'جميع الحقول مطلوبة'}),400
    if Resident.query.filter_by(cin=data['cin']).first():
        return jsonify({'error':'رقم CIN مستخدم مسبقاً'}),400
    if Resident.query.filter_by(email=data['email']).first():
        return jsonify({'error':'البريد الإلكتروني مستخدم مسبقاً'}),400
    if data.get('phone') and Resident.query.filter_by(phone=data['phone']).first():
        return jsonify({'error':'رقم الهاتف مستخدم مسبقاً'}),400
    r = Resident(
        first_name=data.get('first_name',''), last_name=data.get('last_name',''),
        cin=data['cin'], email=data['email'], phone=data['phone'],
        password_hash=generate_password_hash(data['password']),
        residence_id=int(data['residence_id']),
        apartment_id=int(data['apartment_id']) if data.get('apartment_id') else None,
        apt_number=data.get('apt_number',''),
        floor=int(data['floor']) if data.get('floor') else None,
        resident_type=data.get('resident_type','owner'),
        tenant_name=(data.get('tenant_name') or '').strip(),
        tenant_phone=(data.get('tenant_phone') or '').strip(),
        status='pending'
    )
    db.session.add(r)
    db.session.commit()

    # إشعار للسانديك بقاطن جديد
    try:
        res = Residence.query.get(int(data['residence_id']))
        if res:
            syndic = User.query.get(res.user_id)
            if syndic:
                apt_info = data.get('apt_number','—') or '—'
                floor_info = data.get('floor','—') or '—'
                phone_info = data.get('phone','—') or '—'
                snotif = Notification(
                    user_id  = syndic.id,
                    title    = f'👤 قاطن جديد — {data["first_name"]} {data["last_name"]}',
                    body     = f'شقة: {apt_info} | طابق: {floor_info} | هاتف: {phone_info} | الإقامة: {res.name}',
                    type     = 'register',
                    source   = 'resident'
                )
                db.session.add(snotif)
                db.session.commit()
    except Exception:
        pass

    return jsonify({'ok':True,'message':'تم إرسال طلبك، في انتظار الموافقة من السانديك'})

@app.route('/api/resident/login', methods=['POST'])
@rate_limit(max_calls=5, period=300)
def resident_login():
    data = request.get_json() or {}
    r = Resident.query.filter_by(email=data.get('email','')).first()
    if not r or not check_password_hash(r.password_hash, data.get('password','')):
        return jsonify({'error':'بيانات غير صحيحة'}),401
    if r.status != 'approved':
        msgs = {'pending':'حسابك قيد المراجعة','rejected':'تم رفض حسابك'}
        return jsonify({'error':msgs.get(r.status,'حساب غير نشط')}),403
    session['resident_id'] = r.id
    session.permanent = True
    session.pop('user_id', None)
    # Return full resident data so frontend can init without a second API call
    apt = Apartment.query.get(r.apartment_id) if r.apartment_id else None
    res = Residence.query.get(r.residence_id)
    return jsonify({
        'ok': True,
        'id': r.id,
        'full_name': f'{r.first_name} {r.last_name}',
        'email': r.email,
        'phone': r.phone,
        'cin': r.cin,
        'residence_name': res.name if res else '?',
        'residence_id': r.residence_id,
        'apt_number': apt.number if apt else r.apt_number or '—',
        'apartment_id': r.apartment_id,
        'floor': r.floor,
        'status': r.status,
    })

def get_current_resident():
    return Resident.query.get(session['resident_id'])

@app.route('/api/resident/me')
@resident_required
def resident_me():
    r = get_current_resident()
    apt = Apartment.query.get(r.apartment_id) if r.apartment_id else None
    res = Residence.query.get(r.residence_id)
    return jsonify({
        'id': r.id, 'full_name': f'{r.first_name} {r.last_name}',
        'email': r.email, 'phone': r.phone, 'cin': r.cin,
        'residence_name': res.name if res else '?',
        'residence_id': r.residence_id,
        'apt_number': apt.number if apt else r.apt_number or '—',
        'apartment_id': r.apartment_id,
        'floor': r.floor, 'status': r.status,
    })

@app.route('/api/resident/payments')
@resident_required
def resident_payments():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify([])
    pays = Payment.query.filter_by(apartment_id=r.apartment_id).order_by(Payment.year.desc(), Payment.month.desc()).all()
    return jsonify([p.to_dict() for p in pays])

@app.route('/api/resident/balance')
@resident_required
def resident_balance():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify({'paid':0,'unpaid':0,'total_paid':0,'total_due':0})
    apt = Apartment.query.get(r.apartment_id)
    pays = Payment.query.filter_by(apartment_id=r.apartment_id).all()
    paid_count = len([p for p in pays if p.status=='paid'])
    unpaid_count = len([p for p in pays if p.status=='unpaid'])
    total_paid = sum(p.amount for p in pays if p.status=='paid')
    return jsonify({
        'paid': paid_count, 'unpaid': unpaid_count,
        'total_paid': total_paid,
        'monthly_fee': apt.monthly_fee if apt else 0,
        'owner_name': apt.owner_name if apt else '',
    })

@app.route('/api/resident/complaints')
@resident_required
def resident_complaints():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify([])
    comps = Complaint.query.filter_by(apartment_id=r.apartment_id).order_by(Complaint.date_created.desc()).all()
    return jsonify([c.to_dict() for c in comps])

@app.route('/api/resident/complaints', methods=['POST'])
@resident_required
def resident_add_complaint():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify({'error':'لم يتم تعيين شقة لك بعد'}),400
    title = request.form.get('title','') or (request.get_json() or {}).get('title','')
    description = request.form.get('description','') or (request.get_json() or {}).get('description','')
    priority = request.form.get('priority','medium') or (request.get_json() or {}).get('priority','medium')
    photo_path = ''
    if 'photo' in request.files:
        f = request.files['photo']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            import time
            fn = f"{int(time.time())}_{secure_filename(f.filename)}"
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            photo_path = f'static/uploads/{fn}'
    c = Complaint(
        apartment_id=r.apartment_id,
        title=title, description=description,
        priority=priority, status='open',
        photo_path=photo_path
    )
    db.session.add(c)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/resident/notifications')
@resident_required
def resident_notifications():
    r = get_current_resident()
    if not r:
        return jsonify([])
    notifs = Notification.query.filter_by(resident_id=r.id).order_by(Notification.created_at.desc()).limit(20).all()
    globals_ = GlobalNotification.query.filter(
        GlobalNotification.target.in_(['all'])
    ).order_by(GlobalNotification.created_at.desc()).limit(5).all()
    result = [n.to_dict() for n in notifs]
    for g in globals_:
        result.append({'id':f'g{g.id}','title':g.title,'body':g.body or '','type':g.type,'read':False,'created_at':g.created_at.strftime('%d/%m/%Y %H:%M')})
    return jsonify(result)

@app.route('/api/resident/notifications/<int:nid>/read', methods=['POST'])
@resident_required
def resident_read_notif(nid):
    n = Notification.query.get_or_404(nid)
    n.read = True
    db.session.commit()
    return jsonify({'ok':True})



@app.route('/api/resident/syndic-contact')
@resident_required
def resident_syndic_contact():
    r = get_current_resident()
    res = Residence.query.get(r.residence_id)
    if not res:
        return jsonify({'error':'لا توجد إقامة'})
    u = User.query.get(res.user_id)
    if not u:
        return jsonify({'error':'لا يوجد سانديك'})
    return jsonify({
        'syndic_name': u.full_name or u.username,
        'syndic_phone': u.phone or '',
    })


@app.route('/api/resident/update-profile', methods=['PUT'])
@resident_required
def resident_update_profile():
    r = get_current_resident()
    data = request.get_json() or {}
    if 'phone' in data:
        r.phone = data['phone']
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/resident/change-password', methods=['POST'])
@resident_required
def resident_change_password():
    r = get_current_resident()
    data = request.get_json() or {}
    old_pass = data.get('old_password','')
    new_pass = data.get('new_password','')
    from werkzeug.security import check_password_hash, generate_password_hash
    import re
    if not check_password_hash(r.password_hash, old_pass):
        return jsonify({'error':'كلمة المرور الحالية غير صحيحة'}),400
    if len(new_pass) < 6:
        return jsonify({'error':'يجب أن تحتوي على 6 رموز على الأقل'}),400
    if not re.search(r'[a-zA-Z]', new_pass):
        return jsonify({'error':'يجب حرف واحد على الأقل'}),400
    if not re.search(r'[0-9]', new_pass):
        return jsonify({'error':'يجب رقم واحد على الأقل'}),400
    if not re.search(r'[^a-zA-Z0-9]', new_pass):
        return jsonify({'error':'يجب رمز خاص واحد على الأقل'}),400
    r.password_hash = generate_password_hash(new_pass)
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/resident/logout', methods=['POST'])
def resident_logout():
    session.pop('resident_id', None)
    return jsonify({'ok':True})

# ═══════════════════════════════
#  RESIDENCES LIST (public - for registration)
# ═══════════════════════════════

@app.route('/api/public/check-email', methods=['POST'])
def check_email():
    data = request.get_json() or {}
    email = data.get('email','').strip().lower()
    phone = data.get('phone','').strip()
    errors = {}
    if email:
        if User.query.filter_by(email=email).first():
            errors['email'] = 'البريد الإلكتروني مستخدم مسبقاً'
        elif Resident.query.filter_by(email=email).first():
            errors['email'] = 'البريد الإلكتروني مستخدم مسبقاً'
    if phone:
        if User.query.filter_by(phone=phone).first():
            errors['phone'] = 'رقم الهاتف مستخدم مسبقاً'
        elif Resident.query.filter_by(phone=phone).first():
            errors['phone'] = 'رقم الهاتف مستخدم مسبقاً'
    return jsonify({'ok': len(errors)==0, 'errors': errors})

@app.route('/api/public/residences')
def public_residences():
    res = Residence.query.all()
    return jsonify([{'id':r.id,'name':r.name,'city':r.city or ''} for r in res])

# ═══════════════════════════════
#  ADMIN - send notification to resident
# ═══════════════════════════════
@app.route('/api/syndic/notify-resident', methods=['POST'])
@login_required
def notify_resident():
    data = request.get_json() or {}
    n = Notification(
        resident_id=int(data['resident_id']),
        title=data.get('title','إشعار'),
        body=data.get('body',''),
        type=data.get('type','info')
    )
    db.session.add(n)
    db.session.commit()
    return jsonify({'ok':True})

# ═══════════════════════════════
#  نظام الأداء الإلكتروني
# ═══════════════════════════════
import hashlib, qrcode, io, base64
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from flask import send_file, make_response

def gen_receipt_ref(payment_id):
    return f"SP-{datetime.utcnow().year}-{payment_id:06d}"

def gen_qr_base64(data):
    qr = qrcode.QRCode(version=1, box_size=6, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1e40af", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()

# QR Code لشقة معينة
@app.route('/api/payments/qr/<int:apt_id>')
@login_required
def payment_qr(apt_id):
    apt = Apartment.query.get_or_404(apt_id)
    res = Residence.query.get(apt.residence_id)
    data = f"SYNDIKPRO|APT:{apt_id}|RES:{apt.residence_id}|FEE:{apt.monthly_fee}|{res.name if res else ''}"
    qr_b64 = gen_qr_base64(data)
    return jsonify({'qr': qr_b64, 'data': data})

# وصل الدفع PDF
@app.route('/api/payments/<int:pid>/receipt')
@login_required
def payment_receipt(pid):
    pay = Payment.query.get_or_404(pid)
    apt = Apartment.query.get(pay.apartment_id)
    res = Residence.query.get(apt.residence_id) if apt else None
    syndic = User.query.get(res.user_id) if res else None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    ar = ParagraphStyle('ar', fontName='Helvetica', fontSize=11,
                        alignment=1, spaceAfter=6, leading=18)
    title_style = ParagraphStyle('title', fontName='Helvetica-Bold',
                                 fontSize=16, alignment=1, spaceAfter=12)
    small = ParagraphStyle('small', fontName='Helvetica',
                           fontSize=9, alignment=1, textColor=colors.grey)

    MONTHS_FR = ['Janvier','Février','Mars','Avril','Mai','Juin',
                 'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

    ref = gen_receipt_ref(pay.id)
    qr_b64 = gen_qr_base64(f"RECEIPT|{ref}|{pay.amount}|{pay.status}")
    qr_img_data = base64.b64decode(qr_b64)
    qr_buf = io.BytesIO(qr_img_data)

    story = []

    # Header
    story.append(Paragraph("SyndikPro", title_style))
    story.append(Paragraph("Reçu de Paiement / وصل الأداء", ar))
    story.append(Spacer(1, 0.4*cm))

    # Info table
    month_name = MONTHS_FR[(pay.month or 1)-1] if pay.month else '?'
    data_table = [
        ['Référence / المرجع', ref],
        ['Résidence / الإقامة', res.name if res else '?'],
        ['Appartement / الشقة', f"N° {apt.number}" if apt else '?'],
        ['Propriétaire / المالك', apt.owner_name if apt else '?'],
        ['Mois / الشهر', f"{month_name} {pay.year}"],
        ['Montant / المبلغ', f"{pay.amount:.2f} MAD"],
        ['Mode de paiement / طريقة الأداء', pay.method or 'نقدي'],
        ['Date / التاريخ', pay.date_paid.strftime('%d/%m/%Y') if pay.date_paid else '?'],
        ['Statut / الحالة', '✓ Payé / مدفوع' if pay.status=='paid' else pay.status],
    ]

    t = Table(data_table, colWidths=[7*cm, 10*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#eff6ff')),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#dbeafe')),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.6*cm))

    # QR Code
    qr_img = Image(qr_buf, width=3*cm, height=3*cm)
    story.append(Paragraph("QR Code de vérification / رمز التحقق", small))
    story.append(qr_img)
    story.append(Spacer(1, 0.4*cm))

    # Footer
    story.append(Paragraph(f"Syndic: {syndic.full_name if syndic else '?'} | {syndic.phone if syndic else ''}", small))
    story.append(Paragraph(f"Généré par SyndikPro le {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}", small))
    story.append(Paragraph("Ce document est un reçu officiel de paiement.", small))

    doc.build(story)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename="recu_{ref}.pdf"'
    return response

# وصل جماعي — كل مدفوعات شهر
@app.route('/api/residences/<int:rid>/receipt-all')
@login_required
def residence_receipt_all(rid):
    month = int(request.args.get('month', datetime.utcnow().month))
    year  = int(request.args.get('year',  datetime.utcnow().year))
    res   = Residence.query.get_or_404(rid)
    apts  = Apartment.query.filter_by(residence_id=rid).all()

    MONTHS_FR = ['Janvier','Février','Mars','Avril','Mai','Juin',
                 'Juillet','Août','Septembre','Octobre','Novembre','Décembre']

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=14, alignment=1, spaceAfter=8)
    sub_s   = ParagraphStyle('s', fontName='Helvetica', fontSize=10, alignment=1, spaceAfter=4)
    small   = ParagraphStyle('sm', fontName='Helvetica', fontSize=8, textColor=colors.grey, alignment=1)

    story = []
    story.append(Paragraph("SyndikPro — Rapport Mensuel", title_s))
    story.append(Paragraph(f"{res.name} | {MONTHS_FR[month-1]} {year}", sub_s))
    story.append(Spacer(1, 0.4*cm))

    rows = [['N° Appt', 'Propriétaire', 'Montant (MAD)', 'Statut', 'Date', 'Méthode']]
    total_paid = 0
    paid_count = 0

    for a in apts:
        pay = Payment.query.filter_by(apartment_id=a.id, month=month, year=year).first()
        status = '✓ Payé' if pay and pay.status=='paid' else '✗ Impayé'
        amount = f"{pay.amount:.2f}" if pay and pay.status=='paid' else f"{a.monthly_fee:.2f}"
        date   = pay.date_paid.strftime('%d/%m/%Y') if pay and pay.date_paid else '—'
        method = pay.method if pay else '—'
        if pay and pay.status=='paid':
            total_paid += pay.amount
            paid_count += 1
        rows.append([a.number, a.owner_name or '—', amount, status, date, method])

    rows.append(['', 'TOTAL PAYÉ', f"{total_paid:.2f}", f"{paid_count}/{len(apts)}", '', ''])

    t = Table(rows, colWidths=[2*cm, 5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f8fafc')]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#dbeafe')),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph(f"Généré par SyndikPro le {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}", small))

    doc.build(story)
    buf.seek(0)

    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="rapport_{res.name}_{month}_{year}.pdf"'
    return response

# إحصائيات الأداء لإقامة
@app.route('/api/residences/<int:rid>/financial-summary')
@login_required
def financial_summary(rid):
    month = request.args.get('month')
    year  = int(request.args.get('year', datetime.utcnow().year))
    apts  = Apartment.query.filter_by(residence_id=rid).all()
    apt_ids = [a.id for a in apts]
    q = Payment.query.filter(Payment.apartment_id.in_(apt_ids), Payment.year==year, Payment.status=='paid')
    if month:
        q = q.filter(Payment.month==int(month))
    income = db.session.query(db.func.sum(Payment.amount)).filter(
        Payment.apartment_id.in_(apt_ids), Payment.year==year, Payment.status=='paid'
    )
    if month:
        income = income.filter(Payment.month==int(month))
    income = income.scalar() or 0
    eq = Expense.query.filter(Expense.residence_id==rid, db.extract('year', Expense.date)==year)
    if month:
        eq = eq.filter(db.extract('month', Expense.date)==int(month))
    expenses = sum(e.amount for e in eq.all())
    return jsonify({'income': income, 'expenses': expenses, 'net': income - expenses})

@app.route('/api/residences/<int:rid>/payment-stats')
@login_required
def payment_stats(rid):
    apts = Apartment.query.filter_by(residence_id=rid).all()
    result = []
    for i in range(6):
        d = datetime.utcnow().replace(day=1) - timedelta(days=i*30)
        m, y = d.month, d.year
        paid = 0
        for a in apts:
            p = Payment.query.filter_by(apartment_id=a.id, month=m, year=y, status='paid').first()
            if p: paid += p.amount
        result.append({'month': f"{m}/{y}", 'amount': paid})
    return jsonify(list(reversed(result)))


# ══════════════════════════════════════════════════════
#  PASSWORD RESET — استرداد كلمة المرور
# ══════════════════════════════════════════════════════
import random, string
from models import PasswordResetToken

# Rate limiting لطلبات استرداد كلمة المرور
from collections import defaultdict
import time as _time
_forgot_attempts = defaultdict(list)

def _check_rate_limit(ip, max_attempts=3, period=3600):
    now = _time.time()
    attempts = [t for t in _forgot_attempts[ip] if now - t < window]
    _forgot_attempts[ip] = attempts
    if len(attempts) >= max_attempts:
        return False
    _forgot_attempts[ip].append(now)
    return True

GMAIL_USER = os.environ.get("GMAIL_USER", "")
GMAIL_PASS = os.environ.get("GMAIL_PASS", "")

def send_email(destination, subject, body_text):
    """دالة لإرسال رسائل التحقق الإلكترونية"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    try:
        if not GMAIL_USER or not GMAIL_PASS:
            print("⚠️  بيانات Gmail لم تُعيّن")
            return False
            
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = GMAIL_USER
        msg['To'] = destination

        html_body = f"""
        <div dir="rtl" style="font-family:Arial;max-width:480px;margin:auto;padding:24px;border:1px solid #e2e8f0;border-radius:12px">
          <h2 style="color:#1e40af">🏢 SyndikPro</h2>
          <p>مرحباً،</p>
          <p><strong>{subject}</strong></p>
          <div style="font-size:28px;font-weight:900;letter-spacing:8px;color:#1e40af;text-align:center;padding:20px;background:#eff6ff;border-radius:8px;margin:20px 0">
            {body_text}
          </div>
          <p style="color:#64748b;font-size:12px">⏱️ صالح لمدة 10 دقائق فقط</p>
          <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0">
          <p style="color:#94a3b8;font-size:11px;text-align:center">SyndikPro — نظام تدبير الإقامات السكنية</p>
        </div>
        """
        
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.sendmail(GMAIL_USER, destination, msg.as_string())
        
        print(f"✅ رسالة أُرسلت إلى {destination}")
        return True
        
    except smtplib.SMTPAuthenticationError:
        print(f"❌ خطأ في المصادقة: تحقق من GMAIL_USER و GMAIL_PASS")
        return False
    except Exception as e:
        print(f"❌ خطأ في الإرسال: {e}")
        return False


def send_otp_console(method, destination, code, user_type):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    print("═"*55 + "\n")

    if method == 'email' and user_type == 'admin_alert':
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = "🔑 SyndikPro — طلب استرداد كلمة مرور سانديك"
            msg['From']    = GMAIL_USER
            msg['To']      = destination

            body = f"""
            <div dir="rtl" style="font-family:Arial;max-width:480px;margin:auto;padding:24px;border:1px solid #e2e8f0;border-radius:12px">
              <h2 style="color:#dc2626">🔑 SyndikPro — تنبيه للمدير</h2>
              <p>أحد السانديك طلب إعادة تعيين كلمة المرور.</p>
              <p>المرجع: <strong>{code}</strong></p>
              <p>يرجى مراجعة لوحة التحكم والرد على التذكرة.</p>
              <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0">
              <p style="color:#94a3b8;font-size:11px;text-align:center">SyndikPro — نظام تدبير الإقامات السكنية</p>
            </div>
            """
            msg.attach(MIMEText(body, 'html', 'utf-8'))

            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(GMAIL_USER, GMAIL_PASS)
                smtp.sendmail(GMAIL_USER, destination, msg.as_string())
            print(f"  ✅ إيميل تنبيه أُرسل للمدير على {destination}")
        except Exception as e:
            print(f"  ❌ خطأ في إرسال التنبيه: {e}")
        return

    if method == 'email':
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = "🔑 SyndikPro — كود استرداد كلمة المرور"
            msg['From']    = GMAIL_USER
            msg['To']      = destination

            body = f"""
            <div dir="rtl" style="font-family:Arial;max-width:480px;margin:auto;padding:24px;border:1px solid #e2e8f0;border-radius:12px">
              <h2 style="color:#1e40af">🏢 SyndikPro</h2>
              <p>مرحباً،</p>
              <p>هذا هو كود استرداد كلمة المرور الخاص بك:</p>
              <div style="font-size:36px;font-weight:900;letter-spacing:12px;color:#1e40af;text-align:center;padding:20px;background:#eff6ff;border-radius:8px;margin:20px 0">
                {code}
              </div>
              <p style="color:#64748b;font-size:12px">⏱️ صالح لمدة 10 دقائق فقط</p>
              <p style="color:#64748b;font-size:12px">إذا لم تطلب هذا الكود، تجاهل هذا الإيميل.</p>
              <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0">
              <p style="color:#94a3b8;font-size:11px;text-align:center">SyndikPro — نظام تدبير الإقامات السكنية</p>
            </div>
            """
            msg.attach(MIMEText(body, 'html', 'utf-8'))

            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(GMAIL_USER, GMAIL_PASS)
                smtp.sendmail(GMAIL_USER, destination, msg.as_string())
            print(f"  ✅ إيميل أُرسل إلى {destination}")
        except Exception as e:
            print(f"  ❌ خطأ في الإرسال: {e}")
    else:
        pass
def generate_otp(length=6):
    return ''.join(random.choices(string.digits, k=length))


# ── 1. القاطن: طلب OTP بالإيميل ──────────────────────
@app.route('/api/resident/forgot-password', methods=['POST'])
def resident_forgot_password():
    if not _check_rate_limit(request.remote_addr):
        return jsonify({'error': 'تجاوزت الحد المسموح به، حاول بعد ساعة'}), 429
    data  = request.get_json() or {}
    email = data.get('email','').strip().lower()
    if not email:
        return jsonify({'error':'أدخل البريد الإلكتروني'}), 400

    r = Resident.query.filter_by(email=email).first()
    if not r:
        return jsonify({"ok": True, "msg": "إذا كان الإيميل مسجلاً سيصلك الكود"})
        return jsonify({'ok':True, 'msg':'إذا كان الإيميل مسجلاً سيصلك الكود'})

    # إلغاء الأكواد القديمة
    PasswordResetToken.query.filter_by(user_type='resident', user_id=r.id, used=False).update({'used':True})
    db.session.commit()

    code = generate_otp(6)
    tok  = PasswordResetToken(
        user_type='resident', user_id=r.id,
        token=code, method='email',
        expires_at=datetime.utcnow() + timedelta(minutes=10)
    )
    db.session.add(tok)
    db.session.commit()

    send_otp_console('email', email, code, 'resident')
    return jsonify({'ok':True, 'msg':'إذا كان الإيميل مسجلاً سيصلك الكود'})


# ── 2. القاطن: تأكيد OTP وتغيير كلمة المرور ──────────
@app.route('/api/resident/reset-password', methods=['POST'])
def resident_reset_password():
    data     = request.get_json() or {}
    email    = data.get('email','').strip().lower()
    code     = data.get('code','').strip()
    new_pass = data.get('new_password','')

    if not all([email, code, new_pass]):
        return jsonify({'error':'جميع الحقول مطلوبة'}), 400
    if len(new_pass) < 6:
        return jsonify({'error':'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400

    r = Resident.query.filter_by(email=email).first()
    if not r:
        return jsonify({'error':'بيانات غير صحيحة'}), 400

    tok = PasswordResetToken.query.filter_by(
        user_type='resident', user_id=r.id, token=code, used=False
    ).first()

    if not tok or tok.expires_at < datetime.utcnow():
        return jsonify({'error':'الكود غير صالح أو منتهي الصلاحية'}), 400

    tok.used = True
    r.password_hash = generate_password_hash(new_pass)
    db.session.commit()

    print(f"\n  ✅ القاطن [{r.first_name} {r.last_name}] غيّر كلمة المرور بنجاح\n")
    return jsonify({'ok':True, 'msg':'تم تغيير كلمة المرور بنجاح'})


# ── 3. السانديك: طلب استرداد بالهاتف ─────────────────
@app.route('/api/syndic/forgot-password/phone', methods=['POST'])
def syndic_forgot_phone():
    if not _check_rate_limit(request.remote_addr):
        return jsonify({'error': 'تجاوزت الحد المسموح به، حاول بعد ساعة'}), 429
    data  = request.get_json() or {}
    phone = data.get('phone','').strip()
    if not phone:
        return jsonify({'error':'أدخل رقم الهاتف'}), 400

    u = User.query.filter_by(phone=phone, role='syndic').first()
    if not u:
        return jsonify({"ok": True, "msg": "إذا كان الهاتف مسجلاً سيصلك الكود"})

    if u.status != 'active':
        return jsonify({'error':'الحساب غير نشط، تواصل مع المدير'}), 403

    PasswordResetToken.query.filter_by(user_type='syndic', user_id=u.id, used=False).update({'used':True})
    db.session.commit()

    code = generate_otp(8)
    tok  = PasswordResetToken(
        user_type='syndic', user_id=u.id,
        token=code, method='phone',
        expires_at=datetime.utcnow() + timedelta(minutes=10)
    )
    db.session.add(tok)
    db.session.commit()

    send_otp_console('phone', phone, code, 'syndic')
    return jsonify({'ok':True, 'msg':'إذا كان الرقم مسجلاً سيصلك الكود'})


# ── 4. السانديك: تأكيد OTP هاتف وتغيير كلمة المرور ──
@app.route('/api/syndic/reset-password/phone', methods=['POST'])
def syndic_reset_phone():
    data     = request.get_json() or {}
    phone    = data.get('phone','').strip()
    code     = data.get('code','').strip()
    new_pass = data.get('new_password','')

    if not all([phone, code, new_pass]):
        return jsonify({'error':'جميع الحقول مطلوبة'}), 400
    if len(new_pass) < 6:
        return jsonify({'error':'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400

    u = User.query.filter_by(phone=phone, role='syndic').first()
    if not u:
        return jsonify({'error':'بيانات غير صحيحة'}), 400

    tok = PasswordResetToken.query.filter_by(
        user_type='syndic', user_id=u.id, token=code, method='phone', used=False
    ).first()

    if not tok or tok.expires_at < datetime.utcnow():
        return jsonify({'error':'الكود غير صالح أو منتهي الصلاحية'}), 400

    tok.used = True
    u.password_hash = generate_password_hash(new_pass)
    db.session.commit()

    print(f"\n  ✅ السانديك [{u.full_name}] غيّر كلمة المرور عبر الهاتف\n")
    return jsonify({'ok':True, 'msg':'تم تغيير كلمة المرور بنجاح'})


# ── 5. السانديك: طلب استرداد عبر المدير (إيميل) ──────
@app.route('/api/syndic/forgot-password/admin', methods=['POST'])
def syndic_forgot_admin():
    if not _check_rate_limit(request.remote_addr):
        return jsonify({'error': 'تجاوزت الحد المسموح به، حاول بعد ساعة'}), 429
    data  = request.get_json() or {}
    email = data.get('email','').strip().lower()
    if not email:
        return jsonify({'error':'أدخل البريد الإلكتروني'}), 400

    u = User.query.filter_by(email=email, role='syndic').first()
    if not u:
        return jsonify({"ok": True, "msg": "إذا كان الهاتف مسجلاً سيصلك الكود"})

    # تذكرة دعم تلقائية
    ticket = SupportTicket(
        syndic_id=u.id,
        category='technical',
        title=f'🔑 طلب إعادة تعيين كلمة المرور — {u.username}',
        description=f'السانديك {u.full_name} (إيميل: {email}) طلب إعادة تعيين كلمة المرور.',
        status='open'
    )
    db.session.add(ticket)

    # إشعار فوري للمدير
    gn = GlobalNotification(
        title=f'🔑 طلب استرداد كلمة المرور — {u.full_name}',
        body=f'السانديك {u.full_name} ({email}) طلب إعادة تعيين كلمة المرور.',
        type='alert',
        target='admin'
    )
    db.session.add(gn)
    db.session.commit()

    # إيميل للمدير
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"🔑 SyndikPro — طلب استرداد كلمة المرور من {u.full_name}"
        msg['From'] = GMAIL_USER
        msg['To']   = GMAIL_USER
        body = f"""<div dir="rtl" style="font-family:Arial;max-width:480px;margin:auto;padding:24px;border:1px solid #e2e8f0;border-radius:12px">
          <h2 style="color:#dc2626">🔑 طلب استرداد كلمة المرور</h2>
          <p>السانديك <strong>{u.full_name}</strong> طلب إعادة تعيين كلمة المرور.</p>
          <p>البريد: {email}</p>
          <p>الهاتف: {u.phone or '—'}</p>
          <p>رقم التذكرة: <strong>#{ticket.id}</strong></p>
          <p>يرجى الدخول للوحة التحكم والرد على التذكرة.</p>
        </div>"""
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.sendmail(GMAIL_USER, GMAIL_USER, msg.as_string())
        print("  ✅ إيميل أُرسل للمدير")
    except Exception as e:
        print(f"  ❌ خطأ: {e}")

    return jsonify({'ok':True, 'msg':'تم إرسال طلبك للمدير، سيتواصل معك قريباً'})


@app.route('/api/admin/syndics/<int:sid>/reset-password', methods=['POST'])
@admin_required
def admin_reset_syndic_password(sid):
    data     = request.get_json() or {}
    new_pass = data.get('new_password','')
    if not new_pass or len(new_pass) < 6:
        return jsonify({'error':'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400

    u = User.query.get_or_404(sid)
    if u.role == 'admin':
        return jsonify({'error':'لا يمكن تعديل حساب المدير'}), 403

    u.password_hash = generate_password_hash(new_pass)
    db.session.commit()

    admin = get_current_user()
    print(f"\n  🛡️  المدير [{admin.full_name}] أعاد تعيين كلمة مرور السانديك [{u.full_name}]\n")
    return jsonify({'ok':True, 'msg':f'تم إعادة تعيين كلمة مرور {u.full_name}'})

# عدد طلبات استرداد كلمة المرور المعلقة
@app.route('/api/admin/reset-requests-count')
@admin_required
def reset_requests_count():
    count = SupportTicket.query.filter(
        SupportTicket.title.like('%إعادة تعيين كلمة المرور%'),
        SupportTicket.status == 'open'
    ).count()
    return jsonify({'count': count})

@app.route('/api/syndic/forgot-username', methods=['POST'])
def syndic_forgot_username():
    if not _check_rate_limit(request.remote_addr):
        return jsonify({"error": "تجاوزت الحد المسموح به، حاول بعد ساعة"}), 429
    data  = request.get_json() or {}
    email = data.get('email','').strip().lower()
    phone = data.get('phone','').strip()

    # تحقق من صحة البريد
    import re
    if email and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        return jsonify({"error": "ادخل بريد الكتروني صحيح"}), 400

    u = None
    if email:
        u = User.query.filter_by(email=email, role='syndic').first()
    if not u and phone:
        u = User.query.filter_by(phone=phone, role='syndic').first()

    if not u:
        return jsonify({"ok": True, "msg": "إذا كان البريد أو الهاتف مسجلاً سيصلك الكود"})

    ticket = SupportTicket(
        syndic_id=u.id,
        category='technical',
        title=f'طلب استرداد اسم المستخدم - {u.full_name}',
        description=f'السانديك {u.full_name} طلب استرداد اسم المستخدم.',
        status='open'
    )
    db.session.add(ticket)

    gn = GlobalNotification(
        title=f'طلب استرداد اسم المستخدم - {u.full_name}',
        body=f'البريد: {u.email} | الهاتف: {u.phone or "-"}',
        type='alert',
        target='admin'
    )
    db.session.add(gn)
    db.session.commit()

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = "SyndikPro - طلب استرداد اسم المستخدم"
        msg['From'] = GMAIL_USER
        msg['To']   = GMAIL_USER
        body = (
            '<div dir="rtl" style="font-family:Arial;padding:20px">'
            '<h2>طلب استرداد اسم المستخدم</h2>'
            '<p>السانديك: <strong>' + u.full_name + '</strong></p>'
            '<p>البريد: ' + u.email + '</p>'
            '<p>الهاتف: ' + (u.phone or '-') + '</p>'
            '<p>التذكرة: #' + str(ticket.id) + '</p>'
            '</div>'
        )
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.sendmail(GMAIL_USER, GMAIL_USER, msg.as_string())
    except Exception as e:
        print(f"email error: {e}")

    return jsonify({'ok': True, 'msg': 'تم ارسال الطلب للمدير، سيتواصل معك قريبا'})

# ———— السانديك: نسيت كلمة المرور عبر البريد ————
@app.route('/api/syndic/forgot-password/email', methods=['POST'])
def syndic_forgot_password_email():
    if not _check_rate_limit(request.remote_addr):
        return jsonify({"error": "تجاوزت الحد المسموح به، حاول بعد ساعة"}), 429
    data  = request.get_json() or {}
    email = data.get('email','').strip().lower()
    if not email:
        return jsonify({'error':'أدخل البريد الإلكتروني'}), 400
    u = User.query.filter_by(email=email, role='syndic').first()
    if not u:
        return jsonify({"ok": True, "msg": "إذا كان البريد مسجلاً سيصلك الكود"})
    PasswordResetToken.query.filter_by(user_type='syndic', user_id=u.id, used=False).update({'used':True})
    db.session.commit()
    code = generate_otp(6)
    tok  = PasswordResetToken(
        user_type='syndic', user_id=u.id,
        token=code, method='email',
        expires_at=datetime.utcnow() + timedelta(minutes=10)
    )
    db.session.add(tok)
    db.session.commit()
    send_otp_console('email', email, code, 'syndic')
    return jsonify({'ok':True, 'msg':'إذا كان الإيميل مسجلاً سيصلك الكود'})

@app.route('/api/syndic/reset-password/email', methods=['POST'])
def syndic_reset_password_email():
    data     = request.get_json() or {}
    email    = data.get('email','').strip().lower()
    code     = data.get('code','').strip()
    new_pass = data.get('new_password','')
    if not all([email, code, new_pass]):
        return jsonify({'error':'جميع الحقول مطلوبة'}), 400
    if len(new_pass) < 6:
        return jsonify({'error':'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400
    u = User.query.filter_by(email=email, role='syndic').first()
    if not u:
        return jsonify({'error':'بيانات غير صحيحة'}), 400
    tok = PasswordResetToken.query.filter_by(
        user_type='syndic', user_id=u.id, token=code, used=False
    ).first()
    if not tok or tok.expires_at < datetime.utcnow():
        return jsonify({'error':'الكود غير صالح أو منتهي الصلاحية'}), 400
    tok.used = True
    u.password_hash = generate_password_hash(new_pass)
    db.session.commit()
    return jsonify({'ok':True, 'msg':'تم تغيير كلمة المرور بنجاح'})

@app.route('/api/admin/export_revenue_xlsx', methods=['POST'])
@admin_required
def export_revenue_xlsx():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file
    data = request.get_json()
    month = data.get('month', '')
    year = data.get('year', '')
    # جلب البيانات مباشرة من DB
    from datetime import datetime
    month_names = ['يناير','فبراير','مارس','أبريل','ماي','يونيو','يوليو','غشت','شتنبر','أكتوبر','نونبر','دجنبر']
    month_num = month_names.index(month)+1 if month in month_names else datetime.utcnow().month
    year_num = int(year) if year else datetime.utcnow().year
    syndics = User.query.filter_by(role='syndic', status='active').all()
    details = []
    for s in syndics:
        if s.approved_at and s.approved_at.month == month_num and s.approved_at.year == year_num:
            res = Residence.query.filter_by(user_id=s.id).first()
            details.append({
                'syndic_name': s.full_name or s.username,
                'residence': res.name if res else '',
                'city': s.city or '',
                'syndic_phone': s.phone or '',
                'payments_count': 0,
                'total': s.total_amount or 0,
            })
    wb = Workbook()
    ws = wb.active
    ws.title = 'الإيرادات'
    ws.sheet_view.rightToLeft = True
    headers = ['اسم السانديك','الإقامة','المدينة','الهاتف','عدد الدفعات','الإجمالي (درهم)','الشهر','السنة']
    header_fill = PatternFill(start_color='6C3FC7', end_color='6C3FC7', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for row, d in enumerate(details, 2):
        ws.cell(row=row, column=1, value=d.get('syndic_name',''))
        ws.cell(row=row, column=2, value=d.get('residence',''))
        ws.cell(row=row, column=3, value=d.get('city',''))
        ws.cell(row=row, column=4, value=d.get('syndic_phone',''))
        ws.cell(row=row, column=5, value=d.get('payments_count',0))
        ws.cell(row=row, column=6, value=d.get('total',0))
        ws.cell(row=row, column=7, value=month)
        ws.cell(row=row, column=8, value=year)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'إيرادات_{month}_{year}.xlsx')

@app.route('/api/syndic/check-subscription-pending')
@login_required
def check_subscription_pending():
    from models import SubscriptionRequest
    user = get_current_user()
    if not user or user.role != 'syndic':
        return jsonify({'error': 'Unauthorized'}), 403
    
    pending = SubscriptionRequest.query.filter_by(user_id=user.id, status='pending').first()
    return jsonify({'has_pending': bool(pending)})

@app.route('/api/syndic/plan-limits')
@login_required
def syndic_plan_limits():
    u = get_current_user()
    plan_obj = SubscriptionPlan.query.filter_by(name=u.plan).first()
    current_res = Residence.query.filter_by(user_id=u.id).count()
    res_ids = [r.id for r in Residence.query.filter_by(user_id=u.id).all()]
    current_apts = Apartment.query.filter(Apartment.residence_id.in_(res_ids)).count() if res_ids else 0
    max_res  = plan_obj.max_residences if plan_obj else 1
    max_apts = plan_obj.max_apartments if plan_obj else 20
    return jsonify({
        'plan': u.plan,
        'plan_label': plan_obj.label if plan_obj else u.plan,
        'max_residences': max_res,
        'max_apartments': max_apts,
        'current_residences': current_res,
        'current_apartments': current_apts,
        'can_add_residence': current_res < max_res,
        'can_add_apartment': current_apts < max_apts,
    })

@app.route('/api/public/plans')
def public_plans():
    from models import SubscriptionPlan
    plans = SubscriptionPlan.query.filter_by(is_active=True).order_by(SubscriptionPlan.id).all()
    return jsonify([p.to_dict() for p in plans])

# ══════════════════════════════
#  SECURITY MONITORING
# ══════════════════════════════
@app.route('/api/admin/security-log')
@admin_required
def admin_security_log():
    from security import get_security_log
    return jsonify(get_security_log())

@app.route('/api/admin/blocked-ips')
@admin_required
def admin_blocked_ips():
    from security import _attempts
    import time
    now = time.time()
    blocked = {ip: len(attempts) for ip, attempts in _attempts.items()
               if len(attempts) >= 5}
    return jsonify(blocked)


# ═══════════════════════════════
#  PUBLIC APIs — تسجيل القاطن التدريجي
# ═══════════════════════════════

@app.route("/api/public/cities")
def public_cities():
    from sqlalchemy import distinct
    from morocco_cities import MOROCCO_CITIES
    
    # احصل على قائمة المدن من MOROCCO_CITIES
    all_cities = list(MOROCCO_CITIES.keys())
    
    # أضف أي مدن من قاعدة البيانات لم تكن في القائمة
    db_cities = db.session.query(distinct(Residence.city)).filter(
        Residence.city != None, Residence.city != ""
    ).all()
    
    for c in db_cities:
        if c[0] and c[0] not in all_cities:
            all_cities.append(c[0])
    
    from flask import Response
    import json
    return Response(json.dumps(sorted(all_cities), ensure_ascii=False), mimetype='application/json')

@app.route("/api/public/neighborhoods")
def public_neighborhoods():
    city = request.args.get("city","")
    from sqlalchemy import distinct
    hoods = db.session.query(distinct(Residence.neighborhood)).filter(
        Residence.city == city,
        Residence.neighborhood != None,
        Residence.neighborhood != ""
    ).all()
    result = [h[0] for h in hoods if h[0]]
    if not result:
        try:
            from morocco_cities import MOROCCO_CITIES
            result = MOROCCO_CITIES.get(city, [])
        except:
            pass
    return jsonify(result if result else ['_all_'])

@app.route("/api/public/residences-by-hood")
def public_residences_by_hood():
    city = request.args.get("city","")
    hood = request.args.get("neighborhood","")
    q = Residence.query.filter_by(city=city)
    if hood:
        q = q.filter_by(neighborhood=hood)
    return jsonify([{"id":r.id,"name":r.name} for r in q.all()])

@app.route("/api/public/buildings-by-residence")
def public_buildings_by_residence():
    rid = request.args.get("residence_id","")
    buildings = Building.query.filter_by(residence_id=rid).all()
    return jsonify([{"id":b.id,"name":b.name,"floors":b.total_floors} for b in buildings])

@app.route("/api/public/floors-by-building")
def public_floors_by_building():
    bid = request.args.get("building_id","")
    from sqlalchemy import distinct
    floors = db.session.query(distinct(Apartment.floor)).filter(
        Apartment.building_id == bid
    ).order_by(Apartment.floor).all()
    return jsonify([f[0] for f in floors if f[0] is not None])

@app.route("/api/public/apartments-by-floor")
def public_apartments_by_floor():
    bid = request.args.get("building_id","")
    floor = request.args.get("floor","")
    if not bid:
        return jsonify([])
    if floor:
        apts = Apartment.query.filter_by(building_id=int(bid), floor=int(floor)).all()
    else:
        apts = Apartment.query.filter_by(building_id=int(bid)).all()
    return jsonify([{"id":a.id,"number":a.number} for a in apts])


# ══════════════════════════════════════════════════════════
#  PLATFORM SETTINGS — إعدادات منصة المدير 💳
# ══════════════════════════════════════════════════════════

# قائمة الإعدادات الافتراضية
PLATFORM_DEFAULTS = [
    # ── معلومات الأداء ──
    {'key': 'bank_name',        'label': '🏦 اسم البنك',               'category': 'payment',  'value': 'CIH Bank'},
    {'key': 'bank_rib',         'label': '🔢 رقم الحساب (RIB)',         'category': 'payment',  'value': ''},
    {'key': 'bank_owner',       'label': '👤 اسم صاحب الحساب',          'category': 'payment',  'value': 'SyndikPro SARL'},
    {'key': 'paypal_email',     'label': '🌐 حساب PayPal',              'category': 'payment',  'value': 'payments@syndikpro.ma'},
    {'key': 'paypal_name',      'label': '👤 اسم حساب PayPal',          'category': 'payment',  'value': 'SyndikPro'},
    {'key': 'cmi_merchant',     'label': '💳 رقم التاجر CMI',           'category': 'payment',  'value': ''},
    {'key': 'payment_ref_note', 'label': '📝 ملاحظة المرجع',            'category': 'payment',  'value': 'ضع اسم المستخدم كمرجع للتحويل'},
    # ── معلومات الدعم ──
    {'key': 'support_phone',    'label': '📞 هاتف خدمة العملاء',        'category': 'support',  'value': ''},
    {'key': 'support_whatsapp', 'label': '💬 واتساب خدمة العملاء',      'category': 'support',  'value': ''},
    {'key': 'support_email',    'label': '📧 بريد خدمة العملاء',        'category': 'support',  'value': 'support@syndikpro.ma'},
    {'key': 'support_hours',    'label': '🕐 أوقات الدعم',              'category': 'support',  'value': 'الاثنين-الجمعة 9:00-18:00'},
    # ── معلومات عامة ──
    {'key': 'platform_name',    'label': '🏷️ اسم المنصة',              'category': 'general',  'value': 'SyndikPro'},
    {'key': 'platform_website', 'label': '🌍 الموقع الإلكتروني',        'category': 'general',  'value': 'https://syndikpro.ma'},
    {'key': 'platform_slogan',  'label': '💬 الشعار',                   'category': 'general',  'value': 'نظام تدبير الإقامات السكنية الذكي'},
]

def init_platform_settings():
    """تهيئة الإعدادات الافتراضية عند أول تشغيل"""
    try:
        from models import PlatformSettings
        for item in PLATFORM_DEFAULTS:
            if not PlatformSettings.query.filter_by(key=item['key']).first():
                s = PlatformSettings(**item)
                db.session.add(s)
        db.session.commit()
    except Exception as e:
        print(f"init_platform_settings error: {e}")

def get_all_settings():
    """جلب كل الإعدادات كـ dict"""
    try:
        from models import PlatformSettings
        settings = PlatformSettings.query.all()
        return {s.key: s.value for s in settings}
    except:
        return {}

# ── GET: جلب جميع الإعدادات ──────────────────────────
@app.route('/api/admin/platform-settings', methods=['GET'])
@admin_required
def admin_get_platform_settings():
    from models import PlatformSettings
    init_platform_settings()
    settings = PlatformSettings.query.order_by(PlatformSettings.category, PlatformSettings.id).all()
    # تجميع حسب الفئة
    result = {}
    for s in settings:
        cat = s.category
        if cat not in result:
            result[cat] = []
        result[cat].append(s.to_dict())
    return jsonify(result)

# ── PUT: تحديث إعداد واحد ────────────────────────────
@app.route('/api/admin/platform-settings/<key>', methods=['PUT'])
@admin_required
def admin_update_setting(key):
    from models import PlatformSettings
    data = request.get_json() or {}
    s = PlatformSettings.query.filter_by(key=key).first()
    if not s:
        s = PlatformSettings(key=key, label=data.get('label', key),
                             category=data.get('category', 'general'))
        db.session.add(s)
    s.value = data.get('value', '')
    s.updated_at = datetime.utcnow()
    s.updated_by = session.get('user_id')
    db.session.commit()
    return jsonify({'ok': True, 'key': key, 'value': s.value})

# ── PUT: تحديث دفعي لعدة إعدادات ─────────────────────
@app.route('/api/admin/platform-settings', methods=['PUT'])
@admin_required
def admin_update_all_settings():
    from models import PlatformSettings
    data = request.get_json() or {}
    updated = 0
    for key, value in data.items():
        s = PlatformSettings.query.filter_by(key=key).first()
        if s:
            s.value = value
            s.updated_at = datetime.utcnow()
            s.updated_by = session.get('user_id')
            updated += 1
    db.session.commit()
    return jsonify({'ok': True, 'updated': updated})

# ── GET PUBLIC: للسانديك والقاطنين ───────────────────
@app.route('/api/public/platform-info')
def public_platform_info():
    """معلومات الأداء العامة — مرئية للسانديك عند الدفع"""
    try:
        from models import PlatformSettings
        settings = PlatformSettings.query.filter(
            PlatformSettings.category.in_(['payment', 'support'])
        ).all()
        return jsonify({s.key: s.value for s in settings})
    except:
        return jsonify({})


@app.route('/api/resident/pay-multi', methods=['POST'])
@resident_required
def resident_pay_multi():
    r = get_current_resident()
    if not r or not r.apartment_id:
        return jsonify({'error': 'لم يتم تعيين شقة لك بعد'}), 400

    method = request.form.get('method', '')
    months_data = request.form.get('months_data', '[]')
    total_amount = float(request.form.get('total_amount', 0))
    months_labels = request.form.get('months_labels', '')

    import json
    months = json.loads(months_data)
    if not months:
        return jsonify({'error': 'اختر شهراً على الأقل'}), 400

    proof_path = ''
    if 'proof' in request.files:
        f = request.files['proof']
        if f and f.filename and allowed_file(f.filename):
            from werkzeug.utils import secure_filename
            import time
            fn = f"{int(time.time())}_{secure_filename(f.filename)}"
            fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
            f.save(fp)
            proof_path = f'static/uploads/{fn}'

    # إنشاء طلب واحد لكل شهر
    created = []
    for m in months:
        existing = OnlinePaymentRequest.query.filter_by(
            apartment_id=r.apartment_id, month=m['month'], year=m['year'], status='pending'
        ).first()
        if existing:
            continue
        req = OnlinePaymentRequest(
            resident_id=r.id,
            apartment_id=r.apartment_id,
            month=int(m['month']), year=int(m['year']),
            amount=float(m['amount']),
            method=method,
            tx_ref=request.form.get('tx_ref',''),
            proof_path=proof_path,
            status='pending'
        )
        db.session.add(req)
        created.append(m)

    # إشعار واحد للقاطن
    notif = Notification(
        resident_id=r.id,
        title='✅ تم إرسال طلب الأداء',
        body=f'طلب أداء {total_amount} درهم ({months_labels}) قيد المراجعة.',
        type='info'
    )
    db.session.add(notif)

    # إشعار السانديك — كان مفقوداً لأشهر 2024/2025
    if created:
        apt = Apartment.query.get(r.apartment_id)
        res = Residence.query.get(r.residence_id)
        syndic_notif = GlobalNotification(
            title=f'💰 طلب أداء من {r.first_name} {r.last_name}',
            body=f'الإقامة: {res.name if res else "?"} | شقة: {apt.number if apt else "?"} | المبلغ: {total_amount} درهم | الأشهر: {months_labels}',
            type='payment',
            target='all'
        )
        db.session.add(syndic_notif)

    db.session.commit()

    return jsonify({'ok': True, 'created': len(created)})


# ══════════════════════════════════════════
#  NEIGHBOR POSTS — إعلانات الجيران
# ══════════════════════════════════════════
@app.route('/api/resident/neighbor-posts', methods=['GET'])
@resident_required
def get_neighbor_posts():
    r = get_current_resident()
    type_filter = request.args.get('type', '')
    from sqlalchemy import or_
    if r is None: return jsonify([])
    q = NeighborPost.query.filter(
        NeighborPost.is_active==True,
        or_(NeighborPost.residence_id==r.residence_id, NeighborPost.scope=='public')
    )
    if type_filter:
        q = q.filter(NeighborPost.type==type_filter)
    posts = q.order_by(NeighborPost.created_at.desc()).limit(50).all()
    return jsonify([p.to_dict(me_id=r.id) for p in posts])

@app.route('/api/resident/neighbor-posts', methods=['POST'])
@resident_required
def add_neighbor_post():
    r = get_current_resident()
    data = request.get_json() or {}
    if not data.get('title','').strip():
        return jsonify({'error': 'العنوان مطلوب'}), 400
    post = NeighborPost(
        residence_id=r.residence_id,
        resident_id=r.id,
        type=data.get('type','sell'),
        title=data['title'].strip(),
        description=data.get('description','').strip(),
        phone=data.get('phone','').strip(),
        city=data.get('city','').strip(),
        scope=data.get('scope','private'),
        contact_name=data.get('contact_name','').strip(),
        images=data.get('images',''),
        author_name=data.get('contact_name','').strip(),
        is_active=True,
    )
    db.session.add(post)
    db.session.commit()
    return jsonify({'ok': True, 'id': post.id})

@app.route('/api/resident/neighbor-posts/<int:pid>', methods=['DELETE'])
@resident_required
def delete_neighbor_post(pid):
    r = get_current_resident()
    post = NeighborPost.query.get_or_404(pid)
    if post.resident_id != r.id:
        return jsonify({'error': 'غير مصرح'}), 403
    post.is_active = False
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/resident/neighbor-posts/<int:pid>/sold', methods=['POST'])
@resident_required
def mark_neighbor_sold(pid):
    r = get_current_resident()
    post = NeighborPost.query.get_or_404(pid)
    if post.resident_id != r.id:
        return jsonify({'error': 'غير مصرح'}), 403
    data = request.get_json() or {}
    post.sold = data.get('sold', True)
    db.session.commit()
    return jsonify({'ok': True})

# ============ CHAT ROUTES ============
@app.route('/api/resident/chat/messages')
# ============= NEIGHBOR POST CHAT (private, anonymous) =============
@app.route('/api/resident/neighbor-posts/<int:pid>/chat')
@resident_required
def get_neighbor_post_chat(pid):
    r = get_current_resident()
    post = NeighborPost.query.get_or_404(pid)
    owner_id = post.resident_id
    if r.id == owner_id:
        other_id = request.args.get('with', type=int)
        if not other_id:
            return jsonify([])
    else:
        other_id = owner_id

    msgs = NeighborPostMessage.query.filter(
        NeighborPostMessage.post_id == pid,
        db.or_(
            db.and_(NeighborPostMessage.sender_resident_id==r.id, NeighborPostMessage.receiver_resident_id==other_id),
            db.and_(NeighborPostMessage.sender_resident_id==other_id, NeighborPostMessage.receiver_resident_id==r.id)
        )
    ).order_by(NeighborPostMessage.created_at.asc()).limit(200).all()

    for m in msgs:
        if m.receiver_resident_id == r.id and not m.is_read:
            m.is_read = True
    db.session.commit()

    return jsonify([m.to_dict(me_id=r.id) for m in msgs])


@app.route('/api/resident/neighbor-posts/<int:pid>/chat', methods=['POST'])
@resident_required
def send_neighbor_post_chat(pid):
    r = get_current_resident()
    post = NeighborPost.query.get_or_404(pid)
    data2 = request.get_json() or {}
    msg = (data2.get('message') or '').strip()
    if not msg or len(msg) > 1000:
        return jsonify({'error': 'invalid message'}), 400

    owner_id = post.resident_id
    if r.id == owner_id:
        other_id = data2.get('with')
        if not other_id:
            return jsonify({'error': 'recipient required'}), 400
    else:
        other_id = owner_id

    if other_id == r.id:
        return jsonify({'error': 'cannot message yourself'}), 400

    nm = NeighborPostMessage(
        post_id=pid,
        sender_resident_id=r.id,
        receiver_resident_id=other_id,
        message=msg
    )
    db.session.add(nm)
    db.session.commit()
    return jsonify({'ok': True, 'id': nm.id})


@app.route('/api/resident/neighbor-chats')
@resident_required
def get_neighbor_chats():
    r = get_current_resident()
    msgs = NeighborPostMessage.query.filter(
        db.or_(
            NeighborPostMessage.sender_resident_id == r.id,
            NeighborPostMessage.receiver_resident_id == r.id
        )
    ).order_by(NeighborPostMessage.created_at.desc()).limit(200).all()

    threads = {}
    for m in msgs:
        other_id = m.receiver_resident_id if m.sender_resident_id == r.id else m.sender_resident_id
        key = (m.post_id, other_id)
        if key not in threads:
            post2 = NeighborPost.query.get(m.post_id)
            unread = NeighborPostMessage.query.filter_by(
                post_id=m.post_id, sender_resident_id=other_id,
                receiver_resident_id=r.id, is_read=False
            ).count()
            threads[key] = {
                'post_id': m.post_id,
                'post_title': post2.title if post2 else '',
                'with': other_id,
                'last_message': m.message,
                'last_at': m.created_at.strftime('%d/%m/%Y %H:%M'),
                'unread': unread,
            }
    return jsonify(list(threads.values()))
app.route('/api/resident/chat/messages')
def get_chat_messages():
    if 'resident_id' not in session:
        return jsonify({'error': 'unauthorized'}), 401
    import sqlite3
    from datetime import datetime
    resident = Apartment.query.filter_by(id=session['resident_id']).first()
    if not resident:
        return jsonify({'error': 'not found'}), 404
    conn = sqlite3.connect('/home/Hicham/syndikpro/databasesyndic.db')
    conn.row_factory = sqlite3.Row
    since = request.args.get('since', 0)
    rows = conn.execute('''SELECT * FROM chat_message 
        WHERE residence_id=? AND id>? 
        ORDER BY created_at ASC LIMIT 50''',
        (resident.residence_id, since)).fetchall()
    conn.close()
    return jsonify({'messages': [dict(r) for r in rows]})

@app.route('/api/resident/chat/send', methods=['POST'])
def send_chat_message():
    if 'resident_id' not in session:
        return jsonify({'error': 'unauthorized'}), 401
    import sqlite3
    data = request.get_json()
    msg = (data.get('message') or '').strip()
    if not msg or len(msg) > 500:
        return jsonify({'error': 'invalid'}), 400
    resident = Apartment.query.filter_by(id=session['resident_id']).first()
    if not resident:
        return jsonify({'error': 'not found'}), 404
    conn = sqlite3.connect('/home/Hicham/syndikpro/databasesyndic.db')
    cur = conn.execute('''INSERT INTO chat_message 
        (residence_id, resident_id, resident_name, apartment, message)
        VALUES (?,?,?,?,?)''',
        (resident.residence_id, resident.id, 
         resident.owner_name or 'ساكن', 
         resident.number or '?', msg))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({'ok': True, 'id': new_id})

@app.route('/api/public/apartments-by-building')
def apartments_by_building():
    building_id = request.args.get('building_id')
    if not building_id:
        return jsonify([])
    from models import Apartment
    apts = Apartment.query.filter_by(building_id=int(building_id)).order_by(Apartment.number).all()
    return jsonify([{'id': a.id, 'number': a.number, 'floor': a.floor} for a in apts])


# ============================================================
# Excel Export / Import Routes for Syndic
# ============================================================
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file
import io

@app.route('/api/syndic/residents/export-excel')
def syndic_export_residents_excel():
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 403
    syndic_id = session['user_id']
    from models import Residence, Apartment, User
    syndic_user = User.query.get(syndic_id)
    if not syndic_user or syndic_user.role != 'syndic':
        return jsonify({'error': 'غير مصرح'}), 403
    residence_id_filter = request.args.get('residence_id', type=int)
    residences = Residence.query.filter_by(user_id=syndic_id).all()
    if residence_id_filter:
        residences = [r for r in residences if r.id == residence_id_filter]
    res_ids = [r.id for r in residences]
    apartments = Apartment.query.filter(Apartment.residence_id.in_(res_ids)).all()
    apt_map = {a.id: a for a in apartments}
    res_map = {r.id: r for r in residences}
    wb = Workbook()
    ws = wb.active
    ws.title = "السكان"
    ws.sheet_view.rightToLeft = True
    headers = ['الإقامة', 'العمارة', 'رقم الشقة', 'اسم المالك', 'هاتف المالك', 'اسم القاطن', 'هاتف القاطن']
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row, apt in enumerate(apartments, 2):
        res = res_map.get(apt.residence_id) if apt else None
        from models import Building
        building = Building.query.get(apt.building_id) if apt.building_id else None
        ws.cell(row=row, column=1, value=res.name if res else '')
        ws.cell(row=row, column=2, value=building.name if building else '')
        ws.cell(row=row, column=3, value=apt.number if apt else '')
        ws.cell(row=row, column=4, value=apt.owner_name or '')
        ws.cell(row=row, column=5, value=apt.owner_phone or '')
        ws.cell(row=row, column=6, value=apt.tenant_name or '')
        ws.cell(row=row, column=7, value=apt.tenant_phone or '')
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    from openpyxl.styles import Protection
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.protection = Protection(locked=True)
    for row in range(2, ws.max_row + 1):
        for col in [4, 5, 6, 7]:
            ws.cell(row=row, column=col).protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.password = "syndikpro2026"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='residents.xlsx')

@app.route('/api/syndic/residents/import-excel', methods=['POST'])
def syndic_import_residents_excel():
    file = request.files.get('file')
    residence_id = request.form.get('residence_id', type=int)
    if not file:
        return jsonify({'error': 'لم يتم رفع ملف'}), 400
    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        imported = 0
        skipped  = 0
        errors   = []

        all_res = Residence.query.all()
        if residence_id:
            all_res = [r for r in all_res if r.id == residence_id]
        res_map = {r.name.strip().lower(): r for r in all_res}
        apt_map = {(a.residence_id, str(int(a.number)) if isinstance(a.number, float) else str(a.number).strip()): a
                   for a in Apartment.query.all()}

        def handle_person(full_name, phone, role, apt, res):
            nonlocal imported, skipped
            if not full_name or not full_name.strip():
                return
            full_name = full_name.strip()
            phone = str(phone or '').strip()
            if role == 'owner':
                apt.owner_name = full_name
                apt.owner_phone = phone
            else:
                apt.tenant_name = full_name
                apt.tenant_phone = phone
            imported += 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            res_name  = str(row[0] or '').strip().lower()
            apt_num   = str(int(float(str(row[2])))) if row[2] is not None and str(row[2]).strip() != '' else ''
            owner_name  = str(row[3] or '').strip()
            owner_phone = str(row[4] or '').strip()
            tenant_name  = str(row[5] or '').strip()
            tenant_phone = str(row[6] or '').strip()

            if not res_name or not apt_num:
                errors.append(f"سطر ناقص: {row}")
                skipped += 1
                continue

            res = res_map.get(res_name)
            if not res:
                errors.append(f"مبنى غير موجود: {row[0]}")
                skipped += 1
                continue

            apt = apt_map.get((res.id, apt_num))
            if not apt:
                errors.append(f"شقة غير موجودة: {apt_num} في {row[0]}")
                skipped += 1
                continue

            # المالك
            handle_person(owner_name, owner_phone, 'owner', apt, res)

            # القاطن — فقط إذا مختلف عن المالك
            if tenant_name and tenant_name.lower() != owner_name.lower():
                handle_person(tenant_name, tenant_phone, 'tenant', apt, res)

        db.session.commit()
        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'message': f'✅ تم استيراد {imported} ساكن — تجاوز {skipped} شق فارغة',
            'errors': errors,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'خطأ في معالجة الملف: {str(e)}'}), 500


# ===== ملاحق المرسوم 2.23.700 =====
from annexes import build_pdf, get_building_category

@app.route('/api/syndic/annexes/info')
def annexes_info():
    u = get_current_user()
    if not u:
        return jsonify({'error': 'غير مصرح'}), 401
    year = request.args.get('year', 2024, type=int)
    rid = request.args.get('rid', type=int)
    if not rid:
        res = Residence.query.filter_by(user_id=u.id).first()
        if not res:
            return jsonify({'error': 'لا توجد إقامة'}), 404
        rid = res.id
    category, required, total = get_building_category(rid, year, db)
    return jsonify({'category': category, 'annexes': required, 'total_charges': total})

@app.route('/api/syndic/annexes/download/<annexe_num>')
def download_annexe(annexe_num):
    u = get_current_user()
    if not u:
        return jsonify({'error': 'غير مصرح'}), 401
    year = request.args.get('year', 2024, type=int)
    rid = request.args.get('rid', type=int)
    if not rid:
        res = Residence.query.filter_by(user_id=u.id).first()
        if not res:
            return jsonify({'error': 'لا توجد إقامة'}), 404
        rid = res.id
    buf = build_pdf(rid, year, annexe_num, db)
    return send_file(buf, mimetype='application/pdf',
        as_attachment=True, download_name=f'Annexe_{annexe_num}_{year}.pdf')

@app.route('/api/syndic/annexes/download-all')
def download_all_annexes():
    import zipfile, io as _io
    u = get_current_user()
    if not u:
        return jsonify({'error': 'غير مصرح'}), 401
    year = request.args.get('year', 2024, type=int)
    rid = request.args.get('rid', type=int)
    if not rid:
        res = Residence.query.filter_by(user_id=u.id).first()
        if not res:
            return jsonify({'error': 'لا توجد إقامة'}), 404
        rid = res.id
    category, required, total = get_building_category(rid, year, db)
    zip_buf = _io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w') as zf:
        for ann in required:
            pdf_buf = build_pdf(rid, year, ann, db)
            zf.writestr(f'Annexe_{ann}_{year}.pdf', pdf_buf.read())
    zip_buf.seek(0)
    return send_file(zip_buf, mimetype='application/zip',
        as_attachment=True, download_name=f'Annexes_2.23.700_{year}.zip')


# ============================================================
# ACCOUNTING - المحاسبة
# ============================================================
@app.route('/api/accounting')
def api_accounting():
    if 'user_id' not in session:
        return jsonify({'error': 'unauthorized'}), 401
    year  = request.args.get('year',  type=int)
    month = request.args.get('month', type=int)
    syndic_id = session['user_id']
    residences = Residence.query.filter_by(user_id=syndic_id).all()
    res_ids = [r.id for r in residences]
    apts = Apartment.query.filter(Apartment.residence_id.in_(res_ids)).all()
    apt_ids = [a.id for a in apts]

    # الإيرادات
    pq = Payment.query.filter(Payment.apartment_id.in_(apt_ids), Payment.status=='paid')
    if year:  pq = pq.filter(Payment.year==year)
    if month: pq = pq.filter(Payment.month==month)
    payments = pq.all()

    # المصاريف
    eq = Expense.query.filter(Expense.residence_id.in_(res_ids))
    if year:  eq = eq.filter(db.func.strftime('%Y', Expense.date)==str(year))
    if month: eq = eq.filter(db.func.strftime('%m', Expense.date)==f'{month:02d}')
    expenses = eq.all()

    total_income  = sum(p.amount for p in payments)
    total_expense = sum(e.amount for e in expenses)

    monthly = {}
    for p in payments:
        k = f"{p.year}-{p.month:02d}"
        monthly.setdefault(k, {"income": 0, "expense": 0})
        monthly[k]["income"] += p.amount
    for e in expenses:
        k = e.date.strftime("%Y-%m")
        monthly.setdefault(k, {"income": 0, "expense": 0})
        monthly[k]["expense"] += e.amount

    return jsonify({
        "total_income":  total_income,
        "total_expense": total_expense,
        "net":           total_income - total_expense,
        "payments":      [p.to_dict() for p in payments],
        "expenses":      [e.to_dict() for e in expenses],
        "monthly":       monthly
    })

@app.route('/accounting')
def accounting_page():
    if 'user_id' not in session:
        return jsonify({'error': 'unauthorized'}), 401
    return render_template('accounting.html')




@app.route('/api/daily-fact')
def get_daily_fact():
    import random
    facts = DailyFact.query.all()
    if not facts:
        return jsonify({'fact_text': 'لا توجد معلومات'})
    return jsonify(random.choice([f.to_dict() for f in facts]))

@app.route('/api/admin/daily-facts')
@admin_required
def admin_facts():
    return jsonify([f.to_dict() for f in DailyFact.query.all()])

@app.route('/api/admin/daily-facts', methods=['POST'])
@admin_required
def admin_add_fact():
    data = request.get_json()
    f = DailyFact(fact_text=data.get('fact_text',''))
    db.session.add(f)
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/admin/daily-facts/<int:fid>', methods=['DELETE'])
@admin_required
def admin_del_fact(fid):
    f = DailyFact.query.get(fid)
    if f:
        db.session.delete(f)
        db.session.commit()
    return jsonify({'ok': True})

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ADMIN: SYNDIC APPROVAL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@app.route('/api/admin/pending-syndics')
@admin_required
def admin_pending_syndics():
    pending = User.query.filter_by(status='pending', role='syndic').all()
    return jsonify([u.to_dict() for u in pending])

@app.route('/api/admin/syndics/<int:uid>/approve', methods=['POST'])
@admin_required
def admin_approve_reject_syndic(uid):
    u = User.query.get(uid)
    if not u or u.role != 'syndic':
        return jsonify({'error': 'غير موجود'}), 404
    
    data = request.get_json() or {}
    action = data.get('action', 'approve')
    
    if action == 'approve':
        u.status = 'active'
        u.subscription_confirmed = True
        u.renewal_pending = False
    u.approved_by = session.get('user_id')
    u.approved_at = datetime.utcnow()
    u.subscription_start = datetime.utcnow()
    u.subscription_end = datetime.utcnow() + timedelta(days=30)
    db.session.commit()
    
    # إشعار للسانديك
    notif = Notification(
        resident_id=u.id,
        title='تم قبول طلبك',
        body='تم الموافقة على حسابك! يمكنك الآن استخدام كل الميزات',
        type='info'
    )
    db.session.add(notif)
    db.session.commit()
    
    return jsonify({'ok': True, 'message': 'تمت الموافقة'})

@app.route('/api/admin/syndics/<int:uid>/reject', methods=['POST'])
@admin_required
def admin_reject_syndic(uid):
    u = User.query.get(uid)
    if not u:
        return jsonify({'error': 'غير موجود'}), 404
    
    u.status = 'rejected'
    u.subscription_confirmed = False
    u.renewal_pending = False
    u.approved_by = session.get('user_id')
    u.approved_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({'ok': True})

# ══════════════════════════════════════════
# SUBSCRIPTION REQUESTS - طلبات الاشتراك
# ══════════════════════════════════════════
@app.route('/api/subscriptions/request', methods=['POST'])
def create_subscription_request():
    if 'user_id' not in session:
        return jsonify({'error': 'غير مصرح'}), 401
    
    user_id = session.get('user_id')
    try:
            from models import SubscriptionRequest, SubscriptionDiscount, User
            from werkzeug.utils import secure_filename
            import os
    
            plan_name = request.form.get('plan_id') or request.form.get('plan')
            months = int(request.form.get('months', 1))
            payment_method = request.form.get('payment_method')
            notes = request.form.get('notes', '')
            receipt_file = request.files.get('receipt')
    
            if not plan_name:
                return jsonify({'error': 'خطة غير محددة'}), 400
    
    except Exception as e:
        print(f"Subscription error: {e}")
        return jsonify({"error": str(e)}), 500
    if not payment_method:
        return jsonify({'error': 'طريقة الدفع مطلوبة'}), 400
    
    if not receipt_file:
        return jsonify({'error': 'الوصل مطلوب'}), 400
    
    allowed_extensions = {'jpg', 'jpeg', 'png', 'gif', 'pdf'}
    if '.' not in receipt_file.filename or receipt_file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        return jsonify({'error': 'نوع الملف غير مدعوم'}), 400
    
    plan = SubscriptionPlan.query.filter_by(name=plan_name).first()
    if not plan:
        return jsonify({'error': 'خطة غير موجودة'}), 404
    
    discount = SubscriptionDiscount.query.filter(
        SubscriptionDiscount.min_months <= months,
        SubscriptionDiscount.max_months >= months
    ).first()
    
    base_price = plan.price_monthly * months
    discount_percent = discount.discount_percent if discount else 0
    discount_amount = base_price * (discount_percent / 100)
    final_price = base_price - discount_amount
    
    receipt_filename = secure_filename(f"receipt_{current_user.id}_{int(__import__('time').time())}.{receipt_file.filename.rsplit('.', 1)[1].lower()}")
    receipts_dir = os.path.join(os.path.dirname(__file__), 'uploads', 'receipts')
    os.makedirs(receipts_dir, exist_ok=True)
    receipt_path = os.path.join(receipts_dir, receipt_filename)
    receipt_file.save(receipt_path)
    
    req = SubscriptionRequest(
        user_id=user_id,
        plan_id=plan_id,
        months=months,
        base_price=base_price,
        discount_percent=discount_percent,
        discount_amount=discount_amount,
        final_price=final_price,
        payment_method=payment_method,
        receipt_path=receipt_path,
        notes=notes,
        status='pending'
    )
    db.session.add(req)
    db.session.commit()
    
    # إشعار واحد للأدمين
    try:
        u_req = User.query.get(session['user_id'])
        gn = GlobalNotification(
            title=f'📋 طلب اشتراك جديد — {u_req.full_name if u_req else "سانديك"}',
            body=f'الخطة: {plan_obj.label} | {months} شهر | {final_price:.2f} درهم | الطريقة: {payment_method}',
            type='subscription',
            target='admin'
        )
        db.session.add(gn)
        db.session.commit()
    except Exception as e:
        print(f'Admin notification error: {e}')

    return jsonify({'ok': True, 'request_id': req.id, 'final_price': final_price})

@app.route('/api/admin/subscription-requests', methods=['GET'])
@admin_required
def get_subscription_requests():
    from models import SubscriptionRequest, SubscriptionPlan
    reqs = SubscriptionRequest.query.filter_by(status='pending').order_by(SubscriptionRequest.requested_at.desc()).all()
    result = []
    for r in reqs:
        u = User.query.get(r.user_id)
        p = SubscriptionPlan.query.get(r.plan_id)
        result.append({
            'id': r.id,
            'user_id': r.user_id,
            'syndic_name': u.full_name if u else '—',
            'syndic_email': u.email if u else '—',
            'syndic_phone': u.phone if u else '—',
            'syndic_city': u.city if u else '—',
            'plan_name': p.name if p else '—',
            'plan_label': p.label if p else '—',
            'months': r.months or 1,
            'base_price': r.base_price or 0,
            'discount_percent': r.discount_percent or 0,
            'final_price': r.final_price or 0,
            'requested_at': r.requested_at.strftime('%d/%m/%Y %H:%M') if r.requested_at else '—',
        })
    return jsonify(result)
@app.route('/api/admin/subscription-requests/<int:req_id>/approve', methods=['POST'])
@admin_required
def approve_subscription_request(req_id):
    from models import SubscriptionRequest
    req = SubscriptionRequest.query.get_or_404(req_id)
    req.status = 'approved'
    req.approved_by = session.get('user_id')
    req.approved_at = datetime.utcnow()
    # تفعيل الاشتراك على المستخدم
    u = User.query.get(req.user_id)
    if u:
        plan_obj = SubscriptionPlan.query.get(req.plan_id)
        u.plan = plan_obj.name if plan_obj else 'basic'
        u.subscription_confirmed = True
        u.subscription_start = datetime.utcnow()
        u.subscription_end = datetime.utcnow() + timedelta(days=30 * (req.months or 1))
        u.status = 'active'
        db.session.commit()
        # إيميل للسانديك
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            plan_label = {'free':'مجانية','basic':'أساسية','pro':'احترافية','enterprise':'غير محدود'}.get(u.plan, u.plan)
            end_date = u.subscription_end.strftime('%d/%m/%Y')
            msg = MIMEMultipart('alternative')
            msg['Subject'] = 'SyndikPro — تم تفعيل اشتراكك'
            msg['From'] = GMAIL_USER
            msg['To'] = u.email
            body = f'''<div dir="rtl" style="font-family:Arial,sans-serif;max-width:520px;margin:auto;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#1e40af,#3b82f6);padding:28px;text-align:center">
    <div style="font-size:40px">🏢</div>
    <h1 style="color:#fff;margin:6px 0;font-size:20px">SyndikPro</h1>
  </div>
  <div style="padding:24px">
    <p style="font-size:15px;color:#0f172a">السلام عليكم، <b>{u.full_name}</b></p>
    <div style="background:#d1fae5;border:1px solid #6ee7b7;border-radius:10px;padding:14px;text-align:center;margin:16px 0">
      <div style="font-size:24px">✅</div>
      <b style="color:#065f46;font-size:15px">تم تفعيل اشتراكك بنجاح!</b>
    </div>
    <table style="width:100%;font-size:13px;border-collapse:collapse;margin-bottom:16px">
      <tr style="background:#f8fafc"><td style="padding:8px;color:#64748b">الخطة</td><td style="padding:8px;font-weight:700">{plan_label}</td></tr>
      <tr><td style="padding:8px;color:#64748b">المدة</td><td style="padding:8px;font-weight:700">{req.months or 1} شهر</td></tr>
      <tr style="background:#f8fafc"><td style="padding:8px;color:#64748b">تاريخ الانتهاء</td><td style="padding:8px;font-weight:700;color:#dc2626">{end_date}</td></tr>
    </table>
    <div style="text-align:center">
      <a href="https://hicham.pythonanywhere.com" style="background:linear-gradient(135deg,#1e40af,#3b82f6);color:#fff;padding:12px 32px;border-radius:10px;text-decoration:none;font-weight:700;font-size:14px">الدخول إلى التطبيق ←</a>
    </div>
  </div>
  <div style="background:#f1f5f9;padding:12px;text-align:center;font-size:11px;color:#94a3b8">SyndikPro — نظام تدبير الإقامات السكنية</div>
</div>'''
            msg.attach(MIMEText(body, 'html', 'utf-8'))
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(GMAIL_USER, GMAIL_PASS)
                smtp.sendmail(GMAIL_USER, u.email, msg.as_string())
        except Exception as e:
            print(f'email error approve: {e}')
    return jsonify({'ok': True})

@app.route('/api/admin/subscription-requests/<int:req_id>/reject', methods=['POST'])
@admin_required
def reject_subscription_request(req_id):
    from models import SubscriptionRequest
    req = SubscriptionRequest.query.get_or_404(req_id)
    data = request.get_json() or {}
    reason = data.get('reason', 'لم يتم تحديد السبب')
    req.status = 'rejected'
    req.approved_by = current_user.id
    req.approved_at = datetime.utcnow()
    db.session.commit()
    try:
        syndic = User.query.get(req.user_id)
        if syndic:
            gn = GlobalNotification(
                title='❌ تم رفض طلب اشتراكك',
                body=f'للأسف تم رفض طلبك. السبب: {reason}',
                type='subscription',
                target='user',
                user_id=req.user_id
            )
            db.session.add(gn)
            db.session.commit()
    except Exception as e:
        print(f'reject notif error: {e}')
    return jsonify({'ok': True})


# ===== City & Neighborhood APIs =====
@app.route("/api/cities", methods=['GET'])
def get_cities():
    from morocco_cities import MOROCCO_CITIES
    cities = sorted(list(MOROCCO_CITIES.keys()))
    return jsonify(cities)

@app.route("/api/neighborhoods", methods=['GET'])
def get_neighborhoods():
    from morocco_cities import MOROCCO_CITIES
    city = request.args.get('city', '')
    if city in MOROCCO_CITIES:
        neighborhoods = sorted(MOROCCO_CITIES[city])
        return jsonify(neighborhoods)
    return jsonify([])
